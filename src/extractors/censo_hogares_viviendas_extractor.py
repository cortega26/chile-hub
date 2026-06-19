"""Extrae viviendas y hogares censados por comuna desde Censo 2024."""

import datetime
import sys
from pathlib import Path

import openpyxl
import polars as pl
import requests

UTC = datetime.timezone.utc

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.extractors.base import BaseExtractor, ensure_staging_directories, write_staging_metadata
from src.validation import validate_censo_hogares_viviendas

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
RAW_DIR = DATA_DIR / "raw"
STAGING_DIR = DATA_DIR / "staging"
STAGING_CSV_PATH = STAGING_DIR / "censo_hogares_viviendas.csv"
METADATA_PATH = STAGING_DIR / "censo_hogares_viviendas.metadata.json"
SOURCE_URL = (
    "https://censo2024.ine.gob.cl/wp-content/uploads/2025/03/V1_Viviendas-y-hogares-censados.xlsx"
)
REUSE_POLICY = {
    "status": "open-attribution",
    "license": "CC BY 4.0",
    "license_url": "https://www.ine.gob.cl/terminos-de-uso",
    "attribution_required": True,
    "redistribution_ok": True,
    "summary": "Resultados oficiales del Censo 2024; atribucion requerida.",
}


class CensoHogaresViviendasExtractor(BaseExtractor):
    @property
    def dataset_name(self) -> str:
        return "censo_hogares_viviendas"

    def fetch(self, **kwargs):
        return fetch_workbook()

    def normalize(self, raw_data):
        path, _ = raw_data
        return parse_workbook(path)

    def validate(self, df, metadata: dict) -> dict:
        return validate_censo_hogares_viviendas(df, metadata)

    def write_staging(self, df, metadata: dict) -> Path:
        ensure_staging_directories()
        output = Path(STAGING_CSV_PATH)
        df.write_csv(output)
        full_metadata = {
            **metadata,
            "dataset": self.dataset_name,
            "record_count": df.height,
            "fields": df.columns,
            "reuse_policy": REUSE_POLICY,
        }
        write_staging_metadata(str(METADATA_PATH), full_metadata)
        return output


def fetch_workbook():
    ensure_staging_directories()
    target = (
        RAW_DIR
        / f"ine_censo2024_hogares_viviendas_{datetime.datetime.now(UTC):%Y%m%dT%H%M%SZ}.xlsx"
    )
    try:
        with requests.get(SOURCE_URL, timeout=60) as response:
            response.raise_for_status()
            target.write_bytes(response.content)
        return target, "live"
    except (requests.RequestException, OSError):
        snapshots = sorted(RAW_DIR.glob("ine_censo2024_hogares_viviendas_*.xlsx"))
        if not snapshots:
            raise
        return snapshots[-1], "fallback"


def parse_workbook(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = {}
    for row in workbook["2"].iter_rows(min_row=5, values_only=True):
        if not row[4] or int(row[4]) == 0:
            continue
        code = str(int(row[4])).zfill(5)
        records[code] = {
            "codigo_region": str(int(row[0])).zfill(2),
            "nombre_region": row[1],
            "codigo_provincia": str(int(row[2])).zfill(3),
            "nombre_provincia": row[3],
            "codigo_comuna": code,
            "nombre_comuna": row[5],
            "viviendas_censadas": int(row[6]),
            "viviendas_particulares_ocupadas": int(row[7]),
            "viviendas_particulares_desocupadas": int(row[8]),
            "viviendas_colectivas": int(row[9]),
        }
    for row in workbook["6"].iter_rows(min_row=5, values_only=True):
        if len(row) < 8:
            continue
        if not row[4] or int(row[4]) == 0:
            continue
        code = str(int(row[4])).zfill(5)
        records[code].update(
            {
                "hogares_censados": int(row[6]),
                "promedio_personas_hogar": None if row[7] in ("-", None) else float(row[7]),
            }
        )
    workbook.close()
    return pl.DataFrame(list(records.values())).sort("codigo_comuna")


def process():
    path, source_mode = fetch_workbook()
    df = parse_workbook(path)
    extractor = CensoHogaresViviendasExtractor()
    validation = extractor.validate(df, {"source_mode": source_mode})
    if validation["status"] == "error":
        raise SystemExit(f"Validacion fallida: {validation['errors']}")

    metadata = {
        "dataset": "censo_hogares_viviendas",
        "source_name": "Instituto Nacional de Estadisticas - Censo 2024",
        "source_url": SOURCE_URL,
        "source_mode": source_mode,
        "source_detail": "official_xlsx" if source_mode == "live" else "raw_snapshot_recovery",
        "refreshed_at_utc": datetime.datetime.now(UTC).isoformat(),
        "record_count": df.height,
        "fields": df.columns,
        "notes": [],
        "reuse_policy": REUSE_POLICY,
    }
    extractor.write_staging(df, metadata)
    print(f"Censo hogares y viviendas guardado: {df.height} registros ({source_mode})")


if __name__ == "__main__":
    process()
