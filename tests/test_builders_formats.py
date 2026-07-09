"""Tests golden round-trip para los writers de formatos (formats.py).

Cada test sigue el patron: escribir con el builder → reabrir el artefacto →
verificar columnas, dtypes y conteo de filas.
"""

import datetime
import os
import sqlite3
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

import duckdb
import polars as pl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from src.builders.formats import (
    build_duckdb,
    build_excel,
    build_flat_files,
    build_sqlite,
)

# ---------------------------------------------------------------------------
# Fixtures helpers
# ---------------------------------------------------------------------------


def _make_fixtures() -> dict[str, pl.DataFrame]:
    """Retorna los 7 DataFrames base con columnas canonicas y 2-3 filas cada uno."""
    return {
        "regiones": pl.DataFrame(
            {
                "codigo_region": ["13", "01", "02"],
                "nombre_region": [
                    "Region Metropolitana",
                    "Region de Tarapaca",
                    "Region de Antofagasta",
                ],
            },
            schema={"codigo_region": pl.String, "nombre_region": pl.String},
        ),
        "provincias": pl.DataFrame(
            {
                "codigo_provincia": ["131", "132", "011"],
                "codigo_region": ["13", "13", "01"],
                "nombre_provincia": ["Santiago", "Chacabuco", "Iquique"],
                "nombre_region": [
                    "Region Metropolitana",
                    "Region Metropolitana",
                    "Region de Tarapaca",
                ],
            },
            schema={
                "codigo_provincia": pl.String,
                "codigo_region": pl.String,
                "nombre_provincia": pl.String,
                "nombre_region": pl.String,
            },
        ),
        "comunas": pl.DataFrame(
            {
                "codigo_comuna": ["13101", "13102", "13103"],
                "codigo_provincia": ["131", "131", "131"],
                "codigo_region": ["13", "13", "13"],
                "nombre_comuna": ["Santiago", "Cerrillos", "Cerro Navia"],
                "nombre_comuna_clean": ["Santiago", "Cerrillos", "Cerro Navia"],
            },
            schema={
                "codigo_comuna": pl.String,
                "codigo_provincia": pl.String,
                "codigo_region": pl.String,
                "nombre_comuna": pl.String,
                "nombre_comuna_clean": pl.String,
            },
        ),
        "indicadores": pl.DataFrame(
            {
                "fecha": [
                    datetime.date(2026, 7, 1),
                    datetime.date(2026, 7, 1),
                    datetime.date(2026, 7, 2),
                ],
                "codigo_indicador": ["UF", "DOLAR", "UF"],
                "valor": [38000.0, 950.0, 38005.0],
            },
            schema={
                "fecha": pl.Date,
                "codigo_indicador": pl.String,
                "valor": pl.Float64,
            },
        ),
        "censo_comunal": pl.DataFrame(
            {
                "codigo_comuna": ["13101", "13102", "13103"],
                "codigo_provincia": ["131", "131", "131"],
                "codigo_region": ["13", "13", "13"],
                "nombre_comuna": ["Santiago", "Cerrillos", "Cerro Navia"],
                "poblacion_censada": [500_000, 80_000, 130_000],
                "hombres": [240_000, 39_000, 63_000],
                "mujeres": [260_000, 41_000, 67_000],
            },
            schema={
                "codigo_comuna": pl.String,
                "codigo_provincia": pl.String,
                "codigo_region": pl.String,
                "nombre_comuna": pl.String,
                "poblacion_censada": pl.Int64,
                "hombres": pl.Int64,
                "mujeres": pl.Int64,
            },
        ),
        "establecimientos_salud": pl.DataFrame(
            {
                "codigo_establecimiento": ["E001", "E002", "E003"],
                "nombre_establecimiento": [
                    "Hospital Salvador",
                    "CESFAM Lo Prado",
                    "Posta Central",
                ],
                "tipo_establecimiento": ["Hospital", "CESFAM", "Posta"],
                "codigo_region": ["13", "13", "13"],
                "codigo_comuna": ["13101", "13102", "13103"],
            },
            schema={
                "codigo_establecimiento": pl.String,
                "nombre_establecimiento": pl.String,
                "tipo_establecimiento": pl.String,
                "codigo_region": pl.String,
                "codigo_comuna": pl.String,
            },
        ),
        "establecimientos_educacionales": pl.DataFrame(
            {
                "rbd": ["10001", "10002", "10003"],
                "dv_rbd": ["0", "1", "k"],
                "nombre_establecimiento": [
                    "Liceo A",
                    "Colegio B",
                    "Escuela C",
                ],
                "codigo_region": ["13", "13", "13"],
                "codigo_comuna": ["13101", "13102", "13103"],
                "estado_funcionamiento": [
                    "Funcionando",
                    "Funcionando",
                    "Funcionando",
                ],
            },
            schema={
                "rbd": pl.String,
                "dv_rbd": pl.String,
                "nombre_establecimiento": pl.String,
                "codigo_region": pl.String,
                "codigo_comuna": pl.String,
                "estado_funcionamiento": pl.String,
            },
        ),
    }


def _make_extra_table() -> dict[str, pl.DataFrame]:
    """Retorna una tabla extra (censo_hogares_viviendas) para pruebas."""
    return {
        "censo_hogares_viviendas": pl.DataFrame(
            {
                "codigo_comuna": ["13101", "13102", "13103"],
                "codigo_region": ["13", "13", "13"],
                "viviendas_particulares": [200_000, 30_000, 50_000],
                "hogares": [190_000, 28_000, 48_000],
            },
            schema={
                "codigo_comuna": pl.String,
                "codigo_region": pl.String,
                "viviendas_particulares": pl.Int64,
                "hogares": pl.Int64,
            },
        ),
    }


def _make_oversized_table() -> pl.DataFrame:
    """Crea un DataFrame que supera el umbral de 500k filas."""
    return pl.DataFrame({"id": pl.arange(0, 500_001, eager=True)})


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestBuildFlatFiles:
    """build_flat_files: escribe .parquet y .json, sin .tmp residuales."""

    def test_parquet_round_trip(self):
        """Los Parquet reabiertos conservan columnas, dtypes y conteo de filas."""
        dfs = _make_fixtures()
        extra = _make_extra_table()

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized = os.path.join(tmpdir, "normalized")
            os.makedirs(normalized, exist_ok=True)

            with patch("src.builders.formats.NORMALIZED_DIR", normalized):
                build_flat_files(
                    dfs["regiones"],
                    dfs["provincias"],
                    dfs["comunas"],
                    dfs["indicadores"],
                    dfs["censo_comunal"],
                    dfs["establecimientos_salud"],
                    dfs["establecimientos_educacionales"],
                    extra_tables=extra,
                )

            # Reabrir cada Parquet y verificar
            for table_name, expected_df in {**dfs, **extra}.items():
                if table_name == "censo_hogares_viviendas":
                    parquet_name = "censo_hogares_viviendas.parquet"
                elif table_name == "establecimientos_salud":
                    parquet_name = "establecimientos_salud.parquet"
                elif table_name == "establecimientos_educacionales":
                    parquet_name = "establecimientos_educacionales.parquet"
                elif table_name == "censo_comunal":
                    parquet_name = "censo_comunal.parquet"
                elif table_name == "indicadores":
                    parquet_name = "indicadores.parquet"
                else:
                    parquet_name = f"{table_name}.parquet"

                parquet_path = os.path.join(normalized, parquet_name)
                assert os.path.exists(parquet_path), f"Falta {parquet_path}"

                df_read = pl.read_parquet(parquet_path)
                assert df_read.shape[0] == expected_df.shape[0], (
                    f"{table_name}: esperadas {expected_df.shape[0]} filas, "
                    f"leidas {df_read.shape[0]}"
                )
                # Verificar dtypes: codigos CUT son String
                for col in expected_df.columns:
                    if "codigo" in col.lower():
                        assert df_read[col].dtype == pl.String, (
                            f"{table_name}.{col}: dtype esperado String, "
                            f"obtenido {df_read[col].dtype}"
                        )

    def test_no_tmp_files_left_behind(self):
        """No quedan archivos .tmp tras build_flat_files."""
        dfs = _make_fixtures()

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized = os.path.join(tmpdir, "normalized")
            os.makedirs(normalized, exist_ok=True)

            with patch("src.builders.formats.NORMALIZED_DIR", normalized):
                build_flat_files(
                    dfs["regiones"],
                    dfs["provincias"],
                    dfs["comunas"],
                    dfs["indicadores"],
                    dfs["censo_comunal"],
                    dfs["establecimientos_salud"],
                    dfs["establecimientos_educacionales"],
                )

            tmp_files = [f for f in os.listdir(normalized) if f.endswith(".tmp")]
            assert len(tmp_files) == 0, f"Archivos .tmp residuales: {tmp_files}"

    def test_json_files_written(self):
        """build_flat_files tambien escribe los archivos JSON esperados."""
        dfs = _make_fixtures()

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized = os.path.join(tmpdir, "normalized")
            os.makedirs(normalized, exist_ok=True)

            with patch("src.builders.formats.NORMALIZED_DIR", normalized):
                build_flat_files(
                    dfs["regiones"],
                    dfs["provincias"],
                    dfs["comunas"],
                    dfs["indicadores"],
                    dfs["censo_comunal"],
                    dfs["establecimientos_salud"],
                    dfs["establecimientos_educacionales"],
                )

            expected_jsons = [
                "regiones.json",
                "provincias.json",
                "comunas.json",
                "indicadores_hoy.json",
            ]
            for jname in expected_jsons:
                jpath = os.path.join(normalized, jname)
                assert os.path.exists(jpath), f"Falta JSON: {jpath}"
                data = pl.read_json(jpath)
                assert data.shape[0] > 0, f"JSON vacio: {jpath}"


class TestBuildDuckDB:
    """build_duckdb: base de datos DuckDB con tablas y dtypes correctos."""

    def test_duckdb_tables_exist(self):
        """Las tablas base existen en DuckDB y tienen las filas esperadas."""
        dfs = _make_fixtures()
        extra = _make_extra_table()

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.duckdb")
            build_duckdb(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=extra,
                output_path=db_path,
            )

            assert os.path.exists(db_path)

            con = duckdb.connect(db_path)
            try:
                row = con.execute("SELECT COUNT(*) FROM comunas").fetchone()
                assert row[0] == 3, f"comunas: esperadas 3 filas, obtenidas {row[0]}"

                row = con.execute("SELECT COUNT(*) FROM regiones").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM indicadores").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM censo_comunal").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM establecimientos_salud").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM establecimientos_educacionales").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM censo_hogares_viviendas").fetchone()
                assert row[0] == 3

                row = con.execute("SELECT COUNT(*) FROM comunas_enriquecidas").fetchone()
                assert row[0] == 3
            finally:
                con.close()

    def test_duckdb_codigo_comuna_is_varchar(self):
        """codigo_comuna se conserva como VARCHAR en DuckDB."""
        dfs = _make_fixtures()

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.duckdb")
            build_duckdb(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables={},
                output_path=db_path,
            )

            con = duckdb.connect(db_path)
            try:
                dtype_info = con.execute(
                    "SELECT data_type FROM information_schema.columns "
                    "WHERE table_name = 'comunas' AND column_name = 'codigo_comuna'"
                ).fetchone()
                assert dtype_info is not None
                dtype = dtype_info[0].upper()
                assert "VARCHAR" in dtype or "TEXT" in dtype, f"codigo_comuna dtype = {dtype}"
            finally:
                con.close()

    def test_duckdb_extra_table_with_index(self):
        """Tabla extra con codigo_comuna recibe indice."""
        dfs = _make_fixtures()
        extra = _make_extra_table()

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.duckdb")
            build_duckdb(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=extra,
                output_path=db_path,
            )

            con = duckdb.connect(db_path)
            try:
                rows = con.execute(
                    "SELECT index_name FROM duckdb_indexes "
                    "WHERE table_name = 'censo_hogares_viviendas'"
                ).fetchall()
                index_names = [r[0] for r in rows]
                assert any("censo_hogares_viviendas" in n for n in index_names), (
                    f"No se encontro indice para censo_hogares_viviendas: {index_names}"
                )
            finally:
                con.close()


class TestBuildSQLite:
    """build_sqlite: base de datos SQLite con tablas y skip de tablas masivas."""

    def test_sqlite_tables_exist(self):
        """Las tablas base existen en SQLite y tienen las filas esperadas."""
        dfs = _make_fixtures()
        extra = _make_extra_table()

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.db")
            build_sqlite(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=extra,
                output_path=db_path,
            )

            assert os.path.exists(db_path)

            conn = sqlite3.connect(db_path)
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM comunas")
                assert cursor.fetchone()[0] == 3

                cursor.execute("SELECT COUNT(*) FROM regiones")
                assert cursor.fetchone()[0] == 3

                cursor.execute("SELECT COUNT(*) FROM indicadores")
                assert cursor.fetchone()[0] == 3

                cursor.execute("SELECT COUNT(*) FROM censo_comunal")
                assert cursor.fetchone()[0] == 3

                cursor.execute("SELECT COUNT(*) FROM establecimientos_salud")
                assert cursor.fetchone()[0] == 3

                cursor.execute("SELECT COUNT(*) FROM establecimientos_educacionales")
                assert cursor.fetchone()[0] == 3

                cursor.execute(
                    "SELECT name FROM sqlite_master "
                    "WHERE type='view' AND name='comunas_enriquecidas'"
                )
                assert cursor.fetchone() is not None
            finally:
                conn.close()

    def test_sqlite_oversized_table_skipped(self):
        """Tabla extra que supera 500k filas se omite en SQLite."""
        dfs = _make_fixtures()
        oversized = {"oversized_test": _make_oversized_table()}

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.db")
            build_sqlite(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=oversized,
                output_path=db_path,
            )

            conn = sqlite3.connect(db_path)
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='oversized_test'"
                )
                assert cursor.fetchone() is None, (
                    "La tabla oversized_test no debio crearse en SQLite"
                )
            finally:
                conn.close()


class TestBuildExcel:
    """build_excel: archivo XLSX con hojas esperadas y skip de tablas masivas."""

    def test_excel_sheet_names(self):
        """Las hojas base existen en el libro Excel."""
        dfs = _make_fixtures()
        extra = _make_extra_table()

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            build_excel(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=extra,
                output_path=xlsx_path,
            )

            assert os.path.exists(xlsx_path)

            import openpyxl

            wb = openpyxl.load_workbook(xlsx_path)
            try:
                sheet_names = wb.sheetnames
                expected_sheets = [
                    "Regiones",
                    "Provincias",
                    "Comunas y Regiones",
                    "Indicadores Diarios",
                    "Censo Comunal",
                    "Establecimientos Salud",
                    "Establecimientos Educacionales",
                ]
                for sname in expected_sheets:
                    assert sname in sheet_names, f"Hoja '{sname}' no encontrada en {sheet_names}"
            finally:
                wb.close()

    def test_excel_oversized_table_skipped(self):
        """Tabla extra que supera 500k filas NO tiene hoja en Excel."""
        dfs = _make_fixtures()
        oversized = {"censo_hogares_viviendas": _make_oversized_table()}

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            build_excel(
                dfs["regiones"],
                dfs["provincias"],
                dfs["comunas"],
                dfs["indicadores"],
                dfs["censo_comunal"],
                dfs["establecimientos_salud"],
                dfs["establecimientos_educacionales"],
                extra_tables=oversized,
                output_path=xlsx_path,
            )

            import openpyxl

            wb = openpyxl.load_workbook(xlsx_path)
            try:
                sheet_names = wb.sheetnames
                for sname in sheet_names:
                    assert "censo_hogares_viviendas" not in sname.lower(), (
                        f"Encontrada hoja inesperada para tabla masiva: {sname}"
                    )
            finally:
                wb.close()
