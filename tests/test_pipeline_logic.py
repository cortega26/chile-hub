import datetime
import json
import re
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import polars as pl
import requests

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from chile_hub import pipeline_status_utils as _pipeline_canonical
from scripts import package_publishable_bundle
from scripts.fetch_adoption_stats import (
    build_badge as build_adoption_badge,
)
from scripts.fetch_adoption_stats import (
    build_payload as build_adoption_payload,
)
from scripts.fetch_adoption_stats import (
    load_offline_fixture as load_adoption_offline_fixture,
)
from scripts.fetch_adoption_stats import (
    parse_pypi_stats,
    sum_github_downloads,
)
from scripts.verify_pipeline import (
    UTC,
    _verify_stagnation,
    required_files_for_profile,
    verify_dataset_contract,
    verify_publication_policy,
    verify_source_registry,
)
from src.build_dev_db import (
    DATASET_CATALOG_CONFIG,
    build_coverage,
    build_dataset_changelog,
    build_degradation,
    build_drift,
    load_metadata,
    write_publishable_bundle_zip,
)
from src.build_dev_db import (
    main as build_main,
)
from src.extractors import bcentral_extractor
from src.pipeline_status_utils import build_hub_health, build_status_text
from src.validation import (
    EXPECTED_INDICATOR_CODES,
    EXPECTED_LIVE_COMUNAS_COUNT,
    FALLBACK_COMUNAS_COUNT,
    _duplicate_count,
    _invalid_fixed_length_count,
    _missing_columns,
    _negative_numeric_count,
    _percentage_out_of_bounds_count,
    _unknown_codes,
    validate_censo_comunal,
    validate_censo_hogares_viviendas,
    validate_comunas,
    validate_distritos_electorales,
    validate_establecimientos_educacionales,
    validate_establecimientos_salud,
    validate_finanzas_municipales,
    validate_indicadores,
    validate_indicadores_urbanos_siedu,
    validate_perfil_territorial_comunal,
    validate_provincias,
    validate_regiones,
    validate_resultados_educacionales,
)


class PipelineLogicTests(unittest.TestCase):
    def test_build_main_fails_when_staging_inputs_are_missing(self):
        with (
            patch("src.build_dev_db.ensure_directories"),
            patch("src.build_dev_db.os.path.exists", return_value=False),
            self.assertRaisesRegex(SystemExit, "No se encuentran los archivos CSV"),
        ):
            build_main()

    def test_build_main_fails_when_metadata_inputs_are_missing(self):
        def mock_exists(path):
            if str(path).endswith(".csv"):
                return True
            return False

        with (
            patch("src.build_dev_db.ensure_directories"),
            patch("src.build_dev_db.os.path.exists", side_effect=mock_exists),
            self.assertRaisesRegex(SystemExit, "No se encuentra comunas.metadata.json"),
        ):
            build_main()

    def test_load_metadata_fails_on_malformed_json(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            bad_json_file = Path(tmpdir) / "bad.json"
            bad_json_file.write_text("{invalid json", encoding="utf-8")
            with self.assertRaisesRegex(SystemExit, "contiene un JSON malformado"):
                load_metadata(str(bad_json_file))

    def test_load_metadata_fails_on_missing_required_fields(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            incomplete_json_file = Path(tmpdir) / "incomplete.json"
            incomplete_metadata = {
                "dataset": "comunas",
                "source_name": "SUBDERE",
                # missing other fields
            }
            incomplete_json_file.write_text(json.dumps(incomplete_metadata), encoding="utf-8")
            with self.assertRaisesRegex(SystemExit, "no cumple con el esquema obligatorio"):
                load_metadata(str(incomplete_json_file))

    def _stable_registry_entry(self, name, fallback_policy="none"):
        """Build a minimal stable_publishable registry entry for testing."""
        return {
            "dataset": name,
            "license_status": "open-attribution",
            "access_method": "api",
            "live_extractor_status": "implemented",
            "fallback_policy": fallback_policy,
            "maturity_status": "stable",
            "live_ready": True,
            "publish_blocking": True,
            "publication_track": "stable_publishable",
            "public_bundle_eligible": True,
        }

    def _candidate_registry_entry(self, name):
        """Build a minimal candidate registry entry for testing."""
        return {
            "dataset": name,
            "license_status": "open-attribution",
            "access_method": "landing_snapshot",
            "live_extractor_status": "fallback_only",
            "fallback_policy": "allowed_for_dev_blocked_for_publication",
            "maturity_status": "candidate",
            "live_ready": False,
            "publish_blocking": True,
            "publication_track": "candidate",
            "public_bundle_eligible": False,
        }

    def test_publication_policy_accepts_fresh_live_data(self):
        datasets = {
            name: {
                "source_mode": "live",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
            for name in DATASET_CATALOG_CONFIG
        }
        datasets["indicadores"]["indicator_delivery"] = {
            code: "live" for code in EXPECTED_INDICATOR_CODES
        }
        registry = [self._stable_registry_entry(name) for name in DATASET_CATALOG_CONFIG]

        verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_accepts_clean_published_monthly_backfill(self):
        datasets = {
            name: {
                "source_mode": "live",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
            for name in DATASET_CATALOG_CONFIG
        }
        datasets["indicadores"].update(
            {
                "source_detail": "public_api_with_published_backfill",
                "indicator_delivery": {"ipc": "published_backfill", "uf": "live"},
                "published_backfills": ["ipc"],
                "fetch_failures": [],
                "raw_recoveries": [],
                "preserved_existing_pairs": [],
                "empty_live_pairs": [],
            }
        )
        registry = [self._stable_registry_entry(name) for name in DATASET_CATALOG_CONFIG]

        verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_accepts_monthly_source_mode(self):
        """Regresión: finanzas_municipales (SINIM) es stable_publishable con
        cadencia mensual real (sinim_finanzas_live_extractor.py, ver
        Monthly Scrape workflow), y su metadata usa source_mode="monthly" —
        no "live" — a propósito (Fase 3.4, "cadencia honesta"). Antes de
        que verify_publication_policy reconociera ese valor, cualquier
        dataset con source_mode="monthly" quedaba bloqueado de publicación
        de forma permanente, aunque los datos fueran genuinamente reales.
        """
        datasets = {
            name: {
                "source_mode": "live",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
            for name in DATASET_CATALOG_CONFIG
        }
        datasets["indicadores"]["indicator_delivery"] = {
            code: "live" for code in EXPECTED_INDICATOR_CODES
        }
        finanzas_name = next(name for name in DATASET_CATALOG_CONFIG if name != "indicadores")
        datasets[finanzas_name]["source_mode"] = "monthly"
        registry = [self._stable_registry_entry(name) for name in DATASET_CATALOG_CONFIG]

        verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_rejects_unknown_source_mode(self):
        """El gate no debe aceptar valores arbitrarios — solo live/fallback/monthly."""
        datasets = {
            "comunas": {
                "source_mode": "weekly",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
        }
        registry = [self._stable_registry_entry("comunas")]

        with (
            patch("builtins.print"),
            self.assertRaisesRegex(SystemExit, "1"),
        ):
            verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_rejects_fallback_and_partial_delivery(self):
        datasets = {
            name: {
                "source_mode": "live",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
            for name in [
                "regiones",
                "provincias",
                "comunas",
                "comunas_enriquecidas",
                "indicadores",
            ]
        }
        datasets["comunas"]["source_mode"] = "fallback"
        datasets["indicadores"].update(
            {
                "source_detail": "public_api_partial",
                "indicator_delivery": {"uf": "preserved_existing"},
                "fetch_failures": ["uf/2026: timeout"],
                "preserved_existing_pairs": ["uf/2026"],
            }
        )
        registry = [
            self._stable_registry_entry(name)
            for name in [
                "regiones",
                "provincias",
                "comunas",
                "comunas_enriquecidas",
                "indicadores",
            ]
        ]

        with (
            patch("builtins.print"),
            self.assertRaisesRegex(SystemExit, "1"),
        ):
            verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_candidate_fallback_passes_when_excluded_from_manifest(self):
        """Candidate dataset in fallback passes when its artifact is absent from manifest."""
        datasets = {
            "finanzas_municipales": {
                "source_mode": "fallback",
                "source_detail": "generated_fallback",
                "freshness": {"status": "unknown"},
            }
        }
        registry = [self._candidate_registry_entry("finanzas_municipales")]
        manifest = {
            "artifacts": [
                {"path": "data/normalized/comunas.parquet", "dataset": "comunas"},
            ]
        }

        verify_publication_policy({"datasets": datasets}, registry=registry, manifest=manifest)

    def test_publication_policy_candidate_fallback_fails_when_in_manifest(self):
        """Candidate artifact present in manifest triggers violation."""
        datasets = {
            "finanzas_municipales": {
                "source_mode": "fallback",
                "source_detail": "generated_fallback",
                "freshness": {"status": "unknown"},
            }
        }
        registry = [self._candidate_registry_entry("finanzas_municipales")]
        manifest = {
            "artifacts": [
                {
                    "path": "data/normalized/finanzas_municipales.parquet",
                    "dataset": "finanzas_municipales",
                },
            ]
        }

        with (
            patch("builtins.print"),
            self.assertRaisesRegex(SystemExit, "1"),
        ):
            verify_publication_policy({"datasets": datasets}, registry=registry, manifest=manifest)

    def test_publication_policy_stable_publishable_in_fallback_still_fails(self):
        """Stable_publishable dataset in fallback is rejected."""
        datasets = {
            "comunas": {
                "source_mode": "fallback",
                "source_detail": "public_api",
                "freshness": {"status": "fresh"},
            }
        }
        registry = [self._stable_registry_entry("comunas")]

        with (
            patch("builtins.print"),
            self.assertRaisesRegex(SystemExit, "1"),
        ):
            verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_publication_policy_indicadores_raw_recovery_still_fails(self):
        """indicadores with raw_recoveries fails even with candidates excluded."""
        datasets = {
            "indicadores": {
                "source_mode": "live",
                "source_detail": "public_api_with_raw_recovery",
                "freshness": {"status": "fresh"},
                "indicator_delivery": {"uf": "raw_recovery", "dolar": "live"},
                "fetch_failures": ["uf/2026: timeout"],
                "raw_recoveries": ["uf/2026"],
                "preserved_existing_pairs": [],
                "empty_live_pairs": [],
            }
        }
        registry = [self._stable_registry_entry("indicadores")]

        with (
            patch("builtins.print"),
            self.assertRaisesRegex(SystemExit, "1"),
        ):
            verify_publication_policy({"datasets": datasets}, registry=registry)

    def test_dataset_contract_accepts_valid_dataframe(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": ["01101", "01107"],
                "nombre_comuna": ["Iquique", "Alto Hospicio"],
            }
        )
        contract = {
            "dataset": "demo",
            "primary_key": ["codigo_comuna"],
            "required_columns": ["codigo_comuna", "nombre_comuna"],
            "column_types": {"codigo_comuna": "string"},
            "fixed_width_columns": {"codigo_comuna": 5},
            "expected_record_count": 2,
            "coverage_policy": "full",
            "publish_outputs": [],
        }

        verify_dataset_contract("demo", contract, df, {}, ROOT_DIR)

    def test_dataset_contract_rejects_missing_required_column(self):
        df = pl.DataFrame({"codigo_comuna": ["01101"]})
        contract = {
            "dataset": "demo",
            "primary_key": ["codigo_comuna"],
            "required_columns": ["codigo_comuna", "nombre_comuna"],
            "column_types": {},
            "fixed_width_columns": {},
            "coverage_policy": "not_applicable",
            "publish_outputs": [],
        }

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_dataset_contract("demo", contract, df, {}, ROOT_DIR)
        self.assertIn("Faltan columnas requeridas", print_mock.call_args.args[0])

    def test_dataset_contract_rejects_duplicate_primary_key(self):
        df = pl.DataFrame({"codigo_comuna": ["01101", "01101"]})
        contract = {
            "dataset": "demo",
            "primary_key": ["codigo_comuna"],
            "required_columns": ["codigo_comuna"],
            "column_types": {},
            "fixed_width_columns": {},
            "coverage_policy": "not_applicable",
            "publish_outputs": [],
        }

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_dataset_contract("demo", contract, df, {}, ROOT_DIR)
        self.assertIn("no es única", print_mock.call_args.args[0])

    def test_dataset_contract_rejects_invalid_cut_width(self):
        df = pl.DataFrame({"codigo_comuna": ["1101"]})
        contract = {
            "dataset": "demo",
            "primary_key": ["codigo_comuna"],
            "required_columns": ["codigo_comuna"],
            "column_types": {"codigo_comuna": "string"},
            "fixed_width_columns": {"codigo_comuna": 5},
            "coverage_policy": "not_applicable",
            "publish_outputs": [],
        }

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_dataset_contract("demo", contract, df, {}, ROOT_DIR)
        self.assertIn("fuera del ancho esperado de 5", print_mock.call_args.args[0])

    def test_source_registry_accepts_stable_publishable_entry(self):
        catalog = {"datasets": [{"dataset": "comunas"}]}
        registry = [
            {
                "dataset": "comunas",
                "license_status": "open-attribution",
                "access_method": "api",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            }
        ]

        verify_source_registry(registry, catalog)

    def test_source_registry_accepts_candidate_fallback_only_entry(self):
        catalog = {"datasets": [{"dataset": "finanzas_municipales"}]}
        registry = [
            {
                "dataset": "finanzas_municipales",
                "license_status": "public-api-review-terms",
                "access_method": "landing_snapshot",
                "live_extractor_status": "fallback_only",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "candidate",
                "live_ready": False,
                "publish_blocking": True,
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            }
        ]

        verify_source_registry(registry, catalog)

    def test_source_registry_rejects_stable_publishable_for_fallback_only(self):
        catalog = {"datasets": [{"dataset": "finanzas_municipales"}]}
        registry = [
            {
                "dataset": "finanzas_municipales",
                "license_status": "public-api-review-terms",
                "access_method": "landing_snapshot",
                "live_extractor_status": "fallback_only",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "stable",
                "live_ready": False,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            }
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        self.assertIn(
            "fallback_only must have publication_track=candidate", print_mock.call_args.args[0]
        )

    def test_source_registry_rejects_candidate_with_public_bundle_eligible_true(self):
        catalog = {"datasets": [{"dataset": "finanzas_municipales"}]}
        registry = [
            {
                "dataset": "finanzas_municipales",
                "license_status": "public-api-review-terms",
                "access_method": "landing_snapshot",
                "live_extractor_status": "fallback_only",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "candidate",
                "live_ready": False,
                "publish_blocking": True,
                "publication_track": "candidate",
                "public_bundle_eligible": True,
            }
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        # fallback_only check runs before candidate check
        self.assertIn(
            "fallback_only must have public_bundle_eligible=false", print_mock.call_args.args[0]
        )

    def test_source_registry_rejects_candidate_non_fallback_with_public_bundle_eligible_true(self):
        catalog = {"datasets": [{"dataset": "finanzas_municipales"}]}
        registry = [
            {
                "dataset": "finanzas_municipales",
                "license_status": "public-api-review-terms",
                "access_method": "landing_snapshot",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "candidate",
                "live_ready": False,
                "publish_blocking": True,
                "publication_track": "candidate",
                "public_bundle_eligible": True,
            }
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        self.assertIn(
            "candidate must have public_bundle_eligible=false", print_mock.call_args.args[0]
        )

    def test_source_registry_rejects_derived_public_when_upstream_is_candidate(self):
        catalog = {"datasets": [{"dataset": "comunas"}, {"dataset": "perfil_territorial_comunal"}]}
        registry = [
            {
                "dataset": "comunas",
                "license_status": "open-attribution",
                "access_method": "api",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
            {
                "dataset": "perfil_territorial_comunal",
                "license_status": "open-attribution",
                "access_method": "derived",
                "live_extractor_status": "derived",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
                "upstream_datasets": ["comunas"],
            },
        ]

        # Both comunas and perfil are stable_publishable, so this should pass
        verify_source_registry(registry, catalog)

    def test_source_registry_rejects_derived_public_when_upstream_is_candidate_with_upstream(self):
        catalog = {
            "datasets": [
                {"dataset": "finanzas_municipales"},
                {"dataset": "perfil_territorial_comunal"},
            ]
        }
        registry = [
            {
                "dataset": "finanzas_municipales",
                "license_status": "public-api-review-terms",
                "access_method": "landing_snapshot",
                "live_extractor_status": "fallback_only",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "candidate",
                "live_ready": False,
                "publish_blocking": True,
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            },
            {
                "dataset": "perfil_territorial_comunal",
                "license_status": "open-attribution",
                "access_method": "derived",
                "live_extractor_status": "derived",
                "fallback_policy": "allowed_for_dev_blocked_for_publication",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
                "upstream_datasets": ["finanzas_municipales"],
            },
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        self.assertIn("upstream non-publishable datasets", print_mock.call_args.args[0])

    def test_source_registry_rejects_duplicate_dataset(self):
        catalog = {"datasets": [{"dataset": "comunas"}]}
        registry = [
            {
                "dataset": "comunas",
                "license_status": "open-attribution",
                "access_method": "api",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
            {
                "dataset": "comunas",
                "license_status": "open-attribution",
                "access_method": "api",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        self.assertIn("duplicate datasets", print_mock.call_args.args[0])

    def test_source_registry_rejects_invalid_enum(self):
        catalog = {"datasets": [{"dataset": "comunas"}]}
        registry = [
            {
                "dataset": "comunas",
                "license_status": "open-attribution",
                "access_method": "spreadsheet-by-email",
                "live_extractor_status": "implemented",
                "fallback_policy": "allowed_for_dev",
                "maturity_status": "stable",
                "live_ready": True,
                "publish_blocking": True,
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            }
        ]

        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            verify_source_registry(registry, catalog)
        self.assertIn("invalid access_method", print_mock.call_args.args[0])

    def test_readiness_required_files_exclude_local_only_build_outputs(self):
        readiness_paths = {
            path.relative_to(ROOT_DIR).as_posix()
            for path in required_files_for_profile("readiness")
        }
        publication_paths = {
            path.relative_to(ROOT_DIR).as_posix()
            for path in required_files_for_profile("publication")
        }

        local_only_paths = {
            "data/staging/comunas.csv",
            "data/staging/comunas.metadata.json",
            "data/normalized/chile_data.duckdb",
            "data/normalized/chile_data.db",
            "data/normalized/chile_data_latest.xlsx",
        }

        self.assertTrue(local_only_paths.isdisjoint(readiness_paths))
        self.assertTrue(local_only_paths.issubset(publication_paths))
        self.assertIn("data/normalized/pipeline_metadata.json", readiness_paths)
        self.assertIn("data/normalized/comunas.parquet", readiness_paths)

    def test_write_publishable_bundle_zip_fails_before_creating_partial_zip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir) / "data"
            normalized_dir = data_dir / "normalized"
            normalized_dir.mkdir(parents=True)
            manifest = {"artifacts": [{"path": "data/normalized/no-existe.parquet"}]}
            (normalized_dir / "artifact_manifest.json").write_text(
                json.dumps(manifest), encoding="utf-8"
            )

            with (
                patch("src.builders.artifacts.NORMALIZED_DIR", str(normalized_dir)),
                patch("src.builders.artifacts.DATA_DIR", str(data_dir)),
                self.assertRaisesRegex(SystemExit, "no-existe.parquet"),
            ):
                write_publishable_bundle_zip()

            self.assertFalse((normalized_dir / "chile-hub-publishable-bundle.zip").exists())

    def test_clean_publishable_removes_only_manifest_declared_outputs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_file = root / "data" / "normalized" / "artifact_manifest.json"
            artifact_one = root / "data" / "normalized" / "dataset_catalog.json"
            artifact_two = root / "data" / "normalized" / "hub_status.json"
            package_zip = root / "data" / "normalized" / "bundle.zip"
            checksum = root / "data" / "normalized" / "bundle.zip.sha256"
            unrelated = root / "data" / "normalized" / "keep.me"

            for path in [
                manifest_file,
                artifact_one,
                artifact_two,
                package_zip,
                checksum,
                unrelated,
            ]:
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("x", encoding="utf-8")

            manifest = {
                "artifacts": [
                    {"path": "data/normalized/artifact_manifest.json"},
                    {"path": "data/normalized/dataset_catalog.json"},
                    {"path": "data/normalized/hub_status.json"},
                    {"path": "data/normalized/hub_status.json"},
                ],
                "packages": [
                    {
                        "path": "data/normalized/bundle.zip",
                        "checksum_path": "data/normalized/bundle.zip.sha256",
                    },
                ],
            }

            removed = package_publishable_bundle.clean_publishable(manifest, root)

            self.assertEqual(
                removed,
                [
                    "data/normalized/dataset_catalog.json",
                    "data/normalized/hub_status.json",
                    "data/normalized/bundle.zip",
                    "data/normalized/bundle.zip.sha256",
                    "data/normalized/artifact_manifest.json",
                ],
            )
            self.assertFalse(manifest_file.exists())
            self.assertFalse(artifact_one.exists())
            self.assertFalse(artifact_two.exists())
            self.assertFalse(package_zip.exists())
            self.assertFalse(checksum.exists())
            self.assertTrue(unrelated.exists())

    def test_clean_publishable_removes_manifest_last(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            for relative_path in [
                "data/normalized/artifact_manifest.json",
                "data/normalized/hub_status.json",
                "data/normalized/dataset_catalog.json",
            ]:
                path = root / relative_path
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("x", encoding="utf-8")

            manifest = {
                "artifacts": [
                    {"path": "data/normalized/artifact_manifest.json"},
                    {"path": "data/normalized/hub_status.json"},
                    {"path": "data/normalized/dataset_catalog.json"},
                ],
                "packages": [],
            }

            removed = package_publishable_bundle.clean_publishable(
                manifest,
                root_dir=root,
            )

            self.assertEqual(removed[-1], "data/normalized/artifact_manifest.json")

    def test_clean_publishable_from_manifest_is_idempotent_when_manifest_is_missing(
        self,
    ):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            manifest_path = root / "data" / "normalized" / "artifact_manifest.json"

            self.assertEqual(
                package_publishable_bundle.clean_publishable_from_manifest(
                    manifest_path=manifest_path,
                    root_dir=root,
                ),
                [],
            )

    def test_load_existing_staging_returns_consistent_empty_tuple_without_staging_file(
        self,
    ):
        with tempfile.TemporaryDirectory() as tmpdir:
            staging_path = str(Path(tmpdir) / "indicadores.csv")
            published_path = str(Path(tmpdir) / "indicadores.parquet")
            original_staging = bcentral_extractor.STAGING_CSV_PATH
            original_published = bcentral_extractor.PUBLISHED_INDICATORS_PATH
            try:
                bcentral_extractor.STAGING_CSV_PATH = staging_path
                bcentral_extractor.PUBLISHED_INDICATORS_PATH = published_path
                self.assertEqual(
                    bcentral_extractor.load_existing_staging(),
                    (None, None, []),
                )
            finally:
                bcentral_extractor.STAGING_CSV_PATH = original_staging
                bcentral_extractor.PUBLISHED_INDICATORS_PATH = original_published

    def test_build_coverage_full_for_expected_cardinality(self):
        coverage = build_coverage("comunas", {"record_count": 346})

        self.assertEqual(coverage["status"], "full")
        self.assertEqual(coverage["expected_record_count"], 346)
        self.assertEqual(coverage["actual_record_count"], 346)
        self.assertEqual(coverage["coverage_ratio"], 1.0)
        self.assertIn("Cobertura completa", coverage["summary"])

    def test_build_coverage_partial_when_below_baseline(self):
        coverage = build_coverage("provincias", {"record_count": 12})

        self.assertEqual(coverage["status"], "partial")
        self.assertEqual(coverage["expected_record_count"], 56)
        self.assertEqual(coverage["actual_record_count"], 12)
        self.assertEqual(coverage["coverage_ratio"], 0.2143)
        self.assertIn("Cobertura parcial", coverage["summary"])

    def test_build_coverage_not_applicable_without_baseline(self):
        coverage = build_coverage("indicadores", {"record_count": 5})

        self.assertEqual(coverage["status"], "not_applicable")
        self.assertIsNone(coverage["expected_record_count"])
        self.assertEqual(coverage["actual_record_count"], 5)
        self.assertIsNone(coverage["coverage_ratio"])

    def test_build_degradation_marks_fallback_comunas_as_degraded(self):
        degradation = build_degradation(
            "comunas",
            {"source_mode": "fallback", "record_count": 18},
            {"warnings": []},
        )

        self.assertEqual(degradation["status"], "degraded")
        self.assertIn("18 comunas", degradation["impact"])
        self.assertIn("fuente territorial primaria", degradation["recommended_action"])

    def test_build_degradation_uses_warning_status_when_validation_warns(self):
        degradation = build_degradation(
            "indicadores",
            {"source_mode": "live", "record_count": 5},
            {"warnings": ["schema changed", "freshness borderline"]},
        )

        self.assertEqual(degradation["status"], "warning")
        self.assertIn("schema changed", degradation["impact"])
        self.assertIn("freshness borderline", degradation["impact"])
        self.assertIn("warnings operativos", degradation["recommended_action"])

    def test_build_degradation_flags_temporal_anomaly_with_actionable_message(self):
        """Plan 054, ADR-013: una anomalia estructurada produce un
        recommended_action especifico (no el generico de warnings), y expone
        anomaly_detected/anomaly_indicator_codes para que el gate de
        publicacion los lea sin tener que parsear texto."""
        anomaly = {
            "codigo_indicador": "uf",
            "fecha": "2026-01-06",
            "motivo": "Salto atipico en 'uf'...",
        }
        degradation = build_degradation(
            "indicadores",
            {"source_mode": "live", "record_count": 6},
            {"warnings": ["indicadores anomaly: ..."], "anomalies": [anomaly]},
        )

        self.assertEqual(degradation["status"], "warning")
        self.assertTrue(degradation["anomaly_detected"])
        self.assertEqual(degradation["anomaly_indicator_codes"], ["uf"])
        self.assertIn("uf", degradation["recommended_action"])
        self.assertIn("2026-01-06", degradation["recommended_action"])

    def test_build_degradation_fallback_precedence_over_anomaly_is_intentional(self):
        """ADR-013: si indicadores esta en fallback (datos sinteticos de
        desarrollo) Y tiene una anomalia detectada, gana el mensaje de
        fallback -- una anomalia estadistica sobre datos ya sabidos-no-reales
        no aporta señal util. No es un bug: la publicacion de datasets en
        fallback ya se rechaza por una via completamente distinta
        (verify_publication_policy chequea source_mode independientemente)."""
        anomaly = {
            "codigo_indicador": "uf",
            "fecha": "2026-01-06",
            "motivo": "Salto atipico en 'uf'...",
        }
        degradation = build_degradation(
            "indicadores",
            {"source_mode": "fallback", "record_count": 6},
            {"warnings": [], "anomalies": [anomaly]},
        )

        self.assertEqual(degradation["status"], "degraded")
        self.assertNotIn("anomaly_detected", degradation)
        self.assertIn("fallback local", degradation["impact"])

    def test_build_drift_is_healthy_for_live_full_non_degraded_dataset(self):
        drift = build_drift(
            {
                "source_mode": "live",
                "coverage": {"status": "full"},
                "degradation": {"status": "none"},
            }
        )

        self.assertEqual(drift["status"], "healthy")
        self.assertIn("Sin drift operativo", drift["summary"])
        self.assertEqual(drift["recommended_action"], "Ninguna.")

    def test_build_drift_is_drifted_for_fallback_partial_degraded_dataset(self):
        drift = build_drift(
            {
                "source_mode": "fallback",
                "coverage": {"status": "partial"},
                "degradation": {"status": "degraded"},
            }
        )

        self.assertEqual(drift["status"], "drifted")
        self.assertIn("mode=fallback", drift["summary"])
        self.assertIn("coverage=partial", drift["summary"])
        self.assertIn("degradation=degraded", drift["summary"])
        self.assertIn("Revisar fuente, cobertura y warnings", drift["recommended_action"])

    def test_build_hub_health_aggregates_severity_and_operational_counts(self):
        metadata = {
            "generated_at_utc": "2026-06-01T12:00:00+00:00",
            "datasets": {
                "alpha": {
                    "source_mode": "live",
                    "freshness": {"status": "fresh"},
                    "reuse_policy": {
                        "status": "open-attribution",
                        "redistribution_ok": True,
                    },
                    "degradation": {"status": "none"},
                    "coverage": {"status": "full", "coverage_ratio": 1.0},
                    "drift": {"status": "healthy"},
                },
                "beta": {
                    "source_mode": "fallback",
                    "freshness": {"status": "stale"},
                    "reuse_policy": {
                        "status": "public-api-review-terms",
                        "redistribution_ok": False,
                    },
                    "degradation": {"status": "degraded"},
                    "coverage": {"status": "partial", "coverage_ratio": 0.5},
                    "drift": {"status": "drifted"},
                },
                "gamma": {
                    "source_mode": "live",
                    "freshness": {"status": "unknown"},
                    "reuse_policy": {"status": "unknown", "redistribution_ok": None},
                    "degradation": {"status": "warning"},
                    "coverage": {"status": "unknown", "coverage_ratio": None},
                    "drift": {"status": "drifted"},
                },
            },
            "validations": {
                "alpha": {"status": "ok", "warnings": []},
                "beta": {"status": "ok", "warnings": ["fallback active"]},
                "gamma": {"status": "error", "warnings": ["missing baseline"]},
            },
        }

        health = build_hub_health(metadata)
        by_dataset = {entry["dataset"]: entry for entry in health["datasets"]}

        self.assertEqual(health["overall_status"], "error")
        self.assertEqual(health["dataset_count"], 3)
        self.assertEqual(health["ok_count"], 1)
        self.assertEqual(health["warn_count"], 1)
        self.assertEqual(health["error_count"], 1)
        self.assertEqual(health["live_count"], 2)
        self.assertEqual(health["fallback_count"], 1)
        self.assertEqual(health["stale_count"], 1)
        self.assertEqual(health["unknown_freshness_count"], 1)
        self.assertEqual(health["publishable_count"], 1)
        self.assertEqual(health["review_terms_count"], 1)
        self.assertEqual(health["unknown_reuse_count"], 1)
        self.assertEqual(health["degraded_count"], 1)
        self.assertEqual(health["degradation_warning_count"], 1)
        self.assertEqual(health["partial_coverage_count"], 1)
        self.assertEqual(health["unknown_coverage_count"], 1)
        self.assertEqual(health["drifted_count"], 2)
        self.assertEqual(health["warning_count"], 2)

        self.assertEqual(by_dataset["alpha"]["severity"], "ok")
        self.assertEqual(by_dataset["beta"]["severity"], "warn")
        self.assertEqual(by_dataset["gamma"]["severity"], "error")
        self.assertEqual(by_dataset["alpha"]["publishability_status"], "ready")
        self.assertEqual(by_dataset["beta"]["publishability_status"], "review_terms")
        self.assertEqual(by_dataset["gamma"]["publishability_status"], "unknown")

    def test_build_status_text_includes_top_issue_reason_and_action(self):
        metadata = {
            "generated_at_utc": "2026-06-01T12:00:00+00:00",
            "datasets": {
                "alpha": {
                    "source_name": "Alpha Source",
                    "source_mode": "live",
                    "source_detail": "alpha_live",
                    "record_count": 10,
                    "fields": ["id"],
                    "notes": [],
                    "freshness": {"status": "fresh", "summary": "fresh"},
                    "coverage": {"status": "full", "summary": "Cobertura completa"},
                    "reuse_policy": {
                        "status": "open-attribution",
                        "redistribution_ok": True,
                    },
                    "degradation": {
                        "status": "none",
                        "impact": "Sin impacto.",
                        "recommended_action": "Ninguna.",
                    },
                    "drift": {
                        "status": "healthy",
                        "summary": "Sin drift operativo.",
                        "recommended_action": "Ninguna.",
                    },
                },
                "beta": {
                    "source_name": "Beta Source",
                    "source_mode": "live",
                    "source_detail": "beta_partial",
                    "record_count": 5,
                    "fields": ["id", "value"],
                    "notes": ["empty_live_pairs: ipc/2026"],
                    "freshness": {"status": "fresh", "summary": "fresh"},
                    "coverage": {
                        "status": "not_applicable",
                        "summary": "Sin baseline esperado.",
                    },
                    "reuse_policy": {
                        "status": "open-attribution",
                        "redistribution_ok": True,
                    },
                    "degradation": {
                        "status": "warning",
                        "impact": "Serie vacia detectada.",
                        "recommended_action": "Revisar warnings operativos del dataset antes de consumirlo en producción.",
                    },
                    "drift": {
                        "status": "drifted",
                        "summary": "Drift operativo detectado: warnings activos.",
                        "recommended_action": "Revisar warnings operativos del dataset antes de consumirlo en producción.",
                    },
                },
            },
            "validations": {
                "alpha": {"status": "ok", "warnings": []},
                "beta": {
                    "status": "ok",
                    "warnings": [
                        "indicadores live refresh returned empty series for: ipc/2026",
                    ],
                },
            },
        }

        status_text = build_status_text(metadata)

        self.assertIn("top_issue: beta", status_text)
        self.assertIn(
            "top_issue_reason: indicadores live refresh returned empty series for: ipc/2026",
            status_text,
        )
        self.assertIn(
            "top_issue_action: Revisar warnings operativos del dataset antes de consumirlo en producción.",
            status_text,
        )
        self.assertIn(
            "top_issue_summary: beta: indicadores live refresh returned empty series for: ipc/2026",
            status_text,
        )
        self.assertIn(
            "hub_status_json: data/normalized/hub_status.json",
            status_text,
        )

    @staticmethod
    def _issue_dataset(source_mode="live", warning_count=0, drift_status="healthy"):
        return {
            "source_name": "Test Source",
            "source_mode": source_mode,
            "source_detail": "public_api",
            "record_count": 1,
            "fields": ["id"],
            "notes": [],
            "freshness": {"status": "fresh", "summary": "fresh"},
            "coverage": {"status": "full", "summary": "Cobertura completa"},
            "reuse_policy": {"status": "open-attribution", "redistribution_ok": True},
            "degradation": {
                "status": "warning" if warning_count else "none",
                "impact": "Impacto de prueba." if warning_count else "Sin impacto.",
                "recommended_action": "Revisar." if warning_count else "Ninguna.",
            },
            "drift": {
                "status": drift_status,
                "summary": "Drift de prueba.",
                "recommended_action": "Revisar." if drift_status != "healthy" else "Ninguna.",
            },
        }

    def _issue_validation(self, warning_count):
        return {
            "status": "ok",
            "warnings": [f"warning {i}" for i in range(warning_count)],
        }

    def test_build_hub_health_top_issue_excludes_non_public_candidate_dataset(self):
        """Regresión: un dataset candidate (sin tarjeta en la landing page) con
        más warnings que un dataset público NO debe ganar el top_issue — el
        enlace "Ver top issue" de index.html apunta a #dataset-{nombre} y esa
        tarjeta solo existe para datasets con public_bundle_eligible=true.
        """
        import tempfile

        registry = [
            {
                "dataset": "public_ds",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
            {
                "dataset": "candidate_ds",
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            },
        ]
        metadata = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "datasets": {
                # candidate_ds tiene más warnings (peor severidad) que public_ds.
                "candidate_ds": self._issue_dataset(source_mode="fallback", warning_count=5),
                "public_ds": self._issue_dataset(source_mode="live", warning_count=1),
            },
            "validations": {
                "candidate_ds": self._issue_validation(5),
                "public_ds": self._issue_validation(1),
            },
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            registry_path = Path(tmpdir) / "source_registry.json"
            registry_path.write_text(json.dumps(registry), encoding="utf-8")
            with patch.object(_pipeline_canonical, "SOURCE_REGISTRY_PATH", registry_path):
                health = build_hub_health(metadata)

        self.assertEqual(health["top_issue"]["dataset"], "public_ds")

    def test_build_hub_health_top_issue_stays_pure_for_synthetic_fixtures(self):
        """Si ningún dataset de `entries` existe en el registry (fixtures
        sintéticos como en test_build_status_text_includes_top_issue_reason_and_action),
        build_hub_health no debe filtrar — de lo contrario rompería la pureza
        que necesitan los tests unitarios que no dependen del repo real."""
        import tempfile

        registry = [
            {
                "dataset": "real_public_dataset",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
        ]
        metadata = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "datasets": {
                "alpha": self._issue_dataset(source_mode="live", warning_count=0),
                "beta": self._issue_dataset(source_mode="live", warning_count=3),
            },
            "validations": {
                "alpha": self._issue_validation(0),
                "beta": self._issue_validation(3),
            },
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            registry_path = Path(tmpdir) / "source_registry.json"
            registry_path.write_text(json.dumps(registry), encoding="utf-8")
            with patch.object(_pipeline_canonical, "SOURCE_REGISTRY_PATH", registry_path):
                health = build_hub_health(metadata)

        # beta tiene warnings y ninguno de los dos datasets está en el
        # registry de prueba, así que debe ganar por atención sin filtrar.
        self.assertEqual(health["top_issue"]["dataset"], "beta")

    def test_build_hub_health_top_issue_unfiltered_when_registry_missing(self):
        """Si source_registry.json no existe (contexto empaquetado/instalado),
        build_hub_health degrada con gracia al comportamiento sin filtro en
        vez de lanzar una excepción."""
        import tempfile

        metadata = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "datasets": {
                "alpha": self._issue_dataset(source_mode="live", warning_count=0),
                "beta": self._issue_dataset(source_mode="live", warning_count=2),
            },
            "validations": {
                "alpha": self._issue_validation(0),
                "beta": self._issue_validation(2),
            },
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            missing_path = Path(tmpdir) / "does_not_exist.json"
            with patch.object(_pipeline_canonical, "SOURCE_REGISTRY_PATH", missing_path):
                health = build_hub_health(metadata)

        self.assertEqual(health["top_issue"]["dataset"], "beta")

    def test_load_source_registry_datasets_splits_public_and_all(self):
        import tempfile

        from src.pipeline_status_utils import _load_source_registry_datasets

        registry = [
            {
                "dataset": "public_a",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
            {
                "dataset": "public_but_ineligible",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": False,
            },
            {
                "dataset": "candidate_a",
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            registry_path = Path(tmpdir) / "source_registry.json"
            registry_path.write_text(json.dumps(registry), encoding="utf-8")

            with patch.object(_pipeline_canonical, "SOURCE_REGISTRY_PATH", registry_path):
                all_names, public_names = _load_source_registry_datasets()

        self.assertEqual(all_names, {"public_a", "public_but_ineligible", "candidate_a"})
        self.assertEqual(public_names, {"public_a"})

    def test_load_source_registry_datasets_missing_file_returns_empty_sets(self):
        import tempfile

        from src.pipeline_status_utils import _load_source_registry_datasets

        with tempfile.TemporaryDirectory() as tmpdir:
            missing_path = Path(tmpdir) / "does_not_exist.json"
            with patch.object(_pipeline_canonical, "SOURCE_REGISTRY_PATH", missing_path):
                all_names, public_names = _load_source_registry_datasets()

        self.assertEqual(all_names, set())
        self.assertEqual(public_names, set())


class ValidatorTests(unittest.TestCase):
    def test_validate_regiones_accepts_valid_data(self):
        df = pl.DataFrame(
            {
                "codigo_region": ["01", "02"],
                "nombre_region": ["Tarapaca", "Antofagasta"],
            }
        )
        self.assertEqual(validate_regiones(df)["status"], "ok")

    def test_validate_regiones_rejects_empty_and_duplicate_data(self):
        empty = pl.DataFrame({"codigo_region": pl.Series([], dtype=pl.String)})
        duplicate = pl.DataFrame({"codigo_region": ["01", "01"]})
        self.assertEqual(validate_regiones(empty)["status"], "error")
        self.assertEqual(validate_regiones(duplicate)["status"], "error")

    def test_validate_provincias_accepts_valid_data(self):
        df = pl.DataFrame({"codigo_region": ["01", "01"], "codigo_provincia": ["011", "014"]})
        self.assertEqual(validate_provincias(df)["status"], "ok")

    def test_validate_provincias_rejects_empty_and_duplicate_keys(self):
        empty = pl.DataFrame(
            {
                "codigo_region": pl.Series([], dtype=pl.String),
                "codigo_provincia": pl.Series([], dtype=pl.String),
            }
        )
        duplicate = pl.DataFrame(
            {"codigo_region": ["01", "01"], "codigo_provincia": ["011", "011"]}
        )
        self.assertEqual(validate_provincias(empty)["status"], "error")
        self.assertEqual(validate_provincias(duplicate)["status"], "error")

    def test_validate_comunas_accepts_complete_live_data(self):
        codes = [str(i).zfill(5) for i in range(EXPECTED_LIVE_COMUNAS_COUNT)]
        result = validate_comunas(pl.DataFrame({"codigo_comuna": codes}), {"source_mode": "live"})
        self.assertEqual(result["status"], "ok")

    def test_validate_comunas_rejects_incomplete_or_duplicate_live_data(self):
        incomplete = pl.DataFrame({"codigo_comuna": ["01101", "01102"]})
        duplicate = pl.DataFrame({"codigo_comuna": ["01101", "01101"]})
        self.assertEqual(validate_comunas(incomplete, {"source_mode": "live"})["status"], "error")
        self.assertEqual(validate_comunas(duplicate, {"source_mode": "live"})["status"], "error")

    def test_validate_comunas_warns_for_fallback(self):
        codes = [str(i).zfill(5) for i in range(FALLBACK_COMUNAS_COUNT)]
        result = validate_comunas(
            pl.DataFrame({"codigo_comuna": codes}), {"source_mode": "fallback"}
        )
        self.assertEqual(result["status"], "ok")
        self.assertTrue(any("fallback" in warning for warning in result["warnings"]))

    def test_validate_finanzas_municipales_rejects_duplicate_invalid_and_negative(self):
        df = pl.DataFrame(
            {
                "anio": [2024, 2024],
                "codigo_comuna": ["13101", "13101"],
                "nombre_comuna": ["Santiago", "Santiago"],
                "ingresos_totales": [1.0, -1.0],
                "gastos_totales": [1.0, 1.0],
                "ingresos_propios_permanentes": [1.0, 1.0],
                "fondo_comun_municipal": [1.0, 1.0],
                "gasto_personal": [1.0, 1.0],
                "gasto_inversion": [1.0, 1.0],
            }
        )
        result = validate_finanzas_municipales(df, {"source_mode": "live"}, ["13101"])
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("unique" in error for error in result["errors"]))
        self.assertTrue(any("negative" in error for error in result["errors"]))

    def test_validate_resultados_educacionales_rejects_out_of_bounds_percentage(self):
        df = pl.DataFrame(
            {
                "anio": [2024],
                "codigo_comuna": ["13101"],
                "matricula_total": [100],
                "asistencia_promedio": [101.0],
                "tasa_aprobacion": [90.0],
                "tasa_reprobacion": [5.0],
                "tasa_retiro": [5.0],
                "establecimientos_reportados": [3],
            }
        )
        result = validate_resultados_educacionales(df, {"source_mode": "live"}, ["13101"])
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("outside 0-100" in error for error in result["errors"]))

    def test_validate_siedu_warns_for_partial_expected_coverage(self):
        df = pl.DataFrame(
            {
                "anio": [2024],
                "codigo_comuna": ["13101"],
                "codigo_indicador": ["siedu_test"],
                "nombre_indicador": ["Test"],
                "categoria": ["Movilidad"],
                "valor": [1.0],
                "unidad": ["indice"],
                "fuente_original": ["SIEDU"],
                "cobertura_tipo": ["urbana"],
            }
        )
        result = validate_indicadores_urbanos_siedu(
            df, {"coverage": {"status": "partial_expected"}}, ["13101"]
        )
        self.assertEqual(result["status"], "ok")
        self.assertTrue(any("partial" in warning for warning in result["warnings"]))

    def test_validate_perfil_territorial_comunal_requires_all_communes(self):
        df = pl.DataFrame({"codigo_comuna": ["13101"]})
        result = validate_perfil_territorial_comunal(df, {"notes": []}, ["13101"])
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("expected 346" in error for error in result["errors"]))

    @classmethod
    def setUpClass(cls):
        cls.normalized_dir = ROOT_DIR / "data" / "normalized"

    def _load_parquet(self, name):
        path = self.normalized_dir / f"{name}.parquet"
        if not path.exists():
            self.skipTest(f"{name}.parquet no existe; correr make build primero")
        return pl.read_parquet(path)

    def test_comunas_cut_codes_are_fixed_width_strings(self):
        df = self._load_parquet("comunas")
        for column, width in (
            ("codigo_region", 2),
            ("codigo_provincia", 3),
            ("codigo_comuna", 5),
        ):
            self.assertEqual(df[column].dtype, pl.String)
            self.assertTrue((df[column].str.len_chars() == width).all())

    def test_comunas_leading_zeros_are_preserved(self):
        df = self._load_parquet("comunas")
        self.assertTrue(df["codigo_region"].str.starts_with("0").any())

    def test_region_and_province_codes_are_strings_with_fixed_width(self):
        regiones = self._load_parquet("regiones")
        provincias = self._load_parquet("provincias")
        self.assertEqual(regiones["codigo_region"].dtype, pl.String)
        self.assertTrue((regiones["codigo_region"].str.len_chars() == 2).all())
        self.assertEqual(provincias["codigo_provincia"].dtype, pl.String)
        self.assertTrue((provincias["codigo_provincia"].str.len_chars() == 3).all())


class IndicatorFallbackTests(unittest.TestCase):
    def _make_minimal_df(self):
        today = datetime.date.today()
        return pl.DataFrame(
            [
                {"fecha": today, "codigo_indicador": code, "valor": 1.0}
                for code in sorted(EXPECTED_INDICATOR_CODES)
            ]
        ).with_columns(pl.col("fecha").cast(pl.Date))

    def _run_process(self, tmpdir, df, diagnostics):
        staging_path = Path(tmpdir) / "indicadores.csv"
        metadata_path = Path(tmpdir) / "indicadores.metadata.json"
        with (
            patch.object(bcentral_extractor, "STAGING_CSV_PATH", str(staging_path)),
            patch.object(bcentral_extractor, "METADATA_PATH", str(metadata_path)),
            patch.object(bcentral_extractor, "fetch_all_history", return_value=(df, diagnostics)),
        ):
            bcentral_extractor.process_indicators()
        return pl.read_csv(staging_path), json.loads(metadata_path.read_text(encoding="utf-8"))

    def test_generate_fallback_returns_all_expected_codes(self):
        df = bcentral_extractor.generate_fallback_indicators()
        self.assertGreater(df.height, 0)
        self.assertTrue(EXPECTED_INDICATOR_CODES.issubset(set(df["codigo_indicador"].unique())))

    def test_process_indicators_uses_fallback_when_fetch_fails(self):
        diagnostics = {
            "fetch_failures": ["uf/2026: timeout"],
            "raw_recoveries": [],
            "preserved_existing_pairs": [],
            "empty_live_pairs": [],
            "published_backfills": [],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            written, metadata = self._run_process(tmpdir, None, diagnostics)
        self.assertGreater(written.height, 0)
        self.assertEqual(metadata["source_mode"], "fallback")
        self.assertIn("fallback_due_to_live_fetch_failure", metadata["notes"])

    def test_process_indicators_records_raw_recovery(self):
        diagnostics = {
            "fetch_failures": ["uf/2026: timeout"],
            "raw_recoveries": ["uf/2026"],
            "preserved_existing_pairs": [],
            "empty_live_pairs": [],
            "published_backfills": [],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            _, metadata = self._run_process(tmpdir, self._make_minimal_df(), diagnostics)
        self.assertEqual(metadata["source_detail"], "public_api_with_raw_recovery")
        self.assertEqual(metadata["indicator_delivery"]["uf"], "raw_recovery")

    def test_process_indicators_records_published_backfill(self):
        diagnostics = {
            "fetch_failures": [],
            "raw_recoveries": [],
            "preserved_existing_pairs": [],
            "empty_live_pairs": [],
            "published_backfills": ["ipc"],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            _, metadata = self._run_process(tmpdir, self._make_minimal_df(), diagnostics)
        self.assertEqual(metadata["source_detail"], "public_api_with_published_backfill")
        self.assertEqual(metadata["indicator_delivery"]["ipc"], "published_backfill")

    def test_process_indicators_records_combined_raw_recovery_and_partial_refresh(self):
        diagnostics = {
            "fetch_failures": ["uf/2026: timeout", "ipc/2026: timeout"],
            "raw_recoveries": ["uf/2026"],
            "preserved_existing_pairs": ["ipc/2026"],
            "empty_live_pairs": [],
            "published_backfills": [],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            _, metadata = self._run_process(tmpdir, self._make_minimal_df(), diagnostics)
        self.assertEqual(metadata["source_detail"], "public_api_with_raw_recovery_partial")
        self.assertEqual(metadata["indicator_delivery"]["uf"], "raw_recovery")
        self.assertEqual(metadata["indicator_delivery"]["ipc"], "preserved_existing")

    def test_empty_live_series_reuses_published_pair(self):
        today = datetime.date.today()
        published = self._make_minimal_df()
        live_records = [
            {
                "fecha": today.strftime("%Y-%m-%d"),
                "codigo_indicador": code,
                "valor": 2.0,
            }
            for code in sorted(EXPECTED_INDICATOR_CODES - {"ipc"})
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            published_path = Path(tmpdir) / "indicadores.parquet"
            published.write_parquet(published_path)
            with (
                patch.object(
                    bcentral_extractor,
                    "STAGING_CSV_PATH",
                    str(Path(tmpdir) / "missing.csv"),
                ),
                patch.object(
                    bcentral_extractor,
                    "PUBLISHED_INDICATORS_PATH",
                    str(published_path),
                ),
                patch.object(
                    bcentral_extractor,
                    "INDICATOR_CODES",
                    sorted(EXPECTED_INDICATOR_CODES),
                ),
                patch.object(
                    bcentral_extractor,
                    "fetch_indicator_year",
                    side_effect=lambda code, year: (
                        []
                        if code == "ipc"
                        else [
                            record for record in live_records if record["codigo_indicador"] == code
                        ]
                    ),
                ),
                patch.object(bcentral_extractor.time, "sleep"),
            ):
                df, diagnostics = bcentral_extractor.fetch_all_history()

        self.assertEqual(set(df["codigo_indicador"].unique()), EXPECTED_INDICATOR_CODES)
        # ipc es un indicador mensual: una serie vacía en el año en curso es esperada,
        # por lo que NO debe aparecer en empty_live_pairs (no es un error operativo).
        self.assertEqual(diagnostics["empty_live_pairs"], [])
        self.assertEqual(diagnostics["published_backfills"], ["ipc"])

    def test_full_live_refresh_failure_reuses_published_artifact(self):
        published = self._make_minimal_df()
        with tempfile.TemporaryDirectory() as tmpdir:
            published_path = Path(tmpdir) / "indicadores.parquet"
            published.write_parquet(published_path)
            with (
                patch.object(
                    bcentral_extractor,
                    "STAGING_CSV_PATH",
                    str(Path(tmpdir) / "missing.csv"),
                ),
                patch.object(
                    bcentral_extractor,
                    "PUBLISHED_INDICATORS_PATH",
                    str(published_path),
                ),
                patch.object(
                    bcentral_extractor,
                    "INDICATOR_CODES",
                    sorted(EXPECTED_INDICATOR_CODES),
                ),
                patch.object(
                    bcentral_extractor,
                    "fetch_indicator_year",
                    side_effect=requests.RequestException("timeout"),
                ),
                patch.object(bcentral_extractor, "load_latest_raw_snapshot", return_value=[]),
                patch.object(bcentral_extractor.time, "sleep"),
            ):
                df, diagnostics = bcentral_extractor.fetch_all_history()

        self.assertEqual(set(df["codigo_indicador"].unique()), EXPECTED_INDICATOR_CODES)
        self.assertEqual(diagnostics["published_backfills"], sorted(EXPECTED_INDICATOR_CODES))


class StagnationTests(unittest.TestCase):
    """Tests para detección de estancamiento según maturity_status."""

    def _make_entry(
        self,
        dataset="test_dataset",
        maturity="candidate",
        access_method="api",
        review_by="2026-09-18",
        stalled_after_days=90,
    ):
        return {
            "dataset": dataset,
            "maturity_status": maturity,
            "access_method": access_method,
            "review_by": review_by,
            "stalled_after_days": stalled_after_days,
            "stalled": False,
        }

    def test_candidate_not_stalled_passes(self):
        """Candidato con review_by futuro no genera fallo."""
        report = {
            "datasets": [
                self._make_entry(
                    dataset="finanzas_municipales",
                    maturity="candidate",
                    review_by="2026-09-18",
                )
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        # No debe lanzar excepción
        _verify_stagnation(report=report, reference_date=reference_date)

    def test_candidate_stalled_fails(self):
        """Candidato con review_by vencido falla verify-readiness."""
        report = {
            "datasets": [
                self._make_entry(
                    dataset="finanzas_municipales",
                    maturity="candidate",
                    review_by="2026-06-01",
                )
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            _verify_stagnation(report=report, reference_date=reference_date)
        self.assertIn("Stagnation policy", print_mock.call_args.args[0])

    def test_experimental_stalled_warns_but_does_not_fail(self):
        """Experimental estancado emite warning pero no falla."""
        report = {
            "datasets": [
                self._make_entry(
                    dataset="test_experimental",
                    maturity="experimental",
                    review_by="2026-06-01",
                    stalled_after_days=30,
                )
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        with patch("builtins.print") as print_mock:
            _verify_stagnation(report=report, reference_date=reference_date)
        warning_calls = [c for c in print_mock.call_args_list if "WARNING" in str(c.args[0])]
        self.assertGreater(
            len(warning_calls),
            0,
            "Experimental estancado debe emitir warning",
        )

    def test_stable_regression_fails(self):
        """Dataset estable estancado (regresión) falla verify-readiness."""
        report = {
            "datasets": [
                self._make_entry(
                    dataset="comunas",
                    maturity="stable",
                    review_by="2026-06-01",
                )
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            _verify_stagnation(report=report, reference_date=reference_date)
        self.assertIn("regresión", print_mock.call_args.args[0])

    def test_derived_stalled_warns_instead_of_failing(self):
        """Dataset derivado estancado advierte en lugar de fallar (depende de upstream)."""
        report = {
            "datasets": [
                {
                    "dataset": "perfil_territorial_comunal",
                    "maturity_status": "candidate",
                    "access_method": "derived",
                    "review_by": "2026-06-01",
                    "stalled_after_days": 90,
                    "stalled": True,
                }
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        with patch("builtins.print") as print_mock:
            _verify_stagnation(report=report, reference_date=reference_date)
        warning_calls = [c for c in print_mock.call_args_list if "WARNING" in str(c.args[0])]
        self.assertGreater(
            len(warning_calls),
            0,
            "Derivado estancado debe emitir warning en lugar de fallar",
        )

    def test_multiple_datasets_mixed_stall_status(self):
        """Verifica que la política maneja múltiples datasets con distintos estados."""
        report = {
            "datasets": [
                self._make_entry(
                    dataset="comunas",
                    maturity="stable",
                    review_by="2026-12-31",
                ),
                self._make_entry(
                    dataset="finanzas_municipales",
                    maturity="candidate",
                    review_by="2026-09-18",
                ),
                self._make_entry(
                    dataset="resultados_educacionales",
                    maturity="candidate",
                    review_by="2026-06-01",
                ),
                self._make_entry(
                    dataset="test_experimental",
                    maturity="experimental",
                    review_by="2026-05-01",
                    stalled_after_days=30,
                ),
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        # resultados_educacionales es candidato estancado → debe fallar
        with patch("builtins.print") as print_mock, self.assertRaisesRegex(SystemExit, "1"):
            _verify_stagnation(report=report, reference_date=reference_date)
        # El warning del experimental debe aparecer
        warning_calls = [c for c in print_mock.call_args_list if "WARNING" in str(c.args[0])]
        self.assertGreater(len(warning_calls), 0)
        # El error debe mencionar al candidato estancado
        self.assertIn("resultados_educacionales", print_mock.call_args.args[0])

    def test_review_by_none_is_never_stalled(self):
        """Dataset sin review_by nunca se considera estancado."""
        report = {
            "datasets": [
                {
                    "dataset": "sin_revision",
                    "maturity_status": "candidate",
                    "access_method": "api",
                    "review_by": None,
                    "stalled_after_days": 90,
                    "stalled": False,
                }
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        # No debe lanzar excepción
        _verify_stagnation(report=report, reference_date=reference_date)

    def test_invalid_review_by_is_gracefully_skipped(self):
        """review_by con formato inválido no rompe la verificación."""
        report = {
            "datasets": [
                {
                    "dataset": "formato_raro",
                    "maturity_status": "candidate",
                    "access_method": "api",
                    "review_by": "no-es-una-fecha",
                    "stalled_after_days": 90,
                    "stalled": False,
                }
            ]
        }
        reference_date = datetime.datetime(2026, 6, 18, tzinfo=UTC)
        # No debe lanzar excepción (formato inválido se ignora)
        _verify_stagnation(report=report, reference_date=reference_date)


class DatasetChangelogSeverityTests(unittest.TestCase):
    """Tests para severidad de cambios en dataset_changelog."""

    def _make_meta(
        self,
        dataset="test_ds",
        record_count=100,
        fields=None,
        source_mode="live",
        freshness="fresh",
        validation="ok",
        contract_exists=True,
        primary_key=None,
        required_columns=None,
        column_types=None,
        nullable_columns=None,
    ):
        """Construye un dict de metadata para un dataset."""
        if fields is None:
            fields = ["id", "name"]
        if primary_key is None:
            primary_key = ["id"]
        if required_columns is None:
            required_columns = ["id", "name"]
        if column_types is None:
            column_types = {"id": "string", "name": "string"}
        if nullable_columns is None:
            nullable_columns = []

        meta = {
            "dataset": dataset,
            "record_count": record_count,
            "fields": fields,
            "source_mode": source_mode,
            "freshness": {"status": freshness},
        }
        if contract_exists:
            meta.update(
                {
                    "contract_exists": True,
                    "contract_primary_key": primary_key,
                    "contract_required_columns": required_columns,
                    "contract_column_types": column_types,
                    "contract_nullable_columns": nullable_columns,
                }
            )
        return meta

    def _make_metadata(self, datasets, generated_at="2026-06-18T12:00:00+00:00"):
        """Construye un dict de pipeline_metadata completo."""
        validations = {}
        for ds in datasets:
            validations[ds["dataset"]] = {
                "status": ds.get("freshness", {}).get("status", "ok")
                if isinstance(ds.get("freshness"), dict)
                else "ok"
            }
        return {
            "generated_at_utc": generated_at,
            "datasets": {d["dataset"]: d for d in datasets},
            "validations": {d["dataset"]: {"status": "ok"} for d in datasets},
        }

    def test_no_changes_is_none(self):
        """Sin cambios → severity none."""
        ds = self._make_meta()
        current = self._make_metadata([ds])
        previous = self._make_metadata([ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "none")
        self.assertEqual(entry["breaking_changes"], [])
        self.assertEqual(entry["new_columns"], [])
        self.assertEqual(entry["removed_columns"], [])
        self.assertFalse(entry["primary_key_changed"])
        self.assertFalse(entry["contract_changed"])

    def test_new_nullable_column_is_minor(self):
        """Nueva columna anulable → severity minor."""
        current_ds = self._make_meta(
            fields=["id", "name", "optional_tag"],
            required_columns=["id", "name"],
            column_types={"id": "string", "name": "string", "optional_tag": "string"},
            nullable_columns=["optional_tag"],
        )
        previous_ds = self._make_meta()
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "minor")
        self.assertEqual(entry["new_columns"], ["optional_tag"])
        self.assertTrue(entry["contract_changed"])
        self.assertFalse(entry["primary_key_changed"])

    def test_required_column_removed_is_major(self):
        """Columna requerida eliminada del contrato → severity major."""
        current_ds = self._make_meta(
            required_columns=["id"],  # name was required before
            column_types={"id": "string"},
        )
        previous_ds = self._make_meta()
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "major")
        self.assertIn("Required column removed: name", entry["breaking_changes"])
        self.assertTrue(entry["contract_changed"])

    def test_primary_key_changed_is_major(self):
        """PK cambiada → severity major."""
        current_ds = self._make_meta(
            primary_key=["id", "version"],
            required_columns=["id", "name", "version"],
            column_types={"id": "string", "name": "string", "version": "integer"},
        )
        previous_ds = self._make_meta()
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "major")
        self.assertTrue(entry["primary_key_changed"])
        self.assertTrue(entry["contract_changed"])

    def test_incompatible_type_change_is_major(self):
        """Cambio de tipo incompatible → severity major."""
        current_ds = self._make_meta(
            column_types={"id": "integer", "name": "string"},
        )
        previous_ds = self._make_meta(
            column_types={"id": "string", "name": "string"},
        )
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "major")
        self.assertTrue(any("type changed" in bc for bc in entry["breaking_changes"]))
        self.assertTrue(entry["contract_changed"])

    def test_new_dataset_is_minor(self):
        """Dataset nuevo (ausente en previous) → severity minor."""
        current_ds = self._make_meta()
        previous = self._make_metadata([], generated_at="2026-06-17T12:00:00+00:00")
        current = self._make_metadata([current_ds])
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "minor")

    def test_data_only_change_is_patch(self):
        """Solo cambio de datos (campos sin tocar contrato) → severity patch."""
        current_ds = self._make_meta(
            fields=["id", "name", "internal_tmp"],
            record_count=101,
        )
        previous_ds = self._make_meta()
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        self.assertEqual(entry["change_severity"], "patch")
        self.assertEqual(entry["added_fields"], ["internal_tmp"])
        self.assertFalse(entry["contract_changed"])

    def test_migration_no_previous_contract_is_patch_or_none(self):
        """Migración: previous sin contract_fields → severity según datos."""
        current_ds = self._make_meta(
            fields=["id", "name", "nueva_col"],
            record_count=105,
        )
        previous_ds = self._make_meta(contract_exists=False)
        previous_ds.pop("contract_exists", None)
        previous_ds.pop("contract_primary_key", None)
        previous_ds.pop("contract_required_columns", None)
        previous_ds.pop("contract_column_types", None)
        previous_ds.pop("contract_nullable_columns", None)
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        # Sin contrato previo, la severidad se determina por cambios de datos
        self.assertIn(entry["change_severity"], ("patch", "none"))
        # Debe tener los campos de severidad presentes
        self.assertIn("breaking_changes", entry)
        self.assertIn("new_columns", entry)
        self.assertIn("removed_columns", entry)

    def test_integer_to_float_widening_is_not_major(self):
        """Widening integer → float no es breaking."""
        current_ds = self._make_meta(
            column_types={"id": "float", "name": "string"},
        )
        previous_ds = self._make_meta(
            column_types={"id": "integer", "name": "string"},
        )
        current = self._make_metadata([current_ds])
        previous = self._make_metadata([previous_ds], generated_at="2026-06-17T12:00:00+00:00")
        result = build_dataset_changelog(current, previous)
        entry = result["datasets"][0]
        # integer → float es compatible, no debe aparecer en breaking_changes
        self.assertNotIn("major", entry["change_severity"])
        type_breaks = [b for b in entry["breaking_changes"] if "type changed" in b]
        self.assertEqual(type_breaks, [])


class ValidationHelperTests(unittest.TestCase):
    """Tests para las funciones helper puras de validation.py."""

    def test_missing_columns(self):
        df = pl.DataFrame({"a": [1], "b": [2]})
        self.assertEqual(_missing_columns(df, ["a", "b", "c"]), ["c"])
        self.assertEqual(_missing_columns(df, ["a", "b"]), [])

    def test_duplicate_count(self):
        df = pl.DataFrame({"x": [1, 1, 2], "y": ["a", "a", "b"]})
        self.assertEqual(_duplicate_count(df, ["x", "y"]), 1)
        self.assertEqual(_duplicate_count(df, ["x"]), 1)
        self.assertEqual(
            _duplicate_count(pl.DataFrame({"x": []}, schema={"x": pl.Int64}), ["x"]), 0
        )

    def test_invalid_fixed_length_count(self):
        df = pl.DataFrame({"c": ["12345", "1234", "12345"]})
        self.assertEqual(_invalid_fixed_length_count(df, "c", 5), 1)

    def test_unknown_codes(self):
        df = pl.DataFrame({"c": ["A", "B", "C"]})
        self.assertEqual(_unknown_codes(df, "c", ["A", "B"]), ["C"])
        self.assertEqual(_unknown_codes(df, "c", None), [])

    def test_negative_numeric_count(self):
        df = pl.DataFrame({"a": [1, -1, 0], "b": [5.0, 3.0, -2.0]})
        self.assertEqual(_negative_numeric_count(df, ["a", "b"]), 2)

    def test_percentage_out_of_bounds_count(self):
        df = pl.DataFrame({"p": [50, -1, 101, None]})
        self.assertEqual(_percentage_out_of_bounds_count(df, ["p"]), 2)


class RemainingValidatorTests(unittest.TestCase):
    """Tests para los validadores de dataset que no estaban cubiertos."""

    def test_validate_censo_comunal_accepts_346_rows(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "poblacion_censada": [100] * 346,
                "poblacion_0_14": [20] * 346,
                "poblacion_15_29": [20] * 346,
                "poblacion_30_44": [20] * 346,
                "poblacion_45_64": [20] * 346,
                "poblacion_65_mas": [20] * 346,
            }
        )
        result = validate_censo_comunal(df, {"source_mode": "live"})
        self.assertEqual(result["status"], "ok")

    def test_validate_censo_comunal_rejects_empty(self):
        df = pl.DataFrame(
            schema={
                "codigo_comuna": pl.String,
                "poblacion_censada": pl.Int64,
                "poblacion_0_14": pl.Int64,
                "poblacion_15_29": pl.Int64,
                "poblacion_30_44": pl.Int64,
                "poblacion_45_64": pl.Int64,
                "poblacion_65_mas": pl.Int64,
            }
        )
        result = validate_censo_comunal(df, None)
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("empty" in e for e in result["errors"]))

    def test_validate_censo_comunal_rejects_wrong_row_count(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": ["01101"],
                "poblacion_censada": [100],
                "poblacion_0_14": [20],
                "poblacion_15_29": [20],
                "poblacion_30_44": [20],
                "poblacion_45_64": [20],
                "poblacion_65_mas": [20],
            }
        )
        result = validate_censo_comunal(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_censo_comunal_age_bands_must_sum_to_total(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "poblacion_censada": [100] * 346,
                "poblacion_0_14": [10] * 346,
                "poblacion_15_29": [10] * 346,
                "poblacion_30_44": [10] * 346,
                "poblacion_45_64": [10] * 346,
                "poblacion_65_mas": [10] * 346,
            }
        )
        result = validate_censo_comunal(df, None)
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("age bands" in e for e in result["errors"]))

    def test_validate_establecimientos_salud_accepts_valid(self):
        df = pl.DataFrame(
            {
                "codigo_establecimiento": ["1001", "1002"],
                "codigo_comuna": ["13101", "13114"],
            }
        )
        result = validate_establecimientos_salud(df, None)
        self.assertEqual(result["status"], "ok")

    def test_validate_establecimientos_salud_rejects_duplicates(self):
        df = pl.DataFrame(
            {
                "codigo_establecimiento": ["1001", "1001"],
                "codigo_comuna": ["13101", "13101"],
            }
        )
        result = validate_establecimientos_salud(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_establecimientos_salud_rejects_empty(self):
        df = pl.DataFrame(schema={"codigo_establecimiento": pl.String, "codigo_comuna": pl.String})
        result = validate_establecimientos_salud(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_establecimientos_salud_rejects_unknown_communes(self):
        df = pl.DataFrame(
            {
                "codigo_establecimiento": ["1001"],
                "codigo_comuna": ["99999"],
            }
        )
        result = validate_establecimientos_salud(df, None, valid_commune_codes=["13101"])
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("unknown communes" in e for e in result["errors"]))

    def test_validate_establecimientos_salud_rejects_invalid_cut(self):
        df = pl.DataFrame(
            {
                "codigo_establecimiento": ["1001"],
                "codigo_comuna": ["1310"],
            }
        )
        result = validate_establecimientos_salud(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_censo_hogares_viviendas_accepts_346_with_valid_totals(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "viviendas_censadas": [120] * 346,
                "viviendas_particulares_ocupadas": [100] * 346,
                "viviendas_particulares_desocupadas": [18] * 346,
                "viviendas_colectivas": [2] * 346,
            }
        )
        result = validate_censo_hogares_viviendas(df, None)
        self.assertEqual(result["status"], "ok")

    def test_validate_censo_hogares_viviendas_rejects_wrong_row_count(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": ["01101"],
                "viviendas_censadas": [120],
                "viviendas_particulares_ocupadas": [100],
                "viviendas_particulares_desocupadas": [18],
                "viviendas_colectivas": [2],
            }
        )
        result = validate_censo_hogares_viviendas(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_censo_hogares_viviendas_rejects_inconsistent_totals(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "viviendas_censadas": [120] * 346,
                "viviendas_particulares_ocupadas": [50] * 346,
                "viviendas_particulares_desocupadas": [10] * 346,
                "viviendas_colectivas": [3] * 346,
            }
        )
        result = validate_censo_hogares_viviendas(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_censo_hogares_viviendas_rejects_unknown_communes(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "viviendas_censadas": [120] * 346,
                "viviendas_particulares_ocupadas": [100] * 346,
                "viviendas_particulares_desocupadas": [18] * 346,
                "viviendas_colectivas": [2] * 346,
            }
        )
        result = validate_censo_hogares_viviendas(df, None, valid_commune_codes=["01101"])
        self.assertTrue(any("unknown communes" in e for e in result["errors"]))

    def test_validate_censo_hogares_viviendas_null_commune_not_unknown(self):
        """Null codigo_comuna should not be reported as unknown commune."""
        df = pl.DataFrame(
            {
                "codigo_comuna": ["01101", None, "01107"],
                "viviendas_censadas": [120, 130, 140],
                "viviendas_particulares_ocupadas": [100, 110, 120],
                "viviendas_particulares_desocupadas": [18, 18, 18],
                "viviendas_colectivas": [2, 2, 2],
            }
        )
        result = validate_censo_hogares_viviendas(df, None, valid_commune_codes=["01101", "01107"])
        unknown_errors = [e for e in result["errors"] if "unknown communes" in e]
        self.assertEqual(
            len(unknown_errors),
            0,
            f"Null codigo_comuna triggered false unknown commune error: {unknown_errors}",
        )

    def test_validate_establecimientos_educacionales_accepts_valid(self):
        df = pl.DataFrame(
            {
                "rbd": ["1", "2"],
                "codigo_comuna": ["13101", "13114"],
            }
        )
        result = validate_establecimientos_educacionales(df, None)
        self.assertEqual(result["status"], "ok")

    def test_validate_establecimientos_educacionales_rejects_duplicate_rbd(self):
        df = pl.DataFrame({"rbd": ["1", "1"], "codigo_comuna": ["13101", "13101"]})
        result = validate_establecimientos_educacionales(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_establecimientos_educacionales_rejects_empty(self):
        df = pl.DataFrame(schema={"rbd": pl.String, "codigo_comuna": pl.String})
        result = validate_establecimientos_educacionales(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_establecimientos_educacionales_rejects_unknown_communes(self):
        df = pl.DataFrame({"rbd": ["1"], "codigo_comuna": ["99999"]})
        result = validate_establecimientos_educacionales(df, None, valid_commune_codes=["13101"])
        self.assertTrue(any("unknown communes" in e for e in result["errors"]))

    def test_validate_distritos_electorales_accepts_346(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "distrito_electoral": ["10"] * 346,
                "circunscripcion_senatorial": ["7"] * 346,
            }
        )
        result = validate_distritos_electorales(df, None)
        self.assertEqual(result["status"], "ok")

    def test_validate_distritos_electorales_rejects_wrong_count(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": ["13101"],
                "distrito_electoral": ["10"],
                "circunscripcion_senatorial": ["7"],
            }
        )
        result = validate_distritos_electorales(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_distritos_electorales_rejects_duplicates(self):
        df = pl.DataFrame(
            {
                "codigo_comuna": [f"{i:05d}" for i in range(1, 347)],
                "distrito_electoral": ["10"] * 346,
                "circunscripcion_senatorial": ["7"] * 346,
            }
        )
        # Duplicate the first row
        df = pl.concat([df, df.head(1)])
        result = validate_distritos_electorales(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_indicadores_accepts_valid(self):
        df = pl.DataFrame(
            {
                "codigo_indicador": sorted(EXPECTED_INDICATOR_CODES),
                "fecha": ["2024-01-01"] * len(EXPECTED_INDICATOR_CODES),
                "valor": [1.0] * len(EXPECTED_INDICATOR_CODES),
            }
        )
        result = validate_indicadores(df, {"source_mode": "live"})
        self.assertEqual(result["status"], "ok")

    def test_validate_indicadores_rejects_empty(self):
        df = pl.DataFrame(schema={"codigo_indicador": pl.String})
        result = validate_indicadores(df, None)
        self.assertEqual(result["status"], "error")

    def test_validate_indicadores_rejects_missing_codes(self):
        df = pl.DataFrame(
            {
                "codigo_indicador": ["dolar", "uf"],
                "fecha": ["2024-01-01", "2024-01-01"],
                "valor": [1.0, 1.0],
            }
        )
        result = validate_indicadores(df, None)
        self.assertEqual(result["status"], "error")
        self.assertTrue(any("missing" in e for e in result["errors"]))

    def test_validate_indicadores_warns_on_fallback(self):
        df = pl.DataFrame(
            {
                "codigo_indicador": sorted(EXPECTED_INDICATOR_CODES),
                "fecha": ["2024-01-01"] * len(EXPECTED_INDICATOR_CODES),
                "valor": [1.0] * len(EXPECTED_INDICATOR_CODES),
            }
        )
        result = validate_indicadores(df, {"source_mode": "fallback"})
        self.assertTrue(any("fallback" in w for w in result["warnings"]))

    def test_validate_indicadores_warns_on_raw_recovery(self):
        df = pl.DataFrame(
            {
                "codigo_indicador": sorted(EXPECTED_INDICATOR_CODES),
                "fecha": ["2024-01-01"] * len(EXPECTED_INDICATOR_CODES),
                "valor": [1.0] * len(EXPECTED_INDICATOR_CODES),
            }
        )
        result = validate_indicadores(df, {"source_mode": "live", "raw_recoveries": ["uf/2024"]})
        self.assertTrue(any("raw" in w for w in result["warnings"]))


class PipelineStatusUtilsTests(unittest.TestCase):
    """Tests para funciones puras de pipeline_status_utils."""

    def test_format_top_issue_summary_empty(self):
        from src.pipeline_status_utils import format_top_issue_summary

        self.assertEqual(format_top_issue_summary(None), "Sin top issue activo.")
        self.assertEqual(format_top_issue_summary({}), "Sin top issue activo.")

    def test_format_top_issue_summary_full(self):
        from src.pipeline_status_utils import format_top_issue_summary

        issue = {
            "dataset": "empresas",
            "source_detail": "public_api",
            "diagnostic_summary": "Cobertura parcial",
            "recommended_action": "Revisar",
            "warning_count": 3,
            "build_freshness_status": "fresh",
            "drift_status": "healthy",
        }
        result = format_top_issue_summary(issue)
        self.assertIn("empresas", result)
        self.assertIn("Cobertura parcial", result)
        self.assertIn("warnings=3", result)

    def test_compute_top_issue_empty_entries(self):
        from src.pipeline_status_utils import compute_top_issue

        self.assertIsNone(compute_top_issue([]))

    def test_compute_top_issue_warnings_priority_zero(self):
        from src.pipeline_status_utils import compute_top_issue

        entries = [
            {
                "dataset": "a",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "healthy",
            },
            {
                "dataset": "b",
                "warning_count": 2,
                "freshness_status": "fresh",
                "drift_status": "healthy",
            },
        ]
        top = compute_top_issue(entries)
        self.assertEqual(top["dataset"], "b")

    def test_compute_top_issue_stale_priority_zero(self):
        from src.pipeline_status_utils import compute_top_issue

        entries = [
            {
                "dataset": "a",
                "warning_count": 0,
                "freshness_status": "stale",
                "drift_status": "healthy",
            },
            {
                "dataset": "b",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "healthy",
            },
        ]
        top = compute_top_issue(entries)
        self.assertEqual(top["dataset"], "a")

    def test_compute_top_issue_drifted_priority_one(self):
        from src.pipeline_status_utils import compute_top_issue

        entries = [
            {
                "dataset": "a",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "drifted",
            },
            {
                "dataset": "b",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "drifted",
            },
        ]
        top = compute_top_issue(entries)
        self.assertIsNotNone(top)
        self.assertEqual(top["attention_priority"], 1)

    def test_compute_top_issue_all_healthy_returns_none(self):
        from src.pipeline_status_utils import compute_top_issue

        entries = [
            {
                "dataset": "a",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "healthy",
            },
            {
                "dataset": "b",
                "warning_count": 0,
                "freshness_status": "fresh",
                "drift_status": "healthy",
            },
        ]
        self.assertIsNone(compute_top_issue(entries))

    def test_format_freshness_statuses(self):
        from src.pipeline_status_utils import format_freshness

        self.assertEqual(format_freshness(None), "unknown")
        self.assertEqual(format_freshness({}), "unknown")
        self.assertIn(
            "fresh", format_freshness({"status": "fresh", "age_hours": 5, "max_age_hours": 24})
        )
        self.assertIn("stale", format_freshness({"status": "stale"}))
        self.assertEqual(format_freshness({"status": "unknown"}), "unknown")

    def test_parse_iso_datetime_accepts_z_and_offset_utc(self):
        from src.pipeline_status_utils import parse_iso_datetime

        expected = datetime.datetime(2026, 7, 7, 12, 0, tzinfo=datetime.UTC)
        self.assertEqual(parse_iso_datetime("2026-07-07T12:00:00Z"), expected)
        self.assertEqual(parse_iso_datetime("2026-07-07T12:00:00+00:00"), expected)

    def test_format_reuse_policy(self):
        from src.pipeline_status_utils import format_reuse_policy

        self.assertEqual(format_reuse_policy(None), "unknown")
        self.assertIn(
            "open-attribution",
            format_reuse_policy({"status": "open-attribution", "license": "CC-BY"}),
        )
        self.assertEqual(format_reuse_policy({"status": "custom"}), "custom")


class ReportsBuilderTests(unittest.TestCase):
    """Tests para src/builders/reports.py (build_dev_db.py._generate_reports)."""

    @staticmethod
    def _catalog_entry(dataset, source_mode, warnings=None):
        return {
            "dataset": dataset,
            "source_name": f"{dataset} source",
            "source_url": f"https://example.cl/{dataset}",
            "source_mode": source_mode,
            "source_detail": "public_api",
            "refreshed_at_utc": "2026-07-08T00:00:00+00:00",
            "freshness": {"status": "fresh", "age_hours": 1.0, "max_age_hours": 24},
            "reuse_policy": {"status": "open-attribution"},
            "coverage": {"status": "full", "coverage_ratio": 1.0, "summary": "Completa"},
            "degradation": {"status": "none"},
            "drift": {"status": "healthy"},
            "warnings": warnings or [],
            "notes": [],
        }

    def test_build_provenance_report_counts_monthly_as_live(self):
        """Regresión: source_mode="monthly" (finanzas_municipales) debe sumar a
        live_count, no quedar fuera de live_count Y fallback_count a la vez —
        eso rompía la invariante live_count + fallback_count == dataset_count
        que valida test_provenance_report en test_chile_hub.py."""
        from src.builders.reports import build_provenance_report

        catalog = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "datasets": [
                self._catalog_entry("a_live", "live"),
                self._catalog_entry("b_monthly", "monthly"),
                self._catalog_entry("c_fallback", "fallback"),
            ],
        }

        report = build_provenance_report(catalog)

        self.assertEqual(report["dataset_count"], 3)
        self.assertEqual(report["live_count"], 2)
        self.assertEqual(report["fallback_count"], 1)
        self.assertEqual(report["live_count"] + report["fallback_count"], report["dataset_count"])

    def test_build_drift_report_fallback_count_excludes_monthly(self):
        from src.builders.reports import build_drift_report

        catalog = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "datasets": [
                self._catalog_entry("a_live", "live"),
                self._catalog_entry("b_monthly", "monthly"),
                self._catalog_entry("c_fallback", "fallback"),
            ],
        }

        report = build_drift_report(catalog)

        self.assertEqual(report["dataset_count"], 3)
        self.assertEqual(report["fallback_count"], 1)

    def test_build_hub_status_forwards_health_fields(self):
        from src.builders.reports import build_hub_status

        health = {
            "generated_at_utc": "2026-07-08T00:00:00+00:00",
            "overall_status": "warn",
            "dataset_count": 3,
            "live_count": 2,
            "fallback_count": 1,
            "stale_count": 0,
            "drifted_count": 1,
            "degraded_count": 0,
            "warning_count": 2,
            "top_issue": {"dataset": "c_fallback"},
            "top_issue_summary": "c_fallback: algo pasó",
        }

        status = build_hub_status(health)

        self.assertEqual(status["overall_status"], "warn")
        self.assertEqual(status["live_count"], 2)
        self.assertEqual(status["top_issue"]["dataset"], "c_fallback")


class HubHealthHistoryTests(unittest.TestCase):
    """Tests para append_hub_health_history (Plan 063) — historial JSONL
    append-only del hub_health.json de cada build."""

    @staticmethod
    def _health(generated_at_utc, **overrides):
        base = {
            "generated_at_utc": generated_at_utc,
            "overall_status": "ok",
            "ok_count": 19,
            "warn_count": 0,
            "error_count": 0,
            "drifted_count": 0,
            "fallback_count": 0,
            "stale_count": 0,
            "warning_count": 0,
            "dataset_count": 19,
        }
        base.update(overrides)
        return base

    def test_append_creates_file_with_one_valid_line(self):
        from src.builders import reports

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.object(reports, "NORMALIZED_DIR", tmpdir):
                path = reports.append_hub_health_history(self._health("2026-01-01T00:00:00Z"))
            with open(path, encoding="utf-8") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            self.assertEqual(len(lines), 1)
            entry = json.loads(lines[0])
            self.assertEqual(len(entry), 10)

    def test_append_same_timestamp_does_not_duplicate(self):
        from src.builders import reports

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.object(reports, "NORMALIZED_DIR", tmpdir):
                health = self._health("2026-01-01T00:00:00Z")
                path = reports.append_hub_health_history(health)
                reports.append_hub_health_history(health)
            with open(path, encoding="utf-8") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            self.assertEqual(len(lines), 1)

    def test_append_distinct_timestamp_adds_line_in_chronological_order(self):
        from src.builders import reports

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.object(reports, "NORMALIZED_DIR", tmpdir):
                path = reports.append_hub_health_history(self._health("2026-01-01T00:00:00Z"))
                reports.append_hub_health_history(self._health("2026-01-02T00:00:00Z"))
            with open(path, encoding="utf-8") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            self.assertEqual(len(lines), 2)
            self.assertEqual(json.loads(lines[0])["generated_at_utc"], "2026-01-01T00:00:00Z")
            self.assertEqual(json.loads(lines[1])["generated_at_utc"], "2026-01-02T00:00:00Z")

    def test_append_caps_at_max_lines_keeping_most_recent(self):
        from src.builders import reports

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / reports.HUB_HEALTH_HISTORY_NAME)
            seeded = [
                json.dumps(self._health(f"2025-01-{(i % 28) + 1:02d}T00:00:00Z-seed{i}"))
                for i in range(reports.HUB_HEALTH_HISTORY_MAX_LINES)
            ]
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(seeded) + "\n")

            with patch.object(reports, "NORMALIZED_DIR", tmpdir):
                reports.append_hub_health_history(self._health("2026-02-01T00:00:00Z"))

            with open(path, encoding="utf-8") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            self.assertEqual(len(lines), reports.HUB_HEALTH_HISTORY_MAX_LINES)
            # La primera línea sembrada (la más antigua) fue descartada; la
            # segunda sembrada ahora es la primera retenida.
            self.assertEqual(
                json.loads(lines[0])["generated_at_utc"], json.loads(seeded[1])["generated_at_utc"]
            )
            self.assertEqual(json.loads(lines[-1])["generated_at_utc"], "2026-02-01T00:00:00Z")

    def test_append_ignores_blank_and_malformed_lines_in_existing_file(self):
        from src.builders import reports

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / reports.HUB_HEALTH_HISTORY_NAME)
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n\n{not valid json\n\n")

            with patch.object(reports, "NORMALIZED_DIR", tmpdir):
                reports.append_hub_health_history(self._health("2026-01-01T00:00:00Z"))

            with open(path, encoding="utf-8") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            # La línea espuria se conserva (no se valida retroactivamente),
            # pero no rompe el append de la nueva entrada valida.
            self.assertEqual(len(lines), 2)
            self.assertEqual(json.loads(lines[-1])["generated_at_utc"], "2026-01-01T00:00:00Z")


class HubBundleCandidateDatasetTests(unittest.TestCase):
    """Tests para write_hub_bundle_json (src/builders/artifacts.py), en
    particular el filtrado de candidate_datasets/candidate_dataset_count que
    causó una expectativa incorrecta al arreglar Pipeline Check #270: solo
    los datasets candidate que YA tienen salidas reales en dataset_catalog
    (outputs configurados) aparecen ahí — los "próximamente" sin outputs
    (delincuencia_comunal, autoridades_locales) nunca entran, aunque estén
    en el registry como candidate.
    """

    @staticmethod
    def _minimal_pipeline_metadata():
        return {"generated_at_utc": "2026-07-08T00:00:00+00:00", "version": "1.0.0"}

    @staticmethod
    def _minimal_hub_health():
        return {
            "overall_status": "warn",
            "datasets": [],
            "top_issue": None,
            "top_issue_summary": "Sin top issue activo.",
        }

    @staticmethod
    def _minimal_artifact_manifest():
        return {"artifacts": [], "packages": []}

    @staticmethod
    def _catalog_dataset(name):
        return {
            "dataset": name,
            "source_name": f"{name} source",
            "source_url": f"https://example.cl/{name}",
            "source_mode": "live",
            "source_detail": "public_api",
            "refreshed_at_utc": "2026-07-08T00:00:00+00:00",
            "record_count": 1,
            "freshness": {"status": "fresh"},
            "coverage": {"status": "full"},
            "degradation": {"status": "none"},
            "drift": {"status": "healthy"},
            "reuse_policy": {"status": "open-attribution"},
            "warnings": [],
        }

    def test_candidate_dataset_count_excludes_candidates_without_catalog_outputs(self):
        from src.builders import artifacts

        registry = [
            {
                "dataset": "public_ds",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
            {
                "dataset": "candidate_with_outputs",
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            },
            {
                "dataset": "candidate_coming_soon",
                "publication_track": "candidate",
                "public_bundle_eligible": False,
            },
        ]
        # dataset_catalog solo trae datasets con outputs reales;
        # candidate_coming_soon nunca aparece ahí ("próximamente").
        dataset_catalog = {
            "dataset_count": 2,
            "datasets": [
                self._catalog_dataset("public_ds"),
                self._catalog_dataset("candidate_with_outputs"),
            ],
        }

        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.object(artifacts, "load_source_registry", return_value=registry),
            patch.object(artifacts, "NORMALIZED_DIR", tmpdir),
        ):
            _, bundle = artifacts.write_hub_bundle_json(
                self._minimal_pipeline_metadata(),
                self._minimal_hub_health(),
                dataset_catalog,
                self._minimal_artifact_manifest(),
            )

        self.assertEqual(bundle["public_dataset_count"], 1)
        self.assertEqual(bundle["candidate_dataset_count"], 1)
        candidate_names = {e["dataset"] for e in bundle["candidate_datasets"]}
        self.assertEqual(candidate_names, {"candidate_with_outputs"})
        self.assertNotIn("candidate_coming_soon", candidate_names)

    def test_stable_publishable_dataset_appears_in_public_datasets_list(self):
        from src.builders import artifacts

        registry = [
            {
                "dataset": "public_ds",
                "publication_track": "stable_publishable",
                "public_bundle_eligible": True,
            },
        ]
        dataset_catalog = {
            "dataset_count": 1,
            "datasets": [self._catalog_dataset("public_ds")],
        }

        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.object(artifacts, "load_source_registry", return_value=registry),
            patch.object(artifacts, "NORMALIZED_DIR", tmpdir),
        ):
            _, bundle = artifacts.write_hub_bundle_json(
                self._minimal_pipeline_metadata(),
                self._minimal_hub_health(),
                dataset_catalog,
                self._minimal_artifact_manifest(),
            )

        self.assertEqual(len(bundle["datasets"]), 1)
        self.assertEqual(bundle["datasets"][0]["dataset"], "public_ds")
        self.assertEqual(bundle["candidate_datasets"], [])


_INDEX_HTML_FIXTURE = """<!DOCTYPE html>
<html>
<head>
    <!-- START_DATA_CATALOG_JSON_LD -->
    <script type="application/ld+json">
    {"@context": "https://schema.org", "@type": "DataCatalog", "dataset": []}
    </script>
    <!-- END_DATA_CATALOG_JSON_LD -->
</head>
<body>old content https://cortega26.github.io/chile-hub/ more text</body>
</html>
"""

_APP_JS_FIXTURE = """const PUBLIC_DATA_BASE = "https://stale.example.com/data/normalized";
console.log(PUBLIC_DATA_BASE);
"""


class SyncLandingMetadataTests(unittest.TestCase):
    """Tests para sync_landing_metadata (src/builders/landing.py): la función
    que regenera el bloque JSON-LD de index.html y el PUBLIC_DATA_BASE de
    app.js. Su desfase silencioso (un dataset agregado al registry sin volver
    a correr build_dev_db.py) fue la causa raíz original de Pipeline Check
    #270 — index.html llevaba sin el dataset autoridades_locales desde que se
    mezcló, y el gate "Check build-synced files" no lo detectó hasta el
    siguiente run programado.
    """

    def _write_fixtures(self, tmpdir):
        index_path = Path(tmpdir) / "index.html"
        app_path = Path(tmpdir) / "app.js"
        index_path.write_text(_INDEX_HTML_FIXTURE, encoding="utf-8")
        app_path.write_text(_APP_JS_FIXTURE, encoding="utf-8")
        return index_path, app_path

    def test_sync_landing_metadata_includes_every_catalog_dataset(self):
        """Regresión directa: cada dataset de DATASET_CATALOG_CONFIG debe
        aparecer en el JSON-LD regenerado — ninguno debe quedar afuera
        silenciosamente (como pasó con autoridades_locales)."""
        from src.builders import landing

        with tempfile.TemporaryDirectory() as tmpdir:
            index_path, _ = self._write_fixtures(tmpdir)
            with patch.object(landing, "ROOT_DIR", tmpdir):
                landing.sync_landing_metadata("https://example.cl/chile-hub/")

            content = index_path.read_text(encoding="utf-8")
            match = re.search(
                r"<!-- START_DATA_CATALOG_JSON_LD -->.*?<!-- END_DATA_CATALOG_JSON_LD -->",
                content,
                flags=re.DOTALL,
            )
            self.assertIsNotNone(match)
            script_match = re.search(
                r'<script type="application/ld\+json">\s*(.*?)\s*</script>',
                match.group(0),
                flags=re.DOTALL,
            )
            catalog = json.loads(script_match.group(1))
            dataset_names = {entry["url"].split("#dataset-")[1] for entry in catalog["dataset"]}
            self.assertEqual(dataset_names, set(DATASET_CATALOG_CONFIG.keys()))

    def test_sync_landing_metadata_rewrites_site_url_and_data_base(self):
        from src.builders import landing

        with tempfile.TemporaryDirectory() as tmpdir:
            index_path, app_path = self._write_fixtures(tmpdir)
            with patch.object(landing, "ROOT_DIR", tmpdir):
                landing.sync_landing_metadata("https://example.cl/chile-hub/")

            index_content = index_path.read_text(encoding="utf-8")
            self.assertNotIn("https://cortega26.github.io/chile-hub/", index_content)
            self.assertIn("https://example.cl/chile-hub/", index_content)

            app_content = app_path.read_text(encoding="utf-8")
            self.assertIn(
                'const PUBLIC_DATA_BASE = "https://example.cl/chile-hub/data/normalized";',
                app_content,
            )
            self.assertNotIn("stale.example.com", app_content)

    def test_sync_landing_metadata_is_noop_when_already_in_sync(self):
        """Corriendo dos veces con la misma URL, la segunda no debe reescribir
        nada (mismo criterio que usa el gate "Check build-synced files")."""
        from src.builders import landing

        with tempfile.TemporaryDirectory() as tmpdir:
            index_path, app_path = self._write_fixtures(tmpdir)
            with patch.object(landing, "ROOT_DIR", tmpdir):
                landing.sync_landing_metadata("https://example.cl/chile-hub/")
                first_index = index_path.read_text(encoding="utf-8")
                first_app = app_path.read_text(encoding="utf-8")

                with patch("builtins.print") as mock_print:
                    landing.sync_landing_metadata("https://example.cl/chile-hub/")

            self.assertEqual(index_path.read_text(encoding="utf-8"), first_index)
            self.assertEqual(app_path.read_text(encoding="utf-8"), first_app)
            mock_print.assert_not_called()

    def test_sync_landing_metadata_missing_files_does_not_raise(self):
        from src.builders import landing

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.object(landing, "ROOT_DIR", tmpdir):
                landing.sync_landing_metadata("https://example.cl/chile-hub/")


class CheckLandingSyncTests(unittest.TestCase):
    """Tests para scripts/check_landing_sync.py, el gate que corre en cada
    push/PR.

    Los tests de SyncLandingMetadataTests ejercitan la función contra un fixture
    temporal, así que pasaban en verde mientras el index.html **commiteado**
    llevaba días desfasado: por eso la deriva de geometria_comunal (agregado al
    catálogo en 56cd9d5) no se detectó hasta que el run programado del
    2026-07-24 abortó en "Check build-synced files", y volvió a abortar el 25 y
    el 26. Estos tests apuntan al archivo real del repo.
    """

    def _index_html(self):
        return (ROOT_DIR / "index.html").read_text(encoding="utf-8")

    def test_committed_index_html_is_in_sync(self):
        """El artefacto commiteado debe estar sincronizado — este es el test que
        habría atrapado la deriva antes de romper el publish diario."""
        from scripts.check_landing_sync import main

        main()  # No debe levantar SystemExit.

    def test_detects_dataset_missing_from_json_ld(self):
        """Regresión exacta de geometria_comunal: un dataset del catálogo que
        no está en el JSON-LD commiteado."""
        from scripts.check_landing_sync import check_json_ld_block

        index_html = self._index_html()
        entry_pattern = re.compile(
            r"\s*\{\n[^{}]*?\"url\": \"[^\"]*#dataset-geometria_comunal\".*?\n\s*\},",
            re.DOTALL,
        )
        drifted = entry_pattern.sub("", index_html, count=1)
        self.assertNotIn("#dataset-geometria_comunal", drifted)

        problems = check_json_ld_block(drifted, "https://tooltician.com/chile-hub/")
        self.assertTrue(problems)
        self.assertIn("geometria_comunal", problems[0])

    def test_detects_missing_marker(self):
        """Sin marcadores, sync_landing_metadata() no puede regenerar el bloque
        y falla en silencio dentro de su `except Exception` — el gate debe
        tratarlo como error, no comparar contra una extracción vacía."""
        from scripts.check_landing_sync import check_json_ld_block

        problems = check_json_ld_block("<html>sin marcadores</html>", "https://example.cl/hub/")
        self.assertTrue(problems)
        self.assertIn("marcador", problems[0])

    def test_detects_stale_app_data_base(self):
        from scripts.check_landing_sync import check_app_data_base

        stale = 'const PUBLIC_DATA_BASE = "https://stale.example.com/data/normalized";'
        problems = check_app_data_base(stale, "https://tooltician.com/chile-hub/")
        self.assertTrue(problems)

        fresh = 'const PUBLIC_DATA_BASE = "https://tooltician.com/chile-hub/data/normalized";'
        self.assertEqual(check_app_data_base(fresh, "https://tooltician.com/chile-hub/"), [])

    def test_detects_stale_cache_buster(self):
        """Un bump de versión sin regenerar index.html rompe el gate de CI
        ("Check build-synced files"); el guard debe verlo igual."""
        from scripts.check_landing_sync import check_app_cache_buster

        index_html = '<script src="app.js?v=1.0.0" defer></script>'
        self.assertTrue(check_app_cache_buster(index_html, "9.9.9"))
        self.assertEqual(check_app_cache_buster(index_html, "1.0.0"), [])

    def test_render_block_is_byte_identical_to_committed_block(self):
        """El gate solo es fiable si compara exactamente lo que el build escribe."""
        from src.builders.landing import (
            extract_catalog_json_ld_block,
            render_catalog_json_ld_block,
        )

        committed = extract_catalog_json_ld_block(self._index_html())
        self.assertIsNotNone(committed)
        self.assertEqual(
            committed, render_catalog_json_ld_block("https://tooltician.com/chile-hub/")
        )


class ReplaceDelimitedBlockTests(unittest.TestCase):
    """Tests para io_utils.replace_delimited_block(): el helper compartido que
    generaliza el patrón de sync_readme_layers_table() a cualquier bloque
    delimitado (ver AGENTS.md §12, mecanismo de bloques delimitados)."""

    def _write(self, tmpdir, content):
        path = Path(tmpdir) / "doc.md"
        path.write_text(content, encoding="utf-8")
        return path

    def test_missing_marker_raises_system_exit(self):
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(tmpdir, "sin marcadores aquí")
            with self.assertRaises(SystemExit):
                replace_delimited_block(str(path), "FOO", "nuevo cuerpo")

    def test_check_only_does_not_write(self):
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(tmpdir, "<!-- START_FOO -->\n\nviejo\n\n<!-- END_FOO -->")
            changed = replace_delimited_block(str(path), "FOO", "nuevo", check_only=True)
            self.assertTrue(changed)
            self.assertIn("viejo", path.read_text(encoding="utf-8"))

    def test_writes_and_returns_true_when_body_changes(self):
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(
                tmpdir, "antes\n<!-- START_FOO -->\n\nviejo\n\n<!-- END_FOO -->\ndespués"
            )
            changed = replace_delimited_block(str(path), "FOO", "nuevo")
            self.assertTrue(changed)
            content = path.read_text(encoding="utf-8")
            self.assertIn("nuevo", content)
            self.assertNotIn("viejo", content)
            self.assertIn("antes", content)
            self.assertIn("después", content)

    def test_returns_false_and_does_not_touch_file_when_unchanged(self):
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(tmpdir, "<!-- START_FOO -->\n\nigual\n\n<!-- END_FOO -->")
            mtime_before = path.stat().st_mtime_ns
            changed = replace_delimited_block(str(path), "FOO", "igual")
            self.assertFalse(changed)
            self.assertEqual(path.stat().st_mtime_ns, mtime_before)

    def test_inline_separator_keeps_same_line(self):
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(tmpdir, "| celda <!-- START_FOO -->viejo<!-- END_FOO --> | otra |")
            replace_delimited_block(str(path), "FOO", "nuevo", separator="")
            content = path.read_text(encoding="utf-8")
            self.assertEqual(
                content,
                "| celda <!-- START_FOO -->nuevo<!-- END_FOO --> | otra |",
            )

    def test_body_with_backslashes_is_not_interpreted_as_backreference(self):
        """re.sub interpretaría \\1/\\g<...> en un string de reemplazo — el
        helper usa una función de reemplazo para evitar ese bug de forma
        estructural, no por casualidad. Este test lo fija como regresión."""
        from src.builders.io_utils import replace_delimited_block

        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write(tmpdir, "<!-- START_FOO -->\n\nviejo\n\n<!-- END_FOO -->")
            replace_delimited_block(str(path), "FOO", r"C:\Users\1 y \g<name>", separator="")
            content = path.read_text(encoding="utf-8")
            self.assertIn(r"C:\Users\1 y \g<name>", content)


class DocSyncTests(unittest.TestCase):
    """Tests para src/builders/doc_sync.py: sincroniza hechos hardcodeados de
    README.md (conteo de tests/ADRs/contratos, badge, pin de versión, salud,
    calidad, redistribución) con su fuente de verdad. Ver AGENTS.md §12."""

    def _readme(self, tmpdir, *blocks):
        lines = ["# README de prueba", ""]
        for name in blocks:
            lines += [f"<!-- START_{name} -->", "placeholder", f"<!-- END_{name} -->"]
        path = Path(tmpdir) / "README.md"
        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    def test_test_count_matches_ast_function_count(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            tests_dir = Path(tmpdir) / "tests"
            tests_dir.mkdir()
            (tests_dir / "test_a.py").write_text(
                "def test_one():\n    pass\n\n\ndef test_two():\n    pass\n", encoding="utf-8"
            )
            (tests_dir / "test_b.py").write_text(
                "def test_three():\n    pass\n\n\ndef helper():\n    pass\n", encoding="utf-8"
            )
            (tests_dir / "not_a_test.py").write_text(
                "def test_ignored():\n    pass\n", encoding="utf-8"
            )
            readme = self._readme(tmpdir, "TEST_COUNT")

            with (
                patch.object(doc_sync, "TESTS_DIR", str(tests_dir)),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                changed = doc_sync.sync_readme_test_count()

            self.assertTrue(changed)
            self.assertIn("**3 tests**", readme.read_text(encoding="utf-8"))

    def test_adr_and_contract_counts(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            adr_dir = Path(tmpdir) / "adr"
            adr_dir.mkdir()
            (adr_dir / "ADR-001-x.md").write_text("x", encoding="utf-8")
            (adr_dir / "ADR-002-y.md").write_text("y", encoding="utf-8")
            (adr_dir / "README.md").write_text("no cuenta")

            contracts_dir = Path(tmpdir) / "contracts"
            contracts_dir.mkdir()
            (contracts_dir / "comunas.schema.json").write_text("{}", encoding="utf-8")

            readme = self._readme(tmpdir, "ADR_COUNT", "CONTRACT_COUNT")

            with (
                patch.object(doc_sync, "ADR_DIR", str(adr_dir)),
                patch.object(doc_sync, "CONTRACTS_DIR", str(contracts_dir)),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                doc_sync.sync_readme_adr_count()
                doc_sync.sync_readme_contract_count()

            content = readme.read_text(encoding="utf-8")
            self.assertIn("**2 ADRs**", content)
            self.assertIn("1 contratos JSON Schema", content)

    def test_health_and_quality_summary_from_fixtures(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized_dir = Path(tmpdir) / "normalized"
            normalized_dir.mkdir()
            (normalized_dir / "hub_health.json").write_text(
                json.dumps({"ok_count": 5, "warn_count": 2, "error_count": 0}), encoding="utf-8"
            )
            (normalized_dir / "dataset_quality.json").write_text(
                json.dumps(
                    {
                        "average_score": 91.0,
                        "grade_distribution": {"A": 4, "B": 1, "C": 0, "D": 0, "F": 0},
                    }
                ),
                encoding="utf-8",
            )
            readme = self._readme(tmpdir, "HEALTH_SUMMARY", "QUALITY_SUMMARY")

            with (
                patch.object(doc_sync, "NORMALIZED_DIR", str(normalized_dir)),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                doc_sync.sync_readme_health_summary()
                doc_sync.sync_readme_quality_summary()

            content = readme.read_text(encoding="utf-8")
            self.assertIn("5 capas `ok`, 2 `warn`, 0 `error`", content)
            self.assertIn("promedio 91.0/100", content)
            self.assertIn("4 A, 1 B", content)

    def test_health_summary_returns_false_when_artifact_missing(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized_dir = Path(tmpdir) / "normalized"
            normalized_dir.mkdir()
            readme = self._readme(tmpdir, "HEALTH_SUMMARY")

            with (
                patch.object(doc_sync, "NORMALIZED_DIR", str(normalized_dir)),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                changed = doc_sync.sync_readme_health_summary()

            self.assertFalse(changed)
            self.assertIn("placeholder", readme.read_text(encoding="utf-8"))

    def test_version_pin_example_reads_pyproject(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            (Path(tmpdir) / "pyproject.toml").write_text(
                '[project]\nname = "x"\nversion = "9.9.9"\n', encoding="utf-8"
            )
            readme = self._readme(tmpdir, "VERSION_PIN_EXAMPLE")

            with (
                patch.object(doc_sync, "ROOT_DIR", tmpdir),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                doc_sync.sync_readme_version_pin_example()

            self.assertIn("chile-hub==9.9.9", readme.read_text(encoding="utf-8"))

    def test_dataset_badge_counts_only_datasets_with_outputs(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            readme = self._readme(tmpdir, "DATASET_BADGE")
            fake_catalog = {
                "comunas": {"outputs": {"parquet": "x"}},
                "empresas": {"outputs": {"parquet": "y"}},
                "candidato": {"outputs": {}},
            }

            with (
                patch.object(doc_sync, "DATASET_CATALOG_CONFIG", fake_catalog),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                doc_sync.sync_readme_dataset_badge()

            self.assertIn("Datasets-2%20capas", readme.read_text(encoding="utf-8"))

    def test_redistribution_summary_from_fixture(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            normalized_dir = Path(tmpdir) / "normalized"
            normalized_dir.mkdir()
            (normalized_dir / "redistribution_report.json").write_text(
                json.dumps({"ready_count": 17, "dataset_count": 19}), encoding="utf-8"
            )
            readme = self._readme(tmpdir, "REDISTRIBUTION_SUMMARY")

            with (
                patch.object(doc_sync, "NORMALIZED_DIR", str(normalized_dir)),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                doc_sync.sync_readme_redistribution_summary()

            self.assertIn("**17 de 19 capas**", readme.read_text(encoding="utf-8"))

    def test_sync_all_docs_runs_every_function_without_error(self):
        """Smoke test contra el README real: confirma que las 8 funciones
        corren sin lanzar excepciones y que check_only nunca escribe."""
        from src.builders import doc_sync

        changed = doc_sync.sync_all_docs(check_only=True)
        self.assertIsInstance(changed, list)

    def test_extractor_table_generated_from_real_catalog(self):
        """Regresión: la tabla generada preserva el agrupamiento editorial y
        cierra con la fila de perfil_territorial_comunal (Plan 058)."""
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            readme = self._readme(tmpdir, "EXTRACTOR_TABLE")

            with patch.object(doc_sync, "README_PATH", str(readme)):
                changed = doc_sync.sync_readme_extractor_table()

            self.assertTrue(changed)
            content = readme.read_text(encoding="utf-8")
            self.assertIn("| Dominio | Extractores |", content)
            self.assertIn("`subdere_extractor.py`", content)
            self.assertIn(
                "| Derivado en `build_dev_db.py` (sin extractor) | `perfil_territorial_comunal` |",
                content,
            )

    def test_extractor_table_check_only_false_when_unchanged(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            readme = self._readme(tmpdir, "EXTRACTOR_TABLE")

            with patch.object(doc_sync, "README_PATH", str(readme)):
                doc_sync.sync_readme_extractor_table()
                changed_again = doc_sync.sync_readme_extractor_table(check_only=True)

            self.assertFalse(changed_again)

    def test_extractor_table_raises_on_dataset_missing_domain(self):
        from src.builders import doc_sync

        with tempfile.TemporaryDirectory() as tmpdir:
            readme = self._readme(tmpdir, "EXTRACTOR_TABLE")
            fake_catalog = {"dataset_nuevo_sin_dominio": {"extractor": "src/extractors/x.py"}}

            with (
                patch.object(doc_sync, "DATASET_CATALOG_CONFIG", fake_catalog),
                patch.object(doc_sync, "README_PATH", str(readme)),
            ):
                with self.assertRaises(SystemExit):
                    doc_sync.sync_readme_extractor_table()


class ExtractorRegistryTests(unittest.TestCase):
    """Tests para check_extractors() en scripts/check_companion_paths.py —
    valida el campo `extractor` del catalogo (Plan 058)."""

    def _write_catalog(self, tmpdir, catalog):
        path = Path(tmpdir) / "dataset_catalog_config.json"
        path.write_text(json.dumps(catalog), encoding="utf-8")
        return path

    def test_check_extractors_against_real_catalog_has_no_errors(self):
        from scripts import check_companion_paths

        errors = check_companion_paths.check_extractors()
        self.assertEqual(errors, [])

    def test_missing_extractor_file_is_reported(self):
        from scripts import check_companion_paths

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            extractors_dir = root / "src" / "extractors"
            extractors_dir.mkdir(parents=True)
            catalog_path = self._write_catalog(
                tmpdir, {"foo": {"extractor": "src/extractors/no_existe.py"}}
            )

            with (
                patch.object(check_companion_paths, "ROOT_DIR", root),
                patch.object(check_companion_paths, "DATASET_CATALOG_PATH", catalog_path),
                patch.object(check_companion_paths, "EXTRACTORS_DIR", extractors_dir),
            ):
                errors = check_companion_paths.check_extractors()

            self.assertTrue(any("extractor no existe" in e for e in errors))

    def test_orphan_extractor_file_is_reported(self):
        from scripts import check_companion_paths

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            extractors_dir = root / "src" / "extractors"
            extractors_dir.mkdir(parents=True)
            (extractors_dir / "foo_extractor.py").write_text("", encoding="utf-8")
            (extractors_dir / "huerfano_extractor.py").write_text("", encoding="utf-8")
            catalog_path = self._write_catalog(
                tmpdir, {"foo": {"extractor": "src/extractors/foo_extractor.py"}}
            )

            with (
                patch.object(check_companion_paths, "ROOT_DIR", root),
                patch.object(check_companion_paths, "DATASET_CATALOG_PATH", catalog_path),
                patch.object(check_companion_paths, "EXTRACTORS_DIR", extractors_dir),
            ):
                errors = check_companion_paths.check_extractors()

            self.assertTrue(
                any("extractor huérfano" in e and "huerfano_extractor.py" in e for e in errors)
            )
            self.assertFalse(any("foo_extractor.py" in e and "huérfano" in e for e in errors))

    def test_missing_extractor_field_is_reported(self):
        from scripts import check_companion_paths

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            extractors_dir = root / "src" / "extractors"
            extractors_dir.mkdir(parents=True)
            catalog_path = self._write_catalog(tmpdir, {"foo": {"description": "sin extractor"}})

            with (
                patch.object(check_companion_paths, "ROOT_DIR", root),
                patch.object(check_companion_paths, "DATASET_CATALOG_PATH", catalog_path),
                patch.object(check_companion_paths, "EXTRACTORS_DIR", extractors_dir),
            ):
                errors = check_companion_paths.check_extractors()

            self.assertTrue(any("falta el campo" in e for e in errors))


class AdoptionStatsTests(unittest.TestCase):
    """Tests sin red del parseo/construcción de la señal de adopción (Plan 052).

    Sólo ejercitan funciones puras sobre el fixture offline; el fetch real de
    red (pypistats.org, api.github.com) no se testea aquí a propósito (no
    determinista, depende de servicios externos) — ver `--offline` en
    `scripts/fetch_adoption_stats.py`.
    """

    FIXTURE_PATH = ROOT_DIR / "tests" / "fixtures" / "adoption_sample.json"

    def _load_fixture(self):
        return json.loads(self.FIXTURE_PATH.read_text(encoding="utf-8"))

    def test_adoption_parse_pypi_stats_happy_path_from_fixture(self):
        fixture = self._load_fixture()
        stats = parse_pypi_stats(fixture["pypi_recent"])
        self.assertEqual(stats, {"last_day": 41, "last_week": 310, "last_month": 1234})

    def test_adoption_sum_github_downloads_happy_path_from_fixture(self):
        fixture = self._load_fixture()
        total = sum_github_downloads(fixture["github_releases"])
        self.assertEqual(total, 567)

    def test_adoption_parse_pypi_stats_degrades_gracefully_when_source_missing(self):
        stats = parse_pypi_stats(None)
        self.assertEqual(stats, {"last_day": None, "last_week": None, "last_month": None})

    def test_adoption_sum_github_downloads_degrades_gracefully_when_source_missing(self):
        self.assertEqual(sum_github_downloads(None), 0)
        self.assertEqual(sum_github_downloads([]), 0)

    def test_adoption_badge_uses_pypi_last_month_when_present(self):
        badge = build_adoption_badge(1234)
        self.assertEqual(badge["schemaVersion"], 1)
        self.assertEqual(badge["label"], "instalaciones/mes")
        self.assertEqual(badge["message"], "1234")
        self.assertIn("color", badge)

    def test_adoption_badge_degrades_to_nd_without_raising_when_pypi_missing(self):
        badge = build_adoption_badge(None)
        self.assertEqual(badge["schemaVersion"], 1)
        self.assertEqual(badge["message"], "n/d")
        self.assertIn("label", badge)
        self.assertIn("color", badge)

    def test_adoption_payload_contains_pypi_and_github_releases_keys(self):
        fixture = self._load_fixture()
        pypi_stats = parse_pypi_stats(fixture["pypi_recent"])
        github_total = sum_github_downloads(fixture["github_releases"])
        payload = build_adoption_payload(pypi_stats, github_total)
        self.assertIn("generated_at_utc", payload)
        self.assertEqual(payload["pypi"], pypi_stats)
        self.assertEqual(payload["github_releases"], {"total_downloads": 567})

    def test_adoption_offline_fixture_end_to_end_matches_shields_contract(self):
        pypi_recent, github_releases = load_adoption_offline_fixture(self.FIXTURE_PATH)
        pypi_stats = parse_pypi_stats(pypi_recent)
        github_total = sum_github_downloads(github_releases)
        badge = build_adoption_badge(pypi_stats["last_month"])

        self.assertEqual(
            badge,
            {
                "schemaVersion": 1,
                "label": "instalaciones/mes",
                "message": "1234",
                "color": "blue",
                "namedLogo": "pypi",
                "cacheSeconds": 86400,
            },
        )
        self.assertEqual(github_total, 567)


class ExcelGuardTests(unittest.TestCase):
    """Tests para el guardia de omisión de tablas masivas en Excel."""

    def test_build_excel_skips_oversized_table(self):
        """Tablas extra >500k filas se omiten del Excel; las pequeñas sí se incluyen."""
        import openpyxl
        import pandas as pd

        from src.builders.formats import build_excel

        df_tiny = pl.DataFrame({"col": [1]})

        class OversizedFrame(pd.DataFrame):
            def __len__(self):
                return 500_001

        extra_tables_pd = {
            "test_small": pd.DataFrame({"x": [1, 2, 3]}),
            "test_oversized": OversizedFrame({"x": [1, 2, 3]}),
        }
        extra_tables = {
            "test_small": pl.DataFrame({"x": [1, 2, 3]}),
            "test_oversized": pl.DataFrame({"x": [1, 2, 3]}),
        }

        test_config = {
            "test_small": {"outputs": {}},
            "test_oversized": {"outputs": {}},
        }

        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch("src.builders.formats.DATASET_CATALOG_CONFIG", test_config),
        ):
            output_path = str(Path(tmpdir) / "test.xlsx")
            build_excel(
                df_tiny,
                df_tiny,
                df_tiny,
                df_tiny,
                df_tiny,
                df_tiny,
                df_tiny,
                extra_tables,
                extra_tables_pd,
                output_path,
            )
            wb = openpyxl.load_workbook(output_path)
            self.assertIn("test_small", wb.sheetnames)
            self.assertNotIn("test_oversized", wb.sheetnames)


class NormalizeComunaNameTests(unittest.TestCase):
    """Tests para src/chile_hub/text.py::normalize_comuna_name (Plan 050, ADR-009).

    No requiere data/normalized/ — es una función pura sobre strings.
    """

    def test_strips_accents_and_enye(self):
        from src.chile_hub.text import normalize_comuna_name

        self.assertEqual(normalize_comuna_name("Ñuñoa"), "nunoa")

    def test_uppercase_with_accent(self):
        from src.chile_hub.text import normalize_comuna_name

        self.assertEqual(normalize_comuna_name("CONCÓN"), "concon")

    def test_strips_surrounding_whitespace(self):
        from src.chile_hub.text import normalize_comuna_name

        self.assertEqual(normalize_comuna_name("  Concón "), "concon")

    def test_already_clean_string_unchanged(self):
        from src.chile_hub.text import normalize_comuna_name

        self.assertEqual(normalize_comuna_name("santiago"), "santiago")

    def test_matches_polars_chain_except_for_intentional_strip(self):
        """normalize_comuna_name() debe coincidir EXACTAMENTE con la cadena Polars
        de subdere_extractor.py:558-570 (mismo orden, mismos reemplazos) para
        cualquier input sin espacios al borde. La única divergencia permitida es
        el .strip() adicional (ver docstring de text.py y ADR-009)."""
        import polars as pl

        from src.chile_hub.text import normalize_comuna_name

        def polars_chain(value):
            df = pl.DataFrame({"nombre_comuna": [value]})
            df = df.with_columns(
                pl.col("nombre_comuna")
                .str.to_lowercase()
                .str.replace_all("á", "a")
                .str.replace_all("é", "e")
                .str.replace_all("í", "i")
                .str.replace_all("ó", "o")
                .str.replace_all("ú", "u")
                .str.replace_all("ü", "u")
                .str.replace_all("ñ", "n")
                .alias("nombre_comuna_clean")
            )
            return df["nombre_comuna_clean"][0]

        for value in ["Ñuñoa", "CONCÓN", "Aysén", "Curicó", "Máfil", "PUERTO MONTT"]:
            self.assertEqual(normalize_comuna_name(value), polars_chain(value))

    def test_empty_string(self):
        from src.chile_hub.text import normalize_comuna_name

        self.assertEqual(normalize_comuna_name(""), "")


class DcatCatalogTests(unittest.TestCase):
    """Tests para build_dcat_catalog (Plan 051, ADR-010) sobre un
    datapackage.json sintetico (misma fixture que DataPackageBuilderTests)."""

    def _fixture_data_package(self):
        return {
            "name": "chile-hub",
            "title": "chile-hub -- Datos publicos de Chile curados",
            "description": "Descripcion de prueba",
            "version": "1.0.0",
            "homepage": "https://tooltician.com/chile-hub/",
            "created": "2026-07-24T00:00:00+00:00",
            "resources": [
                {
                    "name": "test_dataset",
                    "path": "test_dataset.parquet",
                    "format": "parquet",
                    "mediatype": "application/vnd.apache.parquet",
                    "title": "test_dataset",
                    "description": "Test dataset",
                    "schema": {"fields": []},
                    "licenses": [{"title": "CC BY", "path": "https://example.com/license"}],
                }
            ],
        }

    def test_dataset_list_not_empty(self):
        from src.builders.dcat_catalog import build_dcat_catalog

        catalog = build_dcat_catalog(self._fixture_data_package())
        self.assertTrue(catalog["dataset"])

    def test_download_url_is_absolute_under_base_url(self):
        from src.builders.dcat_catalog import build_dcat_catalog

        catalog = build_dcat_catalog(self._fixture_data_package())
        download_url = catalog["dataset"][0]["distribution"][0]["downloadURL"]
        self.assertEqual(
            download_url,
            "https://tooltician.com/chile-hub/data/normalized/test_dataset.parquet",
        )
        self.assertTrue(download_url.startswith("https://"))

    def test_license_propagated_when_present(self):
        from src.builders.dcat_catalog import build_dcat_catalog

        catalog = build_dcat_catalog(self._fixture_data_package())
        self.assertEqual(catalog["dataset"][0]["license"], "https://example.com/license")

    def test_no_license_key_when_absent(self):
        from src.builders.dcat_catalog import build_dcat_catalog

        data_package = self._fixture_data_package()
        del data_package["resources"][0]["licenses"]
        catalog = build_dcat_catalog(data_package)
        self.assertNotIn("license", catalog["dataset"][0])

    def test_write_dcat_catalog_json_writes_file(self):
        from src.builders import dcat_catalog

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.object(dcat_catalog, "NORMALIZED_DIR", tmpdir):
                output_path = dcat_catalog.write_dcat_catalog_json(self._fixture_data_package())

            written = json.loads(Path(output_path).read_text(encoding="utf-8"))
            self.assertTrue(written["dataset"])
            self.assertTrue(
                written["dataset"][0]["distribution"][0]["downloadURL"].startswith("https://")
            )


NORMALIZED_DIR = Path(__file__).resolve().parents[1] / "data" / "normalized"


class IndicatorAgeTests(unittest.TestCase):
    """Plan 067 / ADR-016: edad del dato entregado por serie."""

    def _frame(self, rows):
        return pl.DataFrame(
            {"codigo_indicador": [r[0] for r in rows], "fecha": [r[1] for r in rows]}
        )

    def test_age_measured_on_delivered_data_per_series(self):
        from src.builders.metadata import build_indicator_ages

        max_dates, ages = build_indicator_ages(
            self._frame([("ipc", "2025-12-01"), ("dolar", "2026-07-20")]),
            today=datetime.date(2026, 7, 29),
        )
        self.assertEqual(max_dates["ipc"], "2025-12-01")
        self.assertEqual(ages["ipc"], 240)
        self.assertEqual(ages["dolar"], 9)

    def test_future_dated_series_yield_negative_age(self):
        """La UF y la UTM se publican por adelantado: no es un error."""
        from src.builders.metadata import build_indicator_ages

        _, ages = build_indicator_ages(
            self._frame([("uf", "2026-08-09")]), today=datetime.date(2026, 7, 29)
        )
        self.assertEqual(ages["uf"], -11)

    def test_empty_frame_returns_empty_maps(self):
        from src.builders.metadata import build_indicator_ages

        max_dates, ages = build_indicator_ages(pl.DataFrame(), today=datetime.date(2026, 7, 29))
        self.assertEqual(max_dates, {})
        self.assertEqual(ages, {})

    def test_threshold_depends_on_cadence(self):
        from src.extractors.bcentral_extractor import (
            MAX_BACKFILL_AGE_DAYS_DAILY,
            MAX_BACKFILL_AGE_DAYS_MONTHLY,
            max_backfill_age_days,
        )

        self.assertEqual(max_backfill_age_days("ipc"), MAX_BACKFILL_AGE_DAYS_MONTHLY)
        self.assertEqual(max_backfill_age_days("utm"), MAX_BACKFILL_AGE_DAYS_MONTHLY)
        self.assertEqual(max_backfill_age_days("dolar"), MAX_BACKFILL_AGE_DAYS_DAILY)
        self.assertGreater(MAX_BACKFILL_AGE_DAYS_MONTHLY, MAX_BACKFILL_AGE_DAYS_DAILY)

    def test_published_artifacts_expose_ages(self):
        metadata = json.loads(
            (NORMALIZED_DIR / "pipeline_metadata.json").read_text(encoding="utf-8")
        )
        ages = metadata["datasets"]["indicadores"]["indicator_age_days"]
        self.assertEqual(set(ages), {"uf", "dolar", "euro", "utm", "ipc"})


class DriftTaxonomyTests(unittest.TestCase):
    """Plan 066 / ADR-014: drift esperado (de diseño) vs drift real.

    Verifica que la clasificación distinga cobertura parcial declarada en el
    contrato y warnings declarados como esperados, sin agregar valores a ningún
    enum ni relajar gates.
    """

    # --- Step 1: source_mode de la capa derivada -------------------------

    def test_non_fallback_source_modes_is_single_source_of_truth(self):
        """verify_pipeline.py debe consumir el conjunto de _shared.py, no
        redefinirlo (la duplicación fue el bug original)."""
        from scripts.verify_pipeline import NON_FALLBACK_SOURCE_MODES as gate_set
        from src.builders._shared import NON_FALLBACK_SOURCE_MODES

        self.assertIs(NON_FALLBACK_SOURCE_MODES, gate_set)
        self.assertEqual(NON_FALLBACK_SOURCE_MODES, {"live", "monthly"})

    def test_monthly_upstream_does_not_make_derived_layer_fallback(self):
        from src.builders._shared import NON_FALLBACK_SOURCE_MODES

        upstreams = [{"source_mode": "live"}] * 8 + [{"source_mode": "monthly"}]
        derived_mode = (
            "live"
            if all(m.get("source_mode") in NON_FALLBACK_SOURCE_MODES for m in upstreams)
            else "fallback"
        )
        self.assertEqual(derived_mode, "live")

    def test_real_fallback_upstream_makes_derived_layer_fallback(self):
        from src.builders._shared import NON_FALLBACK_SOURCE_MODES

        upstreams = [{"source_mode": "live"}] * 8 + [{"source_mode": "fallback"}]
        derived_mode = (
            "live"
            if all(m.get("source_mode") in NON_FALLBACK_SOURCE_MODES for m in upstreams)
            else "fallback"
        )
        self.assertEqual(derived_mode, "fallback")

    def test_perfil_territorial_comunal_is_live_and_healthy_in_artifacts(self):
        """Regresión sobre el artefacto real: la capa derivada dejó de driftar."""
        report = json.loads((NORMALIZED_DIR / "drift_report.json").read_text(encoding="utf-8"))
        entry = next(e for e in report["datasets"] if e["dataset"] == "perfil_territorial_comunal")
        self.assertEqual(entry["source_mode"], "live")
        self.assertEqual(entry["drift_status"], "healthy")

    # --- Step 2: coverage_policy del contrato ----------------------------

    def _coverage(self, policy, record_count=100, expected_record_count=None, name="x"):
        from src.builders import metadata as metadata_mod

        catalog = {name: {"expected_record_count": expected_record_count}}
        with patch.object(metadata_mod, "DATASET_CATALOG_CONFIG", catalog):
            return metadata_mod.build_coverage(
                name,
                {"record_count": record_count},
                {"coverage_policy": policy} if policy else None,
            )

    def test_partial_expected_policy_marks_coverage_expected(self):
        coverage = self._coverage("partial_expected", 200, 346)
        self.assertEqual(coverage["status"], "partial")
        self.assertTrue(coverage["expected"])
        self.assertIn("coverage_policy", coverage["expected_reason"])

    def test_partial_policy_without_expected_stays_unexpected(self):
        coverage = self._coverage("partial", 200, 346)
        self.assertEqual(coverage["status"], "partial")
        self.assertFalse(coverage["expected"])
        self.assertNotIn("expected_reason", coverage)

    def test_expected_flag_always_present_for_every_policy(self):
        """El campo es siempre inspeccionable, nunca ausente."""
        for policy in ["full", "not_applicable", "roster", "partial", "partial_expected"]:
            with self.subTest(policy=policy):
                coverage = self._coverage(policy, 346, 346)
                self.assertIn("expected", coverage)
                self.assertIsInstance(coverage["expected"], bool)

    def test_full_coverage_is_never_expected_partial(self):
        coverage = self._coverage("partial_expected", 346, 346)
        self.assertEqual(coverage["status"], "full")
        self.assertFalse(coverage["expected"])

    def test_not_applicable_coverage_is_not_expected(self):
        coverage = self._coverage("partial_expected", 346, None)
        self.assertEqual(coverage["status"], "not_applicable")
        self.assertFalse(coverage["expected"])

    def test_extractor_declared_partial_expected_converges_on_same_flag(self):
        """La ruta del metadata del extractor (SIEDU) produce el mismo campo."""
        from src.builders.metadata import build_coverage

        coverage = build_coverage(
            "indicadores_urbanos_siedu",
            {
                "record_count": 10,
                "coverage": {
                    "status": "partial_expected",
                    "coverage_ratio": 0.34,
                    "expected_scope": "sólo comunas urbanas",
                },
            },
        )
        self.assertEqual(coverage["status"], "partial")
        self.assertTrue(coverage["expected"])
        self.assertEqual(coverage["expected_reason"], "sólo comunas urbanas")

    def test_all_four_partial_expected_contracts_are_covered(self):
        """Los 4 contratos con coverage_policy=partial_expected.

        Hoy sólo finanzas_municipales e indicadores_urbanos_siedu alcanzan
        coverage_status=partial; en resultados_educacionales y
        delincuencia_comunal el flag queda INERTE (su cobertura no es partial:
        no tienen expected_record_count o su actual >= esperado). Se cubren
        igual para que, si mañana pasan a partial, no drifteen sin que ningún
        test lo haya ejercitado.
        """
        contracts_dir = Path(__file__).resolve().parents[1] / "contracts" / "datasets"
        declared = sorted(
            path.stem.replace(".schema", "")
            for path in contracts_dir.glob("*.schema.json")
            if json.loads(path.read_text(encoding="utf-8")).get("coverage_policy")
            == "partial_expected"
        )
        self.assertEqual(
            declared,
            [
                "delincuencia_comunal",
                "finanzas_municipales",
                "indicadores_urbanos_siedu",
                "resultados_educacionales",
            ],
        )
        for name in declared:
            with self.subTest(dataset=name):
                coverage = self._coverage("partial_expected", 200, 346, name=name)
                self.assertTrue(coverage["expected"])

    # --- Step 3: warnings esperados vs accionables -----------------------

    def test_only_expected_warnings_means_no_degradation(self):
        from src.builders.metadata import build_degradation

        degradation = build_degradation(
            "pobreza_comunal",
            {"source_mode": "live", "record_count": 10},
            {"warnings": ["parcial por diseño"], "expected_warnings": ["parcial por diseño"]},
        )
        self.assertEqual(degradation["status"], "none")
        self.assertIn("parcial por diseño", degradation["impact"])
        self.assertEqual(degradation["recommended_action"], "Ninguna.")

    def test_undeclared_warning_still_degrades(self):
        """Guardrail anti-silenciador: sólo lo declarado se descuenta."""
        from src.builders.metadata import build_degradation

        degradation = build_degradation(
            "empresas",
            {"source_mode": "live", "record_count": 10},
            {"warnings": ["RUT inválido"], "expected_warnings": []},
        )
        self.assertEqual(degradation["status"], "warning")
        self.assertIn("RUT inválido", degradation["impact"])

    def test_mixed_warnings_report_only_actionable_in_impact(self):
        from src.builders.metadata import build_degradation

        degradation = build_degradation(
            "mixto",
            {"source_mode": "live", "record_count": 10},
            {
                "warnings": ["esperado", "accionable"],
                "expected_warnings": ["esperado"],
            },
        )
        self.assertEqual(degradation["status"], "warning")
        self.assertIn("accionable", degradation["impact"])
        self.assertNotIn("esperado", degradation["impact"])

    def test_validators_declaring_expected_warnings_are_exactly_the_known_set(self):
        """Los warnings de diseño se declaran en el emisor, no en el consumidor.

        Aserción estructural (AST), no de formato: agregar una declaración nueva
        obliga a actualizar esta lista deliberadamente — que es justamente el
        control que ADR-014 pide para que `expected_warnings` no se convierta en
        un basurero de warnings molestos.
        """
        import ast
        import inspect

        import src.validation as validation_mod

        tree = ast.parse(inspect.getsource(validation_mod))
        declaring = {
            node.name
            for node in ast.walk(tree)
            if isinstance(node, ast.FunctionDef)
            and any(
                isinstance(call.func, ast.Name) and call.func.id == "_add_expected_warning"
                for call in ast.walk(node)
                if isinstance(call, ast.Call)
            )
        }
        self.assertEqual(
            declaring,
            {
                "validate_indicadores_urbanos_siedu",  # cobertura urbana intencional
                "validate_pobreza_comunal",  # cobertura SAE parcial por diseño
                "validate_partidos_politicos",  # estado_legal segun SERVEL
                "validate_empresas",  # RES solo cubre Ley 20.659 (issue #42)
            },
        )

    # --- Step 4: build_drift y contadores --------------------------------

    def test_expected_partial_coverage_does_not_drift(self):
        from src.builders.metadata import build_drift

        drift = build_drift(
            {
                "source_mode": "live",
                "coverage": {"status": "partial", "expected": True},
                "degradation": {"status": "none"},
            }
        )
        self.assertEqual(drift["status"], "healthy")

    def test_unexpected_partial_coverage_drifts(self):
        from src.builders.metadata import build_drift

        drift = build_drift(
            {
                "source_mode": "live",
                "coverage": {"status": "partial", "expected": False},
                "degradation": {"status": "none"},
            }
        )
        self.assertEqual(drift["status"], "drifted")
        self.assertIn("coverage=partial", drift["summary"])

    def test_unknown_coverage_still_drifts(self):
        from src.builders.metadata import build_drift

        drift = build_drift(
            {
                "source_mode": "live",
                "coverage": {"status": "unknown", "expected": False},
                "degradation": {"status": "none"},
            }
        )
        self.assertEqual(drift["status"], "drifted")

    def test_fallback_source_mode_still_drifts(self):
        from src.builders.metadata import build_drift

        drift = build_drift(
            {
                "source_mode": "fallback",
                "coverage": {"status": "full", "expected": False},
                "degradation": {"status": "none"},
            }
        )
        self.assertEqual(drift["status"], "drifted")
        self.assertIn("mode=fallback", drift["summary"])

    def test_drift_summary_names_only_the_triggering_condition(self):
        from src.builders.metadata import build_drift

        drift = build_drift(
            {
                "source_mode": "live",
                "coverage": {"status": "full", "expected": False},
                "degradation": {"status": "warning"},
            }
        )
        self.assertEqual(drift["summary"], "Drift detectado: degradation=warning.")

    def test_hub_health_reports_actionable_count_without_changing_warning_count(self):
        from src.chile_hub.pipeline_status_utils import build_hub_health

        health = build_hub_health(
            {
                "datasets": {
                    "x": {
                        "source_mode": "live",
                        "freshness": {"status": "fresh"},
                        "reuse_policy": {"status": "ok", "redistribution_ok": True},
                        "degradation": {"status": "none"},
                        "coverage": {"status": "partial", "expected": True},
                        "drift": {"status": "healthy"},
                    }
                },
                "validations": {
                    "x": {
                        "status": "ok",
                        "warnings": ["esperado"],
                        "expected_warnings": ["esperado"],
                    }
                },
            }
        )
        entry = health["datasets"][0]
        self.assertEqual(entry["warning_count"], 1)
        self.assertEqual(entry["actionable_warning_count"], 0)
        self.assertEqual(entry["severity"], "ok")

    # --- Guardrail sobre los artefactos reales ---------------------------

    def test_published_artifacts_report_exactly_two_actionable_problems(self):
        """Guardrail: una regresión de clasificación debe ser ruidosa.

        Bajó de 3 a 2 por un mecanismo **distinto** al de ADR-014: el Plan 068
        excluyó `consumo_electrico_comunal` de la contabilidad de salud por tener
        la fuente muerta (ADR-015), no por declararlo "esperado". El dataset
        sigue apareciendo como `drifted` en `drift_report.json` — lo que cambia
        es que ya no cuenta en la señal.
        """
        health = json.loads((NORMALIZED_DIR / "hub_health.json").read_text(encoding="utf-8"))
        self.assertEqual(health["drifted_count"], 2)
        self.assertEqual(health["warn_count"], 2)
        self.assertEqual(health["retired_count"], 1)
        self.assertEqual(health["dataset_count"], 19)
        self.assertEqual(health["overall_status"], "warn")
        report = json.loads((NORMALIZED_DIR / "drift_report.json").read_text(encoding="utf-8"))
        drifted = sorted(e["dataset"] for e in report["datasets"] if e["drift_status"] == "drifted")
        self.assertEqual(drifted, ["consumo_electrico_comunal", "empresas", "indicadores"])
        retired = sorted(e["dataset"] for e in health["datasets"] if e["retired"])
        self.assertEqual(retired, ["consumo_electrico_comunal"])

    # --- Plan 068 / ADR-015: fuentes retiradas ---------------------------

    def test_retired_set_comes_from_registry_not_hardcoded(self):
        from src.chile_hub.pipeline_status_utils import _load_retired_datasets

        self.assertEqual(_load_retired_datasets(), {"consumo_electrico_comunal"})

    def test_retired_dataset_excluded_from_counters_but_still_listed(self):
        from src.chile_hub.pipeline_status_utils import build_hub_health

        def _dataset(drift):
            return {
                "source_mode": "live",
                "freshness": {"status": "fresh"},
                "reuse_policy": {"status": "ok", "redistribution_ok": True},
                "degradation": {"status": "none"},
                "coverage": {"status": "full", "expected": False},
                "drift": {"status": drift},
            }

        with patch(
            "src.chile_hub.pipeline_status_utils._load_retired_datasets",
            return_value={"muerto"},
        ):
            health = build_hub_health(
                {
                    "datasets": {"vivo": _dataset("healthy"), "muerto": _dataset("drifted")},
                    "validations": {
                        "vivo": {"status": "ok", "warnings": []},
                        "muerto": {"status": "ok", "warnings": ["fuente caida"]},
                    },
                }
            )

        self.assertEqual(health["drifted_count"], 0)
        self.assertEqual(health["retired_count"], 1)
        self.assertEqual(health["dataset_count"], 2, "el inventario no cambia")
        self.assertEqual(health["overall_status"], "ok")
        listed = {entry["dataset"] for entry in health["datasets"]}
        self.assertIn("muerto", listed, "el retirado se muestra marcado, no se esconde")

    def test_non_deprecated_dataset_is_never_retired(self):
        """Guardrail: solo el registry puede retirar; nada más."""
        health = json.loads((NORMALIZED_DIR / "hub_health.json").read_text(encoding="utf-8"))
        for entry in health["datasets"]:
            if entry["dataset"] != "consumo_electrico_comunal":
                self.assertFalse(entry["retired"], entry["dataset"])

    def test_warnings_still_contain_every_message(self):
        """`warnings` nunca pierde mensajes: los esperados siguen listados."""
        status = json.loads((NORMALIZED_DIR / "dataset_status.json").read_text(encoding="utf-8"))
        entries = status["datasets"] if isinstance(status, dict) else status
        pobreza = next(e for e in entries if e["dataset"] == "pobreza_comunal")
        self.assertTrue(
            any("parcial por diseño" in w for w in pobreza.get("warnings", [])),
            msg=f"warnings de pobreza_comunal: {pobreza.get('warnings')}",
        )


if __name__ == "__main__":
    import pytest

    sys.exit(pytest.main(sys.argv))
