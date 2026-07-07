"""Extrae consumo eléctrico comunal desde el portal Energía Abierta de la CNE.

Fuente: Comisión Nacional de Energía (CNE) — Energía Abierta
URL: http://energiaabierta.cl/datasets-estadistica/consumo-electrico-anual-por-comuna-y-tipo-de-cliente/
Formato: Excel (descarga directa, sin API key)
"""

import datetime
import os
import sys
from pathlib import Path

import openpyxl
import polars as pl
import requests

UTC = datetime.timezone.utc

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

try:
    from src.extractors.base import (
        BaseExtractor,
        ensure_staging_directories,
        write_staging_metadata,
    )
except ModuleNotFoundError:
    from base import BaseExtractor, ensure_staging_directories, write_staging_metadata

try:
    from src.extractors.http_utils import fetch_with_retry
except ModuleNotFoundError:
    from http_utils import fetch_with_retry

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data"))
RAW_DIR = os.path.join(DATA_DIR, "raw")
STAGING_DIR = os.path.join(DATA_DIR, "staging")
STAGING_CSV_PATH = os.path.join(STAGING_DIR, "consumo_electrico_comunal.csv")
METADATA_PATH = os.path.join(STAGING_DIR, "consumo_electrico_comunal.metadata.json")

# ── URL de descarga ──────────────────────────────────────────────────────────
# Excel directo desde el portal Energía Abierta (sin API key requerida)
DOWNLOAD_URL = (
    "http://datos.energiaabierta.cl/dataviews/241686/"
    "consumo-electrico-anual-por-comuna-y-tipo-de-cliente/"
)

REUSE_POLICY = {
    "status": "open-attribution",
    "license": "CC BY",
    "license_url": "http://energiaabierta.cl/",
    "attribution_required": True,
    "redistribution_ok": True,
    "summary": (
        "Consumo eléctrico anual por comuna y tipo de cliente, publicado por la "
        "Comisión Nacional de Energía (CNE) en el portal Energía Abierta. "
        "Datos abiertos con atribución requerida."
    ),
}

# ── Fallback mínimo ─────────────────────────────────────────────────────────
FALLBACK_ROWS = [
    {
        "codigo_region": "13",
        "codigo_comuna": "13101",
        "nombre_comuna": "Santiago",
        "anio": 2023,
        "tipo_cliente": "Residencial",
        "consumo_kwh": 1250000000.0,
        "numero_clientes": 150000,
        "fuente": "CNE — Energía Abierta (fallback)",
        "url_fuente": DOWNLOAD_URL,
        "fecha_fuente": "",
    },
    {
        "codigo_region": "08",
        "codigo_comuna": "08101",
        "nombre_comuna": "Concepción",
        "anio": 2023,
        "tipo_cliente": "Residencial",
        "consumo_kwh": 350000000.0,
        "numero_clientes": 60000,
        "fuente": "CNE — Energía Abierta (fallback)",
        "url_fuente": DOWNLOAD_URL,
        "fecha_fuente": "",
    },
    {
        "codigo_region": "13",
        "codigo_comuna": "13101",
        "nombre_comuna": "Santiago",
        "anio": 2023,
        "tipo_cliente": "Comercial",
        "consumo_kwh": 2100000000.0,
        "numero_clientes": 45000,
        "fuente": "CNE — Energía Abierta (fallback)",
        "url_fuente": DOWNLOAD_URL,
        "fecha_fuente": "",
    },
]

REQUIRED_COLUMNS = [
    "codigo_region",
    "codigo_comuna",
    "nombre_comuna",
    "anio",
    "tipo_cliente",
    "consumo_kwh",
    "numero_clientes",
    "fuente",
    "url_fuente",
    "fecha_fuente",
]

# Mapeo de nombres de tipo de cliente a categorías normalizadas
TIPO_CLIENTE_MAP = {
    "residencial": "Residencial",
    "comercial": "Comercial",
    "industrial": "Industrial",
    "agrícola": "Agrícola",
    "agua potable y alcantarillado": "Agua Potable y Alcantarillado",
    "alumbrado público": "Alumbrado Público",
    "otros": "Otros",
    "varios": "Otros",
}


def _snapshot_path() -> Path:
    stamp = datetime.datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    return Path(RAW_DIR) / f"cne_consumo_electrico_comunal_{stamp}.xlsx"


def _download_excel() -> Path:
    """Descarga el Excel de consumo eléctrico y guarda snapshot crudo."""
    target = _snapshot_path()
    with fetch_with_retry(DOWNLOAD_URL, timeout=60) as response:
        response.raise_for_status()
        target.write_bytes(response.content)
    return target


def _parse_excel(path: Path) -> list[dict]:
    """Parsea el Excel de consumo eléctrico comunal.

    El formato esperado tiene columnas: región, comuna, año, tipo_cliente,
    consumo (kWh), número_clientes. Los nombres exactos varían; se intenta
    detectar por posición.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_raw = list(sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()

    rows = []
    for row in rows_raw:
        if not row or sum(1 for c in row if c is not None) < 3:
            continue
        try:
            _nombre_region = str(row[0]).strip() if row[0] else ""
            nombre_comuna = str(row[1]).strip() if row[1] else ""
            anio = int(row[2]) if row[2] is not None else None
            tipo_cliente_raw = str(row[3]).strip().lower() if row[3] else ""
            consumo_kwh = float(row[4]) if len(row) > 4 and row[4] is not None else None
            nclientes = int(row[5]) if len(row) > 5 and row[5] is not None else None
        except (ValueError, TypeError, IndexError):
            continue

        if not nombre_comuna or anio is None or consumo_kwh is None:
            continue

        tipo_cliente = TIPO_CLIENTE_MAP.get(tipo_cliente_raw, tipo_cliente_raw.title())

        rows.append(
            {
                "codigo_region": "",
                "codigo_comuna": "",
                "nombre_comuna": nombre_comuna,
                "anio": anio,
                "tipo_cliente": tipo_cliente,
                "consumo_kwh": consumo_kwh,
                "numero_clientes": nclientes,
                "fuente": "CNE — Energía Abierta",
                "url_fuente": DOWNLOAD_URL,
                "fecha_fuente": datetime.datetime.now(UTC).strftime("%Y-%m-%d"),
            }
        )
    return rows


def _enrich_with_cut(rows: list[dict]) -> list[dict]:
    """Enriquece las filas con codigo_comuna y codigo_region usando la DPA.

    Como el Excel de CNE puede traer nombres de comuna sin CUT, usamos
    el CSV de comunas del staging para hacer match por nombre.
    """
    comunas_csv = os.path.join(STAGING_DIR, "comunas.csv")
    if not os.path.exists(comunas_csv):
        return rows

    comunas_df = pl.read_csv(
        comunas_csv,
        columns=["codigo_region", "codigo_comuna", "nombre_comuna"],
        schema_overrides={"codigo_region": pl.String, "codigo_comuna": pl.String},
    )
    # Normalizar nombres para matching
    comunas_lookup = {}
    for row_data in comunas_df.iter_rows(named=True):
        key = row_data["nombre_comuna"].strip().lower()
        comunas_lookup[key] = (
            row_data["codigo_region"],
            row_data["codigo_comuna"],
            row_data["nombre_comuna"],
        )

    for row in rows:
        key = row["nombre_comuna"].strip().lower()
        if key in comunas_lookup:
            row["codigo_region"] = comunas_lookup[key][0]
            row["codigo_comuna"] = comunas_lookup[key][1]
            row["nombre_comuna"] = comunas_lookup[key][2]

    return rows


def fetch_data() -> tuple[list[dict], str, str, list[str]]:
    """Obtiene datos de consumo eléctrico comunal desde Energía Abierta.

    Retorna (rows, source_mode, source_url, notes).
    """
    ensure_staging_directories()
    notes = []

    try:
        path = _download_excel()
        rows = _parse_excel(path)
        rows = _enrich_with_cut(rows)
        matched = sum(1 for r in rows if r["codigo_comuna"])
        notes.append(
            f"live: {len(rows)} filas descargadas, {matched} con CUT "
            f"({round(matched / max(len(rows), 1) * 100, 1)}% match)"
        )
        return rows, "live", DOWNLOAD_URL, notes
    except (requests.RequestException, OSError, KeyError) as exc:
        snapshots = sorted(Path(RAW_DIR).glob("cne_consumo_electrico_comunal_*.xlsx"))
        if snapshots:
            rows = _parse_excel(snapshots[-1])
            rows = _enrich_with_cut(rows)
            notes.append(f"fallback: {len(rows)} filas desde snapshot {snapshots[-1].name} ({exc})")
            return rows, "fallback", DOWNLOAD_URL, notes
        notes.append(f"fallback: usando datos de muestra ({exc})")
        return FALLBACK_ROWS, "fallback", DOWNLOAD_URL, notes


def normalize_rows(rows: list[dict]) -> pl.DataFrame:
    """Convierte filas de consumo eléctrico a DataFrame canónico."""
    if not rows:
        return pl.DataFrame(schema={col: pl.String for col in REQUIRED_COLUMNS})

    df = pl.DataFrame(rows, strict=False)

    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            if col == "consumo_kwh":
                df = df.with_columns(pl.lit(0.0, dtype=pl.Float64).alias(col))
            elif col == "numero_clientes":
                df = df.with_columns(pl.lit(None, dtype=pl.Int64).alias(col))
            elif col == "anio":
                df = df.with_columns(pl.lit(2023, dtype=pl.Int64).alias(col))
            else:
                df = df.with_columns(pl.lit("", dtype=pl.String).alias(col))

    df = df.with_columns(
        pl.col("codigo_region").cast(pl.String).str.zfill(2),
        pl.col("codigo_comuna").cast(pl.String).str.zfill(5),
        pl.col("nombre_comuna").cast(pl.String),
        pl.col("anio").cast(pl.Int64),
        pl.col("tipo_cliente").cast(pl.String),
        pl.col("consumo_kwh").cast(pl.Float64),
        pl.col("numero_clientes").cast(pl.Int64),
        pl.col("fuente").cast(pl.String),
        pl.col("url_fuente").cast(pl.String),
        pl.col("fecha_fuente").cast(pl.String),
    )

    # Filtrar filas sin CUT (no se pueden cruzar)
    df = df.filter(pl.col("codigo_comuna") != "")

    return df.select(REQUIRED_COLUMNS)


def build_metadata(mode: str, source_url: str, notes: list[str], row_count: int) -> dict:
    return {
        "dataset": "consumo_electrico_comunal",
        "source_name": "CNE — Energía Abierta",
        "source_url": source_url,
        "source_mode": mode,
        "source_detail": "Consumo eléctrico anual por comuna y tipo de cliente",
        "refreshed_at_utc": datetime.datetime.now(UTC).isoformat(),
        "record_count": row_count,
        "fields": REQUIRED_COLUMNS,
        "notes": notes,
        "reuse_policy": REUSE_POLICY,
    }


def process_consumo_electrico() -> dict:
    """Ejecuta el flujo completo de extracción y staging."""
    rows, mode, source_url, notes = fetch_data()
    df = normalize_rows(rows)

    metadata = build_metadata(mode, source_url, notes, df.height)

    ensure_staging_directories()
    df.write_csv(STAGING_CSV_PATH)
    write_staging_metadata(METADATA_PATH, metadata)
    print(f"consumo_electrico_comunal: {df.height} filas escritas en staging (mode={mode})")

    return metadata


# ── Extractor ────────────────────────────────────────────────────────────────


class ConsumoElectricoExtractor(BaseExtractor):
    """Extractor de consumo eléctrico comunal desde Energía Abierta (CNE)."""

    @property
    def dataset_name(self) -> str:
        return "consumo_electrico_comunal"

    def fetch(self, **kwargs):
        return fetch_data()

    def normalize(self, raw_data):
        rows, _mode, _url, _notes = raw_data
        return normalize_rows(rows)

    def validate(self, df, metadata: dict) -> dict:
        from src.validation import validate_consumo_electrico_comunal

        return validate_consumo_electrico_comunal(df, metadata)

    def write_staging(self, df, metadata: dict) -> Path:
        ensure_staging_directories()
        output = Path(STAGING_CSV_PATH)
        df.write_csv(str(output))
        merged = {
            **metadata,
            "dataset": self.dataset_name,
            "refreshed_at_utc": datetime.datetime.now(UTC).isoformat(),
            "record_count": df.height,
        }
        write_staging_metadata(METADATA_PATH, merged)
        return output


if __name__ == "__main__":
    process_consumo_electrico()
