import datetime
import json
import sys
import tempfile
import unittest
from pathlib import Path

import polars as pl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from unittest.mock import MagicMock, patch

import openpyxl
from requests import HTTPError

from src.extractors import (
    autoridades_electas_extractor,
    autoridades_locales_extractor,
    bcentral_extractor,
    censo_extractor,
    censo_hogares_viviendas_extractor,
    electoral_extractor,
    mineduc_establecimientos_extractor,
    mineduc_resultados_extractor,
    partidos_politicos_extractor,
    res_extractor,
    salud_extractor,
    siedu_extractor,
    sinim_finanzas_extractor,
    subdere_extractor,
)
from src.extractors.base import BaseExtractor

# ROOT_DIR is defined above


class _MinimalExtractor(BaseExtractor):
    @property
    def dataset_name(self):
        return "test_dataset"

    def fetch(self, **kwargs):
        return {"records": [{"id": 1}]}

    def normalize(self, raw_data):
        return raw_data["records"]

    def validate(self, df, metadata):
        return {"valid": True, "record_count": len(df), "issues": []}

    def write_staging(self, df, metadata):
        raise AssertionError("write_staging no debe ejecutarse durante dry_run")


BCN_FIXTURE = {
    "features": [
        {
            "attributes": {
                "nom_reg": "Tarapaca",
                "nom_prov": "Iquique",
                "nom_com": "Iquique",
                "cod_comuna": 1101,
                "codregion": 1,
            }
        },
        {
            "attributes": {
                "nom_reg": "Tarapaca",
                "nom_prov": "Iquique",
                "nom_com": "Alto Hospicio",
                "cod_comuna": 1107,
                "codregion": 1,
            }
        },
    ]
}


def mock_response(payload, status_code=200):
    response = MagicMock()
    response.status_code = status_code
    response.json.return_value = payload
    response.__enter__.return_value = response
    if status_code >= 400:
        response.raise_for_status.side_effect = HTTPError(f"HTTP {status_code}")
    return response


def _write_censo_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet_totals = workbook.active
    sheet_totals.title = "2"
    sheet_age = workbook.create_sheet("4")
    sheet_totals.append([])
    sheet_age.append([])
    sheet_totals.cell(row=6, column=1, value=1)
    sheet_totals.cell(row=6, column=2, value="Tarapaca")
    sheet_totals.cell(row=6, column=3, value=11)
    sheet_totals.cell(row=6, column=4, value="Iquique")
    sheet_totals.cell(row=6, column=5, value=1101)
    sheet_totals.cell(row=6, column=6, value="Iquique")
    sheet_totals.cell(row=6, column=7, value=100)
    sheet_totals.cell(row=6, column=8, value=48)
    sheet_totals.cell(row=6, column=9, value=52)
    sheet_totals.cell(row=6, column=10, value=92.3)

    age_rows = [
        ("0 a 4", 10),
        ("15 a 19", 20),
        ("30 a 34", 25),
        ("45 a 49", 30),
        ("65 a 69", 15),
    ]
    for offset, (label, value) in enumerate(age_rows, start=6):
        sheet_age.cell(row=offset, column=5, value=1101)
        sheet_age.cell(row=offset, column=7, value=label)
        sheet_age.cell(row=offset, column=8, value=value)

    workbook.save(path)


def _write_censo_hogares_viviendas_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet_viviendas = workbook.active
    sheet_viviendas.title = "2"
    sheet_hogares = workbook.create_sheet("6")
    sheet_viviendas.cell(row=5, column=1, value=1)
    sheet_viviendas.cell(row=5, column=2, value="Tarapaca")
    sheet_viviendas.cell(row=5, column=3, value=11)
    sheet_viviendas.cell(row=5, column=4, value="Iquique")
    sheet_viviendas.cell(row=5, column=5, value=1101)
    sheet_viviendas.cell(row=5, column=6, value="Iquique")
    sheet_viviendas.cell(row=5, column=7, value=120)
    sheet_viviendas.cell(row=5, column=8, value=100)
    sheet_viviendas.cell(row=5, column=9, value=18)
    sheet_viviendas.cell(row=5, column=10, value=2)

    sheet_hogares.cell(row=5, column=5, value=1101)
    sheet_hogares.cell(row=5, column=7, value=90)
    sheet_hogares.cell(row=5, column=8, value=3.1)

    workbook.save(path)


def _write_salud_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "EstablecimientoCodigo;EstablecimientoGlosa;TipoEstablecimientoGlosa;"
                "DependenciaAdministrativa;NivelAtencionEstabglosa;RegionCodigo;RegionGlosa;"
                "ComunaCodigo;ComunaGlosa;TieneServicioUrgencia;TipoUrgencia;Latitud;Longitud;"
                "EstadoFuncionamiento",
                "1001;Hospital Iquique;Hospital;SNSS;Alta;1;Tarapaca;1101;Iquique;Si;"
                "Urgencia;-20.214;-70.152;Vigente",
            ]
        ),
        encoding="utf-8",
    )


class SubdereExtractorTests(unittest.TestCase):
    def test_fetch_bcn_comunas_normalizes_codes_and_saves_snapshot(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.object(subdere_extractor, "RAW_DIR", tmpdir),
            patch.object(
                subdere_extractor,
                "_stealth_get",
                return_value=mock_response(BCN_FIXTURE),
            ),
        ):
            df, skipped, deduped, supplemental = subdere_extractor.fetch_bcn_comunas()

            self.assertEqual(df.filter(df["codigo_comuna"] == "01101").height, 1)
            self.assertEqual(skipped, 0)
            self.assertEqual(deduped, 0)
            self.assertGreaterEqual(supplemental, 1)
            self.assertEqual(len(list(Path(tmpdir).glob("bcn_comunas_*.json"))), 1)

    def test_fetch_bcn_comunas_raises_on_http_error(self):
        with (
            patch.object(subdere_extractor, "_stealth_get", return_value=mock_response({}, 503)),
            self.assertRaises(HTTPError),
        ):
            subdere_extractor.fetch_bcn_comunas()

    def test_normalize_dpa_writes_required_columns(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            metadata_path = str(Path(tmpdir) / "comunas.metadata.json")
            with (
                patch.object(subdere_extractor, "STAGING_DIR", tmpdir),
                patch.object(subdere_extractor, "METADATA_PATH", metadata_path),
                patch.object(subdere_extractor, "fetch_bcn_comunas") as fetch,
            ):
                fetch.return_value = (
                    subdere_extractor.pl.DataFrame(subdere_extractor.DPA_FALLBACK_DATA[:2]),
                    0,
                    0,
                    0,
                )
                output_path = subdere_extractor.normalize_dpa()

            df = subdere_extractor.pl.read_csv(
                output_path,
                schema_overrides={
                    "codigo_region": subdere_extractor.pl.String,
                    "codigo_provincia": subdere_extractor.pl.String,
                    "codigo_comuna": subdere_extractor.pl.String,
                },
            )
            required = {
                "codigo_comuna",
                "nombre_comuna",
                "nombre_comuna_clean",
                "codigo_provincia",
                "nombre_provincia",
                "codigo_region",
                "nombre_region",
                "latitud_cabecera",
                "longitud_cabecera",
                "poblacion_estimada",
            }
            self.assertTrue(required.issubset(df.columns))
            self.assertEqual(df["codigo_comuna"].dtype, subdere_extractor.pl.String)


class BCentralExtractorTests(unittest.TestCase):
    def _payload(self, codigo="uf", year=2026, count=3):
        return {
            "codigo": codigo,
            "serie": [
                {
                    "fecha": datetime.date(year, 1, index + 1).isoformat() + "T00:00:00.000Z",
                    "valor": 39000.0 + index,
                }
                for index in range(count)
            ],
        }

    def test_fetch_indicator_year_parses_records_and_saves_snapshot(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.object(bcentral_extractor, "RAW_DIR", tmpdir),
            patch.object(
                bcentral_extractor,
                "fetch_with_retry",
                return_value=mock_response(self._payload()),
            ),
        ):
            records = bcentral_extractor.fetch_indicator_year("uf", 2026)

            self.assertEqual(len(records), 3)
            self.assertEqual(records[0]["codigo_indicador"], "uf")
            self.assertEqual(len(list(Path(tmpdir).glob("mindicador_uf_2026_*.json"))), 1)

    def test_fetch_indicator_year_raises_on_http_error(self):
        with (
            patch.object(
                bcentral_extractor, "fetch_with_retry", return_value=mock_response({}, 503)
            ),
            self.assertRaises(HTTPError),
        ):
            bcentral_extractor.fetch_indicator_year("uf", 2026)

    def test_fetch_all_history_continues_after_one_indicator_fails(self):
        current_year = datetime.date.today().year

        def fetch(codigo, year):
            if codigo == "uf":
                raise HTTPError("503")
            return [{"fecha": f"{year}-01-01", "codigo_indicador": codigo, "valor": 1.0}]

        with (
            patch.object(
                bcentral_extractor,
                "load_existing_staging",
                return_value=(None, None, []),
            ),
            patch.object(bcentral_extractor, "HISTORY_START_YEAR", current_year),
            patch.object(bcentral_extractor, "fetch_indicator_year", side_effect=fetch),
            patch.object(bcentral_extractor, "load_latest_raw_snapshot", return_value=[]),
            patch.object(bcentral_extractor.time, "sleep"),
        ):
            df, diagnostics = bcentral_extractor.fetch_all_history()

        self.assertIsNotNone(df)
        self.assertTrue(any(item.startswith("uf/") for item in diagnostics["fetch_failures"]))
        self.assertNotIn("uf", set(df["codigo_indicador"].unique()))
        self.assertEqual(set(df["codigo_indicador"].unique()), {"dolar", "euro", "utm", "ipc"})

    def test_fetch_all_history_incremental_with_monthly_gap_and_empty_series(self):
        """Incremental: monthly gap (ipc) is silent, non-monthly empty (dolar) logs warning."""
        import polars as pl

        current_year = datetime.date.today().year
        existing_df = pl.DataFrame(
            {
                "fecha": [datetime.date(current_year - 1, 1, 1)] * 5,
                "codigo_indicador": ["uf", "dolar", "euro", "utm", "ipc"],
                "valor": [39000.0, 900.0, 1040.0, 70000.0, 0.2],
            }
        ).with_columns(pl.col("fecha").cast(pl.Date))

        def fetch(codigo, year):
            if codigo == "ipc" and year == current_year:
                return []
            if codigo == "dolar":
                return []
            return [{"fecha": f"{year}-01-01", "codigo_indicador": codigo, "valor": 1.0}]

        with (
            patch.object(
                bcentral_extractor,
                "load_existing_staging",
                return_value=(existing_df, None, []),
            ),
            patch.object(bcentral_extractor, "fetch_indicator_year", side_effect=fetch),
            patch.object(bcentral_extractor, "load_latest_raw_snapshot", return_value=[]),
            patch.object(bcentral_extractor.time, "sleep"),
        ):
            df, diagnostics = bcentral_extractor.fetch_all_history()

        self.assertIsNotNone(df)
        ipc_empty = [p for p in diagnostics["empty_live_pairs"] if p.startswith("ipc/")]
        self.assertEqual(len(ipc_empty), 0, "ipc monthly gap should not appear in empty_live_pairs")
        dolar_empty = [p for p in diagnostics["empty_live_pairs"] if p.startswith("dolar/")]
        self.assertEqual(len(dolar_empty), 1, "dolar empty should appear in empty_live_pairs")
        self.assertIn("dolar", diagnostics["published_backfills"])
        self.assertEqual(
            set(df["codigo_indicador"].unique()),
            {"uf", "dolar", "euro", "utm", "ipc"},
        )

    def test_fetch_all_history_raw_recovery_and_preserved_pairs(self):
        """Incremental with fetch failures: raw recovery for one, preserved existing for another."""
        import polars as pl

        current_year = datetime.date.today().year
        existing_df = pl.DataFrame(
            {
                "fecha": [datetime.date(current_year - 1, 1, 1)] * 5,
                "codigo_indicador": ["uf", "dolar", "euro", "utm", "ipc"],
                "valor": [39000.0, 900.0, 1040.0, 70000.0, 0.2],
            }
        ).with_columns(pl.col("fecha").cast(pl.Date))

        def fetch(codigo, year):
            if codigo == "uf":
                raise HTTPError("503 for uf")
            if codigo == "ipc":
                raise HTTPError("503 for ipc")
            return [{"fecha": f"{year}-01-01", "codigo_indicador": codigo, "valor": 1.0}]

        def raw_snapshot(codigo, year):
            if codigo == "uf":
                return [{"fecha": f"{year}-01-01", "codigo_indicador": "uf", "valor": 39001.0}]
            return []

        with (
            patch.object(
                bcentral_extractor,
                "load_existing_staging",
                return_value=(existing_df, None, []),
            ),
            patch.object(bcentral_extractor, "fetch_indicator_year", side_effect=fetch),
            patch.object(bcentral_extractor, "load_latest_raw_snapshot", side_effect=raw_snapshot),
            patch.object(bcentral_extractor.time, "sleep"),
        ):
            df, diagnostics = bcentral_extractor.fetch_all_history()

        self.assertIsNotNone(df)
        self.assertTrue(any("uf/" in f for f in diagnostics["fetch_failures"]))
        self.assertTrue(any("uf/" in r for r in diagnostics["raw_recoveries"]))
        self.assertTrue(any("ipc/" in f for f in diagnostics["fetch_failures"]))
        self.assertTrue(any("ipc/" in p for p in diagnostics["preserved_existing_pairs"]))
        self.assertEqual(
            set(df["codigo_indicador"].unique()),
            {"uf", "dolar", "euro", "utm", "ipc"},
        )
        uf_rows = df.filter(pl.col("codigo_indicador") == "uf")
        self.assertIn(39001.0, uf_rows["valor"].to_list())


class SinimFinanzasExtractorTests(unittest.TestCase):
    def test_normalize_rows_writes_required_schema(self):
        df = sinim_finanzas_extractor.normalize_rows(sinim_finanzas_extractor.FALLBACK_ROWS)
        required = {
            "anio",
            "codigo_comuna",
            "nombre_comuna",
            "ingresos_totales",
            "gastos_totales",
            "ingresos_propios_permanentes",
            "fondo_comun_municipal",
            "gasto_personal",
            "gasto_inversion",
        }
        self.assertTrue(required.issubset(df.columns))
        self.assertEqual(df["codigo_comuna"].dtype, sinim_finanzas_extractor.pl.String)

    def test_run_dry_run_returns_validation_without_writing(self):
        with patch.object(
            sinim_finanzas_extractor,
            "fetch_data",
            return_value=(sinim_finanzas_extractor.FALLBACK_ROWS, "fallback", "url", []),
        ):
            result = sinim_finanzas_extractor.SinimFinanzasExtractor().run(dry_run=True)
        self.assertEqual(result["status"], "ok")


class MineducResultadosExtractorTests(unittest.TestCase):
    def test_normalize_rows_writes_required_schema(self):
        df = mineduc_resultados_extractor.normalize_rows(mineduc_resultados_extractor.FALLBACK_ROWS)
        required = {
            "anio",
            "codigo_comuna",
            "matricula_total",
            "asistencia_promedio",
            "tasa_aprobacion",
            "tasa_reprobacion",
            "tasa_retiro",
            "establecimientos_reportados",
        }
        self.assertTrue(required.issubset(df.columns))
        self.assertEqual(df["codigo_comuna"].dtype, mineduc_resultados_extractor.pl.String)

    def test_run_dry_run_returns_validation_without_writing(self):
        with patch.object(
            mineduc_resultados_extractor,
            "fetch_data",
            return_value=(mineduc_resultados_extractor.FALLBACK_ROWS, "fallback", "url", []),
        ):
            result = mineduc_resultados_extractor.MineducResultadosExtractor().run(dry_run=True)
        self.assertEqual(result["status"], "ok")


class SieduExtractorTests(unittest.TestCase):
    def test_normalize_rows_writes_required_schema(self):
        df = siedu_extractor.normalize_rows(siedu_extractor.FALLBACK_ROWS)
        required = {
            "anio",
            "codigo_comuna",
            "codigo_indicador",
            "nombre_indicador",
            "categoria",
            "valor",
            "unidad",
            "fuente_original",
            "cobertura_tipo",
        }
        self.assertTrue(required.issubset(df.columns))
        self.assertEqual(df["codigo_comuna"].dtype, siedu_extractor.pl.String)

    def test_run_dry_run_returns_validation_without_writing(self):
        with patch.object(
            siedu_extractor,
            "fetch_data",
            return_value=(siedu_extractor.FALLBACK_ROWS, "fallback", "url", []),
        ):
            result = siedu_extractor.SieduExtractor().run(dry_run=True)
        self.assertEqual(result["status"], "ok")


class BaseExtractorContractTests(unittest.TestCase):
    def test_base_class_is_abstract(self):
        with self.assertRaises(TypeError):
            BaseExtractor()

    def test_run_dry_run_returns_validation_without_writing(self):
        result = _MinimalExtractor().run(dry_run=True)
        self.assertTrue(result["valid"])
        self.assertEqual(result["record_count"], 1)

    def test_concrete_extractors_publish_dataset_names(self):
        self.assertEqual(subdere_extractor.SubdereExtractor().dataset_name, "comunas")
        self.assertEqual(bcentral_extractor.BCentralExtractor().dataset_name, "indicadores")
        self.assertEqual(censo_extractor.CensoExtractor().dataset_name, "censo_comunal")
        self.assertEqual(
            censo_hogares_viviendas_extractor.CensoHogaresViviendasExtractor().dataset_name,
            "censo_hogares_viviendas",
        )
        self.assertEqual(salud_extractor.SaludExtractor().dataset_name, "establecimientos_salud")

    def test_censo_parser_preserves_cut_codes_and_age_totals(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "censo_comunal.xlsx"
            _write_censo_workbook(workbook)

            df = censo_extractor.parse_workbook(workbook)

        self.assertEqual(df.height, 1)
        self.assertEqual(df["codigo_comuna"].item(), "01101")
        age_total = sum(df[column] for column in censo_extractor.AGE_BANDS)
        self.assertEqual(df.filter(age_total != df["poblacion_censada"]).height, 0)

    def test_salud_parser_preserves_cut_codes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "salud.csv"
            _write_salud_csv(source)

            df = salud_extractor.parse_csv(source)

        self.assertEqual(df.height, 1)
        self.assertEqual(df["codigo_comuna"].item(), "01101")

    def test_censo_hogares_viviendas_parser_preserves_cut_codes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "censo_hogares_viviendas.xlsx"
            _write_censo_hogares_viviendas_workbook(workbook)

            df = censo_hogares_viviendas_extractor.parse_workbook(workbook)

        self.assertEqual(df.height, 1)
        self.assertEqual(df["codigo_comuna"].item(), "01101")

    def test_censo_hogares_viviendas_write_staging_persists_csv_and_metadata(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "censo_hogares_viviendas.xlsx"
            _write_censo_hogares_viviendas_workbook(workbook)
            df = censo_hogares_viviendas_extractor.parse_workbook(workbook)
            csv_path = Path(tmpdir) / "censo_hogares_viviendas.csv"
            metadata_path = Path(tmpdir) / "censo_hogares_viviendas.metadata.json"
            with (
                patch.object(censo_hogares_viviendas_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(
                    censo_hogares_viviendas_extractor, "METADATA_PATH", str(metadata_path)
                ),
            ):
                censo_hogares_viviendas_extractor.CensoHogaresViviendasExtractor().write_staging(
                    df, {"source_mode": "fallback"}
                )

            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            self.assertTrue(csv_path.exists())
            self.assertEqual(metadata["record_count"], df.height)

    def test_concrete_write_staging_persists_csv_and_metadata(self):
        df = bcentral_extractor.generate_fallback_indicators()
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "indicadores.csv"
            metadata_path = Path(tmpdir) / "indicadores.metadata.json"
            with (
                patch.object(bcentral_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(bcentral_extractor, "METADATA_PATH", str(metadata_path)),
            ):
                bcentral_extractor.BCentralExtractor().write_staging(
                    df, {"source_mode": "fallback"}
                )

            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            self.assertTrue(csv_path.exists())
            self.assertEqual(metadata["dataset"], "indicadores")
            self.assertEqual(metadata["record_count"], df.height)


class ResExtractorTests(unittest.TestCase):
    """Tests para el extractor del Registro de Empresas y Sociedades (RES)."""

    def test_extractor_class_contract(self):
        """El extractor cumple con el contrato de BaseExtractor."""
        extractor = res_extractor.ResExtractor()
        self.assertEqual(extractor.dataset_name, "empresas")

    def test_parse_resources_smoke(self):
        """Parseo de un CSV sintetico con las columnas esperadas del RES."""
        csv_content = (
            "ID;RUT;Razon Social;Fecha de actuacion (1era firma);"
            "Fecha de registro (ultima firma);Fecha de aprobacion x SII;"
            "Anio;Mes;Comuna Tributaria;Region Tributaria;"
            "Codigo de sociedad;Tipo de actuacion;Capital;Comuna Social;Region Social\n"
            "1;76286049-K;Servicios Digitales EIRL;02-05-2022;02-05-2022;02-05-2022;"
            "2022;Mayo;Providencia;13;EIRL;CONSTITUCION;7000000;Providencia;13\n"
            "2;76286055-4;Comercial ABC SpA;03-06-2022;03-06-2022;03-06-2022;"
            "2022;Junio;Las Condes;13;SPA;CONSTITUCION;25000000;Las Condes;13\n"
            "3;96567890-1;Distribuidora Sur SRL;15-06-2022;15-06-2022;15-06-2022;"
            "2022;Junio;Valparaiso;05;SRL;CONSTITUCION;5000000;Valparaiso;05\n"
        )
        contents = [csv_content.encode("utf-8")]
        df = res_extractor.parse_resources(contents)

        self.assertGreater(df.height, 0)
        self.assertIn("rut", df.columns)
        self.assertIn("razon_social", df.columns)
        self.assertIn("codigo_sociedad", df.columns)
        self.assertIn("capital", df.columns)
        self.assertIn("region_tributaria", df.columns)
        self.assertIn("comuna_tributaria", df.columns)

        # Verificar normalizacion de tipos
        ruts = df["rut"].to_list()
        self.assertIn("76286049-K", ruts)

        # codigo_sociedad debe estar mapeado a formato canonico
        sociedades = df["codigo_sociedad"].to_list()
        self.assertIn("EIRL", sociedades)
        self.assertIn("SpA", sociedades)  # SPA -> SpA (title case canonico)
        self.assertIn("SRL", sociedades)

    def test_parse_resources_handles_empty(self):
        """Parseo de lista vacia lanza SystemExit."""
        with self.assertRaises(SystemExit):
            res_extractor.parse_resources([])

    def test_write_staging_dry_run(self):
        """El extractor en dry_run no persiste archivos."""

        # Datos sinteticos minimos para evitar llamadas de red
        csv_content = (
            "ID;RUT;Razon Social;Fecha de actuacion (1era firma);"
            "Fecha de registro (ultima firma);Fecha de aprobacion x SII;"
            "Anio;Mes;Comuna Tributaria;Region Tributaria;"
            "Codigo de sociedad;Tipo de actuacion;Capital;Comuna Social;Region Social\n"
            "1;76286049-K;Test EIRL;02-05-2022;02-05-2022;02-05-2022;"
            "2022;Mayo;Santiago;13;EIRL;CONSTITUCION;1000000;Santiago;13\n"
        )
        fake_contents = [csv_content.encode("utf-8")]

        with patch.object(
            res_extractor, "fetch_resources", return_value=(fake_contents, "live", "test")
        ):
            extractor = res_extractor.ResExtractor()
            result = extractor.run(dry_run=True)
            self.assertIn("status", result)
            self.assertIn(result["status"], ("ok", "error"))

    def test_reuse_policy(self):
        """La politica de reutilizacion esta definida y es abierta."""
        policy = res_extractor.REUSE_POLICY
        self.assertEqual(policy["status"], "open-attribution")
        self.assertTrue(policy["redistribution_ok"])
        self.assertIn("CC-BY", policy["license"])


class SourceAdapterTests(unittest.TestCase):
    """Tests para los helpers reutilizables de extractores (source_adapter.py)."""

    def test_fetch_url_snapshot_success(self):
        """fetch_url_snapshot exitoso retorna success=True con contenido."""
        from src.extractors.source_adapter import fetch_url_snapshot

        mock_content = b"<html><body>datos</body></html>"
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.content = mock_content
        mock_resp.__enter__.return_value = mock_resp
        with patch("src.extractors.source_adapter.fetch_with_retry", return_value=mock_resp):
            with tempfile.TemporaryDirectory() as tmp:
                raw_dir = Path(tmp) / "raw"
                success, content, note, data_parsed = fetch_url_snapshot(
                    "https://example.com/data", raw_dir, "test_dataset"
                )
                self.assertTrue(success)
                self.assertEqual(content, mock_content)
                self.assertEqual(note, "official_landing_snapshot_saved")
                self.assertFalse(data_parsed)  # snapshot crudo, no procesado aún
                # Verifica que el snapshot se guardó en disco
                snapshots = list(raw_dir.glob("test_dataset_*.html"))
                self.assertEqual(len(snapshots), 1)

    def test_fetch_url_snapshot_failure(self):
        """fetch_url_snapshot ante error de red retorna success=False."""
        from src.extractors.source_adapter import fetch_url_snapshot

        with patch(
            "src.extractors.source_adapter.fetch_with_retry",
            side_effect=ConnectionError("timeout"),
        ):
            with tempfile.TemporaryDirectory() as tmp:
                raw_dir = Path(tmp) / "raw"
                success, content, note, data_parsed = fetch_url_snapshot(
                    "https://example.com/data", raw_dir, "test_dataset"
                )
                self.assertFalse(success)
                self.assertIsNone(content)
                self.assertIn("official_landing_unavailable", note)
                self.assertIn("timeout", note)
                self.assertFalse(data_parsed)

    def test_source_mode_from_live_success(self):
        """source_mode_from_live_success: True → live, False → fallback."""
        from src.extractors.source_adapter import source_mode_from_live_success

        # Sin data_parsed, incluso HTTP exitoso es fallback (snapshot no procesado)
        self.assertEqual(source_mode_from_live_success(True), "fallback")
        self.assertEqual(source_mode_from_live_success(False), "fallback")
        # Solo cuando success=True Y data_parsed=True se considera live genuino
        self.assertEqual(source_mode_from_live_success(True, data_parsed=True), "live")
        self.assertEqual(source_mode_from_live_success(False, data_parsed=True), "fallback")
        self.assertEqual(source_mode_from_live_success(True, data_parsed=False), "fallback")

    def test_fallback_metadata_note(self):
        """fallback_metadata_note produce nota con prefijo estandarizado."""
        from src.extractors.source_adapter import fallback_metadata_note

        note = fallback_metadata_note("endpoint no disponible")
        self.assertIn("fallback_curated_rows_used:", note)
        self.assertIn("endpoint no disponible", note)

    def test_build_standard_metadata(self):
        """build_standard_metadata produce dict con todos los campos requeridos."""
        import polars as pl

        from src.extractors.source_adapter import build_standard_metadata

        df = pl.DataFrame({"codigo": ["13101", "13102"], "nombre": ["Santiago", "Iquique"]})
        meta = build_standard_metadata(
            dataset="test_dataset",
            source_name="API de prueba",
            source_url="https://example.com",
            source_mode="live",
            source_detail="public_api",
            df=df,
            notes=["prueba"],
            reuse_policy={"status": "open-attribution", "license": "CC-BY"},
        )
        self.assertEqual(meta["dataset"], "test_dataset")
        self.assertEqual(meta["source_name"], "API de prueba")
        self.assertEqual(meta["source_url"], "https://example.com")
        self.assertEqual(meta["source_mode"], "live")
        self.assertEqual(meta["source_detail"], "public_api")
        self.assertEqual(meta["record_count"], 2)
        self.assertEqual(meta["fields"], ["codigo", "nombre"])
        self.assertIn("prueba", meta["notes"])
        self.assertEqual(meta["reuse_policy"]["status"], "open-attribution")


class BaseExtractorEdgeTests(unittest.TestCase):
    """Tests de borde para el contrato BaseExtractor."""

    def test_ensure_staging_directories_idempotent(self):
        """ensure_staging_directories no falla si los directorios ya existen."""
        from src.extractors.base import ensure_staging_directories

        # Llamar dos veces: la primera crea, la segunda es idempotente
        ensure_staging_directories()
        ensure_staging_directories()
        # Si no lanza excepción, el test pasa


class ElectoralExtractorTests(unittest.TestCase):
    """Tests para el extractor de mapeo electoral (electoral_extractor.py)."""

    @classmethod
    def setUpClass(cls):
        cls._orig_staging_dir = electoral_extractor.STAGING_DIR

    def _write_comunas_stub(self, staging_dir: Path) -> Path:
        """Escribe un CSV de comunas mínimo para alimentar build_electoral_df."""
        csv_path = staging_dir / "comunas.csv"
        csv_path.write_text(
            "codigo_region,codigo_comuna,nombre_comuna,nombre_comuna_clean\n"
            "01,01101,Iquique,iquique\n"
            "13,13101,Santiago,santiago\n"
            "13,13114,Las Condes,las condes\n"
            "13,13120,Nunoa,nunoa\n"
        )
        return csv_path

    def test_build_electoral_df_no_comunas_csv_raises(self):
        """build_electoral_df lanza FileNotFoundError sin comunas.csv."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch.object(electoral_extractor, "STAGING_DIR", tmp):
                with self.assertRaises(FileNotFoundError):
                    electoral_extractor.build_electoral_df()

    def test_build_electoral_df_produces_expected_columns(self):
        """build_electoral_df produce row con columnas requeridas."""
        with tempfile.TemporaryDirectory() as tmp:
            staging = Path(tmp)
            self._write_comunas_stub(staging)
            with patch.object(electoral_extractor, "STAGING_DIR", str(staging)):
                df = electoral_extractor.build_electoral_df()
            self.assertIn("distrito_electoral", df.columns)
            self.assertIn("circunscripcion_senatorial", df.columns)
            self.assertIn("codigo_comuna", df.columns)
            self.assertIn("nombre_comuna", df.columns)
            self.assertGreaterEqual(df.height, 1)

    def test_electoral_extractor_dry_run(self):
        """ElectoralExtractor.run(dry_run=True) retorna validación sin escribir."""
        with tempfile.TemporaryDirectory() as tmp:
            staging = Path(tmp)
            self._write_comunas_stub(staging)
            with patch.object(electoral_extractor, "STAGING_DIR", str(staging)):
                extractor = electoral_extractor.ElectoralExtractor()
                result = extractor.run(dry_run=True)
            self.assertIn("status", result)
            # No debe crear archivos en staging durante dry_run
            csv_candidates = list(staging.glob("distritos_electorales*"))
            self.assertEqual(len(csv_candidates), 0)

    def test_electoral_extractor_dataset_name(self):
        """dataset_name retorna el identificador canónico."""
        self.assertEqual(
            electoral_extractor.ElectoralExtractor().dataset_name,
            "distritos_electorales",
        )


class MineducEstablecimientosExtractorTests(unittest.TestCase):
    """Tests para el extractor de establecimientos educacionales MINEDUC."""

    _CSV_HEADER = (
        "RBD;DGV_RBD;NOM_RBD;COD_REG_RBD;COD_COM_RBD;COD_DEPE2;LATITUD;LONGITUD;ESTADO_ESTAB\n"
    )

    def _write_fixture_csv(self, path: Path, extra_rows: str = "") -> Path:
        # El CSV oficial de MINEDUC usa coma como separador decimal en latitud/longitud
        rows = (
            "1;5;Escuela Uno;13;13101;1;-33,45;-70,65;1\n"
            "2;7;Liceo Dos;13;13114;2;-33,42;-70,60;2\n"
            "99;3;Colegio Cerrado;13;13101;3;-33,44;-70,63;3\n"
        )
        path.write_text(self._CSV_HEADER + rows + extra_rows, encoding="utf-8")
        return path

    def test_parse_csv_filters_closed_establishments(self):
        """parse_csv excluye establecimientos con ESTADO_ESTAB=3 (Cerrado)."""
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "test_ee.csv"
            self._write_fixture_csv(csv_path)
            df = mineduc_establecimientos_extractor.parse_csv(csv_path)
            # El registro con ESTADO_ESTAB=3 debe ser filtrado
            self.assertEqual(df.height, 2)
            rbd_values = df["rbd"].to_list()
            self.assertIn("1", rbd_values)
            self.assertIn("2", rbd_values)
            self.assertNotIn("99", rbd_values)

    def test_parse_csv_required_columns(self):
        """parse_csv produce las columnas del esquema canónico."""
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "test_ee.csv"
            self._write_fixture_csv(csv_path)
            df = mineduc_establecimientos_extractor.parse_csv(csv_path)
            expected = {
                "rbd",
                "dv_rbd",
                "nombre_establecimiento",
                "codigo_region",
                "codigo_comuna",
                "dependencia_administrativa",
                "latitud",
                "longitud",
                "estado_funcionamiento",
            }
            self.assertEqual(set(df.columns), expected)

    def test_parse_csv_cut_codes_are_fixed_width(self):
        """parse_csv asegura codigo_comuna de 5 chars y codigo_region de 2 chars."""
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "test_ee.csv"
            self._write_fixture_csv(csv_path)
            df = mineduc_establecimientos_extractor.parse_csv(csv_path)
            self.assertEqual(df["codigo_comuna"].str.len_chars().unique().to_list(), [5])
            self.assertEqual(df["codigo_region"].str.len_chars().unique().to_list(), [2])

    def test_extractor_dry_run(self):
        """MineducEstablecimientosExtractor.run(dry_run=True) no escribe archivos."""
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "test_ee.csv"
            self._write_fixture_csv(csv_path)
            with (
                patch.object(
                    mineduc_establecimientos_extractor,
                    "fetch_data",
                    return_value=(csv_path, "live", "https://example.com"),
                ),
                patch.object(mineduc_establecimientos_extractor, "STAGING_DIR", tmp),
                patch.object(
                    mineduc_establecimientos_extractor, "METADATA_PATH", Path(tmp) / "meta.json"
                ),
            ):
                extractor = mineduc_establecimientos_extractor.MineducEstablecimientosExtractor()
                result = extractor.run(dry_run=True)
                self.assertIn("status", result)
                # No debe escribir staging CSV
                csv_files = list(Path(tmp).glob("*.csv"))
                self.assertEqual(len(csv_files), 1)  # solo el fixture

    def test_extractor_dataset_name(self):
        """dataset_name retorna el identificador canónico."""
        self.assertEqual(
            mineduc_establecimientos_extractor.MineducEstablecimientosExtractor().dataset_name,
            "establecimientos_educacionales",
        )


# ── B6: Tests extendidos para extractores con cobertura parcial ──


class SaludExtractorExtendedTests(unittest.TestCase):
    """Fetch/write_staging/dry_run para salud_extractor (43% → +cobertura)."""

    def test_write_staging_persists_csv_and_metadata(self):
        df = pl.DataFrame(
            {
                "codigo_establecimiento": ["1001"],
                "nombre_establecimiento": ["Hospital Test"],
                "tipo_establecimiento": ["Hospital"],
                "dependencia_administrativa": ["SNSS"],
                "nivel_atencion": ["Alta"],
                "codigo_region": ["01"],
                "nombre_region": ["Tarapaca"],
                "codigo_comuna": ["01101"],
                "nombre_comuna": ["Iquique"],
                "tiene_servicio_urgencia": ["Si"],
                "tipo_urgencia": ["Urgencia"],
                "latitud": [-20.214],
                "longitud": [-70.152],
                "estado_funcionamiento": ["Vigente"],
            }
        )
        metadata = {
            "dataset": "establecimientos_salud",
            "source_mode": "live",
            "record_count": df.height,
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "establecimientos_salud.csv"
            metadata_path = Path(tmpdir) / "establecimientos_salud.metadata.json"
            with (
                patch.object(salud_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(salud_extractor, "METADATA_PATH", str(metadata_path)),
            ):
                result = salud_extractor.SaludExtractor().write_staging(df, metadata)
            self.assertEqual(result, csv_path)
            self.assertTrue(csv_path.exists())
            self.assertTrue(metadata_path.exists())
            saved_meta = json.loads(metadata_path.read_text(encoding="utf-8"))
            self.assertEqual(saved_meta["record_count"], df.height)

    def test_run_dry_run_returns_validation_without_writing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "test_salud.csv"
            _write_salud_csv(csv_path)
            staging_csv = Path(tmpdir) / "establecimientos_salud.csv"
            staging_meta = Path(tmpdir) / "establecimientos_salud.metadata.json"
            with (
                patch.object(salud_extractor, "STAGING_CSV_PATH", str(staging_csv)),
                patch.object(salud_extractor, "METADATA_PATH", str(staging_meta)),
                patch.object(
                    salud_extractor,
                    "fetch_csv",
                    return_value=(csv_path, "live", "https://example.com"),
                ),
            ):
                extractor = salud_extractor.SaludExtractor()
                result = extractor.run(dry_run=True)
            self.assertIn("status", result)
            self.assertFalse(staging_csv.exists())
            self.assertFalse(staging_meta.exists())


class SieduExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para siedu_extractor (52% → +cobertura)."""

    def test_write_staging_persists_csv_and_metadata(self):
        df = siedu_extractor.normalize_rows(siedu_extractor.FALLBACK_ROWS)
        metadata = {
            "dataset": "indicadores_urbanos_siedu",
            "source_mode": "fallback",
            "record_count": df.height,
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "indicadores_urbanos_siedu.csv"
            metadata_path = Path(tmpdir) / "indicadores_urbanos_siedu.metadata.json"
            with (
                patch.object(siedu_extractor, "STAGING_CSV_PATH", csv_path),
                patch.object(siedu_extractor, "METADATA_PATH", metadata_path),
            ):
                result = siedu_extractor.SieduExtractor().write_staging(df, metadata)
            self.assertEqual(result, csv_path)
            self.assertTrue(csv_path.exists())
            saved_meta = json.loads(metadata_path.read_text(encoding="utf-8"))
            self.assertEqual(saved_meta["record_count"], df.height)


class MineducResultadosExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para mineduc_resultados_extractor (54% → +cobertura)."""

    def test_fetch_and_normalize_produces_expected_schema(self):
        with patch.object(
            mineduc_resultados_extractor,
            "fetch_data",
            return_value=(
                mineduc_resultados_extractor.FALLBACK_ROWS,
                "fallback",
                "https://example.com",
                [],
            ),
        ):
            extractor = mineduc_resultados_extractor.MineducResultadosExtractor()
            raw = extractor.fetch()
            df = extractor.normalize(raw)
        required = {
            "anio",
            "codigo_comuna",
            "matricula_total",
            "asistencia_promedio",
            "tasa_aprobacion",
            "tasa_reprobacion",
            "tasa_retiro",
            "establecimientos_reportados",
        }
        self.assertTrue(required.issubset(df.columns))
        self.assertEqual(df["codigo_comuna"].dtype, pl.String)

    def test_run_dry_run_does_not_write_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            staging_dir = Path(tmp)
            csv_path = staging_dir / "resultados_educacionales.csv"
            meta_path = staging_dir / "resultados_educacionales.metadata.json"
            with (
                patch.object(
                    mineduc_resultados_extractor,
                    "fetch_data",
                    return_value=(
                        mineduc_resultados_extractor.FALLBACK_ROWS,
                        "fallback",
                        "url",
                        [],
                    ),
                ),
                patch.object(mineduc_resultados_extractor, "STAGING_DIR", staging_dir),
                patch.object(mineduc_resultados_extractor, "STAGING_CSV_PATH", csv_path),
                patch.object(mineduc_resultados_extractor, "METADATA_PATH", meta_path),
            ):
                result = mineduc_resultados_extractor.MineducResultadosExtractor().run(dry_run=True)
            self.assertEqual(result["status"], "ok")
            self.assertFalse(csv_path.exists())
            self.assertFalse(meta_path.exists())


class SinimFinanzasExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para sinim_finanzas_extractor (54% → +cobertura)."""

    def test_fetch_and_normalize_produces_expected_schema(self):
        with patch.object(
            sinim_finanzas_extractor,
            "fetch_data",
            return_value=(
                sinim_finanzas_extractor.FALLBACK_ROWS,
                "fallback",
                "https://example.com",
                [],
            ),
        ):
            extractor = sinim_finanzas_extractor.SinimFinanzasExtractor()
            raw = extractor.fetch()
            df = extractor.normalize(raw)
        required = {
            "anio",
            "codigo_comuna",
            "nombre_comuna",
            "ingresos_totales",
            "gastos_totales",
            "ingresos_propios_permanentes",
            "fondo_comun_municipal",
            "gasto_personal",
            "gasto_inversion",
        }
        self.assertTrue(required.issubset(df.columns))
        self.assertEqual(df["codigo_comuna"].dtype, pl.String)


class CensoExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para censo_extractor (56% → +cobertura)."""

    def test_dry_run_does_not_write_files(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "censo.xlsx"
            _write_censo_workbook(workbook)
            csv_path = Path(tmpdir) / "censo_comunal.csv"
            meta_path = Path(tmpdir) / "censo_comunal.metadata.json"
            with (
                patch.object(censo_extractor, "fetch_workbook", return_value=(workbook, "live")),
                patch.object(censo_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(censo_extractor, "METADATA_PATH", str(meta_path)),
            ):
                result = censo_extractor.CensoExtractor().run(dry_run=True)
            self.assertIn("status", result)
            self.assertFalse(csv_path.exists())
            self.assertFalse(meta_path.exists())


class CensoHogaresViviendasExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para censo_hogares_viviendas (60% → +cobertura)."""

    def test_dry_run_does_not_write_files(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "hogares.xlsx"
            _write_censo_hogares_viviendas_workbook(workbook)
            csv_path = Path(tmpdir) / "censo_hogares_viviendas.csv"
            meta_path = Path(tmpdir) / "censo_hogares_viviendas.metadata.json"
            with (
                patch.object(
                    censo_hogares_viviendas_extractor,
                    "fetch_workbook",
                    return_value=(workbook, "live"),
                ),
                patch.object(censo_hogares_viviendas_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(censo_hogares_viviendas_extractor, "METADATA_PATH", str(meta_path)),
            ):
                result = censo_hogares_viviendas_extractor.CensoHogaresViviendasExtractor().run(
                    dry_run=True
                )
            self.assertIn("status", result)
            self.assertFalse(csv_path.exists())


class ResExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para res_extractor (60% → +cobertura)."""

    def test_write_staging_persists_csv_and_metadata(self):
        df = pl.DataFrame(
            {
                "rut": ["76286049-K"],
                "razon_social": ["Test EIRL"],
                "codigo_sociedad": ["EIRL"],
                "tipo_actuacion": ["CONSTITUCION"],
                "capital": [1000000],
                "fecha_actuacion": [datetime.date(2022, 5, 2)],
                "fecha_registro": [datetime.date(2022, 5, 2)],
                "fecha_aprobacion_sii": [datetime.date(2022, 5, 2)],
                "anio": [2022],
                "mes": ["Mayo"],
                "comuna_tributaria": ["Santiago"],
                "region_tributaria": ["13"],
                "comuna_social": ["Santiago"],
                "region_social": ["13"],
            }
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "empresas.csv"
            meta_path = Path(tmpdir) / "empresas.metadata.json"
            with (
                patch.object(res_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(res_extractor, "METADATA_PATH", str(meta_path)),
            ):
                result = res_extractor.ResExtractor().write_staging(df, {"source_mode": "fallback"})
            self.assertEqual(result, csv_path)
            self.assertTrue(csv_path.exists())
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            self.assertEqual(meta["dataset"], "empresas")

    def test_dry_run_does_not_write_files(self):
        csv_content = (
            "ID;RUT;Razon Social;Fecha de actuacion (1era firma);Fecha de registro (ultima firma);"
            "Fecha de aprobacion x SII;Anio;Mes;Comuna Tributaria;Region Tributaria;"
            "Codigo de sociedad;Tipo de actuacion;Capital;Comuna Social;Region Social\n"
            "1;76286049-K;Test EIRL;02-05-2022;02-05-2022;02-05-2022;"
            "2022;Mayo;Santiago;13;EIRL;CONSTITUCION;1000000;Santiago;13\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "empresas.csv"
            meta_path = Path(tmpdir) / "empresas.metadata.json"
            with (
                patch.object(
                    res_extractor,
                    "fetch_resources",
                    return_value=([csv_content.encode("utf-8")], "live", "test"),
                ),
                patch.object(res_extractor, "STAGING_CSV_PATH", str(csv_path)),
                patch.object(res_extractor, "METADATA_PATH", str(meta_path)),
            ):
                result = res_extractor.ResExtractor().run(dry_run=True)
                self.assertIn("status", result)
                self.assertFalse(csv_path.exists())
                self.assertFalse(meta_path.exists())


class SubdereExtractorExtendedTests(unittest.TestCase):
    """Tests extendidos para subdere_extractor (54% → +cobertura)."""

    def test_dry_run_does_not_write_files(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            meta_path = Path(tmpdir) / "comunas.metadata.json"
            with (
                patch.object(subdere_extractor, "STAGING_DIR", tmpdir),
                patch.object(subdere_extractor, "METADATA_PATH", str(meta_path)),
                patch.object(subdere_extractor, "fetch_bcn_comunas") as fetch,
            ):
                fetch.return_value = (
                    pl.DataFrame(subdere_extractor.DPA_FALLBACK_DATA[:2]),
                    0,
                    0,
                    0,
                )
                result = subdere_extractor.SubdereExtractor().run(dry_run=True)
                self.assertIn("status", result)
                self.assertFalse(meta_path.exists())

    def test_write_staging_persists_csv_and_metadata(self):
        df = pl.DataFrame(subdere_extractor.DPA_FALLBACK_DATA[:2])
        with tempfile.TemporaryDirectory() as tmpdir:
            meta_path = Path(tmpdir) / "comunas.metadata.json"
            with (
                patch.object(subdere_extractor, "STAGING_DIR", tmpdir),
                patch.object(subdere_extractor, "METADATA_PATH", str(meta_path)),
            ):
                result = subdere_extractor.SubdereExtractor().write_staging(
                    df, {"source_mode": "live"}
                )
            output_csv = Path(tmpdir) / "comunas.csv"
            self.assertEqual(result, output_csv)
            self.assertTrue(output_csv.exists())
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            self.assertEqual(meta["dataset"], "comunas")


class HttpUtilsRetryTests(unittest.TestCase):
    """Tests de reintentos HTTP para http_utils.fetch_with_retry e _is_retryable."""

    def test_succeeds_on_first_attempt(self):
        """Sin fallos, retorna la respuesta directamente sin reintentar."""
        from src.extractors.http_utils import fetch_with_retry

        ok_resp = mock_response({"ok": True})
        mock_get = MagicMock(return_value=ok_resp)
        result = fetch_with_retry("https://example.com", get_fn=mock_get, timeout=5)
        self.assertEqual(result.status_code, 200)
        mock_get.assert_called_once_with("https://example.com", timeout=5)

    @patch("tenacity.nap.sleep")
    def test_retries_on_connection_error_and_succeeds(self, _sleep):
        """Reintenta en ConnectionError transitorio y retorna respuesta al tercer intento."""
        import requests as _req

        from src.extractors.http_utils import fetch_with_retry

        ok_resp = mock_response({"ok": True})
        mock_get = MagicMock(
            side_effect=[
                _req.exceptions.ConnectionError("fallo transitorio 1"),
                _req.exceptions.ConnectionError("fallo transitorio 2"),
                ok_resp,
            ]
        )
        result = fetch_with_retry("https://example.com", get_fn=mock_get, max_attempts=3, timeout=5)
        self.assertEqual(mock_get.call_count, 3)
        self.assertEqual(result.status_code, 200)

    @patch("tenacity.nap.sleep")
    def test_raises_after_max_attempts_exhausted(self, _sleep):
        """Propaga ConnectionError tras agotar todos los reintentos (reraise=True)."""
        import requests as _req

        from src.extractors.http_utils import fetch_with_retry

        mock_get = MagicMock(side_effect=_req.exceptions.ConnectionError("sin red"))
        with self.assertRaises(_req.exceptions.ConnectionError):
            fetch_with_retry("https://example.com", get_fn=mock_get, max_attempts=2, timeout=5)
        self.assertEqual(mock_get.call_count, 2)

    def test_does_not_retry_4xx(self):
        """No reintenta errores del cliente (4xx) — los retorna tal cual."""
        from src.extractors.http_utils import fetch_with_retry

        not_found = mock_response(None, status_code=404)
        mock_get = MagicMock(return_value=not_found)
        result = fetch_with_retry("https://example.com", get_fn=mock_get, timeout=5)
        self.assertEqual(result.status_code, 404)
        mock_get.assert_called_once()

    @patch("tenacity.nap.sleep")
    def test_retries_on_5xx_and_succeeds(self, _sleep):
        """Reintenta respuestas 5xx hasta recuperarse en el segundo intento."""
        import requests as _req

        from src.extractors.http_utils import fetch_with_retry

        err_resp = MagicMock()
        err_resp.status_code = 503
        err_resp.raise_for_status.side_effect = _req.exceptions.HTTPError(
            "503 Service Unavailable", response=err_resp
        )
        ok_resp = mock_response({"ok": True})
        mock_get = MagicMock(side_effect=[err_resp, ok_resp])
        result = fetch_with_retry("https://example.com", get_fn=mock_get, max_attempts=3, timeout=5)
        self.assertEqual(mock_get.call_count, 2)
        self.assertEqual(result.status_code, 200)

    def test_is_retryable_connection_error(self):
        """_is_retryable: True para requests.exceptions.ConnectionError."""
        import requests as _req

        from src.extractors.http_utils import _is_retryable

        self.assertTrue(_is_retryable(_req.exceptions.ConnectionError()))

    def test_is_retryable_timeout(self):
        """_is_retryable: True para requests.exceptions.Timeout."""
        import requests as _req

        from src.extractors.http_utils import _is_retryable

        self.assertTrue(_is_retryable(_req.exceptions.Timeout()))

    def test_is_retryable_5xx(self):
        """_is_retryable: True para HTTPError con status >= 500."""
        import requests as _req

        from src.extractors.http_utils import _is_retryable

        resp = MagicMock()
        resp.status_code = 503
        exc = _req.exceptions.HTTPError(response=resp)
        self.assertTrue(_is_retryable(exc))

    def test_is_not_retryable_4xx(self):
        """_is_retryable: False para HTTPError con status < 500."""
        import requests as _req

        from src.extractors.http_utils import _is_retryable

        resp = MagicMock()
        resp.status_code = 404
        exc = _req.exceptions.HTTPError(response=resp)
        self.assertFalse(_is_retryable(exc))

    def test_is_not_retryable_generic_exception(self):
        """_is_retryable: False para excepciones no HTTP."""
        from src.extractors.http_utils import _is_retryable

        self.assertFalse(_is_retryable(ValueError("algo inesperado")))


class PobrezaComunalExtractorTests(unittest.TestCase):
    """Tests unitarios para el extractor de pobreza comunal (CASEN / SAE)."""

    def test_normalize_rows_writes_required_schema(self):
        """normalize_rows produce las columnas requeridas con tipos correctos."""
        from src.extractors.pobreza_extractor import FALLBACK_ROWS, normalize_rows

        df = normalize_rows(FALLBACK_ROWS)
        required = {
            "codigo_region",
            "codigo_comuna",
            "nombre_comuna",
            "anio",
            "dimension",
            "tasa",
            "limite_inferior",
            "limite_superior",
            "metodologia",
            "fuente",
        }
        self.assertTrue(required.issubset(set(df.columns)))
        self.assertEqual(df["codigo_comuna"].dtype, pl.String)
        self.assertEqual(df["codigo_region"].dtype, pl.String)
        self.assertGreater(df.height, 0)

    def test_normalize_rows_dimensions(self):
        """Las filas de fallback tienen ambas dimensiones (ingresos y multidimensional)."""
        from src.extractors.pobreza_extractor import FALLBACK_ROWS, normalize_rows

        df = normalize_rows(FALLBACK_ROWS)
        dims = set(df["dimension"].unique().to_list())
        self.assertTrue({"ingresos", "multidimensional"}.issubset(dims))

    def test_dataset_name(self):
        """El extractor reporta el nombre canónico."""
        from src.extractors.pobreza_extractor import PobrezaComunalExtractor

        extractor = PobrezaComunalExtractor()
        self.assertEqual(extractor.dataset_name, "pobreza_comunal")

    def test_run_dry_run_returns_validation_without_writing(self):
        """Dry run ejecuta fetch + normalize + validate sin persistir."""
        import tempfile
        from unittest.mock import patch

        from src.extractors import pobreza_extractor

        with tempfile.TemporaryDirectory() as tmpdir:
            staging_csv = Path(tmpdir) / "pobreza_comunal.csv"
            metadata_path = Path(tmpdir) / "pobreza_comunal.metadata.json"
            with (
                patch.object(pobreza_extractor, "STAGING_CSV_PATH", str(staging_csv)),
                patch.object(pobreza_extractor, "METADATA_PATH", str(metadata_path)),
                patch.object(pobreza_extractor, "RAW_DIR", tmpdir),
                patch.object(pobreza_extractor, "STAGING_DIR", tmpdir),
                patch.object(
                    pobreza_extractor,
                    "fetch_data",
                    return_value=(
                        pobreza_extractor.FALLBACK_ROWS,
                        "fallback",
                        pobreza_extractor.POBREZA_INGRESOS_URL,
                        ["test"],
                    ),
                ),
            ):
                result = pobreza_extractor.PobrezaComunalExtractor().run(dry_run=True)
                self.assertEqual(result["status"], "ok")
                self.assertFalse(staging_csv.exists())

    def test_normalize_empty_rows(self):
        """normalize_rows con lista vacía retorna DataFrame con columnas esperadas."""
        from src.extractors.pobreza_extractor import normalize_rows

        df = normalize_rows([])
        required = {"codigo_region", "codigo_comuna", "nombre_comuna", "anio", "dimension"}
        self.assertTrue(required.issubset(set(df.columns)))
        self.assertEqual(df.height, 0)


class ConsumoElectricoExtractorTests(unittest.TestCase):
    """Tests unitarios para el extractor de consumo eléctrico comunal (CNE)."""

    def test_normalize_rows_writes_required_schema(self):
        from src.extractors.consumo_electrico_extractor import FALLBACK_ROWS, normalize_rows

        df = normalize_rows(FALLBACK_ROWS)
        required = {
            "codigo_region",
            "codigo_comuna",
            "nombre_comuna",
            "anio",
            "tipo_cliente",
            "consumo_kwh",
        }
        self.assertTrue(required.issubset(set(df.columns)))
        self.assertEqual(df["codigo_comuna"].dtype, pl.String)
        self.assertGreater(df.height, 0)

    def test_dataset_name(self):
        from src.extractors.consumo_electrico_extractor import ConsumoElectricoExtractor

        extractor = ConsumoElectricoExtractor()
        self.assertEqual(extractor.dataset_name, "consumo_electrico_comunal")

    def test_run_dry_run_returns_validation_without_writing(self):
        import tempfile
        from unittest.mock import patch

        from src.extractors import consumo_electrico_extractor

        with tempfile.TemporaryDirectory() as tmpdir:
            staging_csv = Path(tmpdir) / "consumo_electrico_comunal.csv"
            metadata_path = Path(tmpdir) / "consumo_electrico_comunal.metadata.json"
            with (
                patch.object(consumo_electrico_extractor, "STAGING_CSV_PATH", str(staging_csv)),
                patch.object(consumo_electrico_extractor, "METADATA_PATH", str(metadata_path)),
                patch.object(consumo_electrico_extractor, "RAW_DIR", tmpdir),
                patch.object(consumo_electrico_extractor, "STAGING_DIR", tmpdir),
                patch.object(
                    consumo_electrico_extractor,
                    "fetch_data",
                    return_value=(
                        consumo_electrico_extractor.FALLBACK_ROWS,
                        "fallback",
                        consumo_electrico_extractor.DOWNLOAD_URL,
                        ["test"],
                    ),
                ),
            ):
                result = consumo_electrico_extractor.ConsumoElectricoExtractor().run(dry_run=True)
                self.assertEqual(result["status"], "ok")
                self.assertFalse(staging_csv.exists())

    def test_normalize_empty_rows(self):
        from src.extractors.consumo_electrico_extractor import normalize_rows

        df = normalize_rows([])
        required = {"codigo_region", "codigo_comuna", "tipo_cliente", "consumo_kwh"}
        self.assertTrue(required.issubset(set(df.columns)))
        self.assertEqual(df.height, 0)


class PartidosPoliticosExtractorTests(unittest.TestCase):
    """Tests del extractor de partidos políticos (Plan 023 · Ola B)."""

    XML = (
        '<?xml version="1.0" encoding="utf-8"?>'
        "<PartidosPoliticosColeccion "
        'xmlns="http://opendata.camara.cl/camaradiputados/v1">'
        "<PartidoPolitico><Id>FA</Id><Nombre>Frente Amplio</Nombre>"
        "<Alias>FA</Alias></PartidoPolitico>"
        "<PartidoPolitico><Id>DC</Id><Nombre>Partido Demócrata Cristiano</Nombre>"
        "<Alias>DC</Alias></PartidoPolitico>"
        "<PartidoPolitico><Id>FA</Id><Nombre>Frente Amplio</Nombre>"
        "<Alias>FA</Alias></PartidoPolitico>"  # duplicado: debe deduplicarse
        "<PartidoPolitico><Id></Id><Nombre>Sin id</Nombre></PartidoPolitico>"  # se descarta
        "</PartidosPoliticosColeccion>"
    ).encode("utf-8")

    def test_parse_partidos_maps_schema_and_dedupes(self):
        df = partidos_politicos_extractor.parse_partidos(self.XML)
        self.assertEqual(df.height, 2)  # deduplicado + descarta el sin id
        self.assertEqual(
            df.columns,
            [
                "id_partido",
                "nombre",
                "sigla",
                "estado_legal",
                "fecha_constitucion",
                "ambito",
                "fuente",
                "url_fuente",
                "fecha_consulta",
            ],
        )
        self.assertEqual(sorted(df["id_partido"].to_list()), ["DC", "FA"])
        # Campos no provistos por la fuente quedan nulos en v1.
        self.assertTrue(df["estado_legal"].is_null().all())

    def test_no_personal_data_columns(self):
        df = partidos_politicos_extractor.parse_partidos(self.XML)
        personales = {"rut", "run", "domicilio", "rutdv"}
        self.assertFalse(personales & {c.lower() for c in df.columns})

    def test_validate_flags_low_count(self):
        df = partidos_politicos_extractor.parse_partidos(self.XML)
        result = partidos_politicos_extractor.PartidosPoliticosExtractor().validate(df, {})
        self.assertEqual(result["status"], "error")  # 2 < MIN_EXPECTED_PARTIES
        self.assertEqual(result["record_count"], 2)

    def test_parse_partidos_con_servel_lookup(self):
        """El join con SERVEL completa estado_legal/fecha_constitucion; ambito queda nulo."""
        servel_lookup = {"partido democrata cristiano": ("constituido", "1988-05-02")}
        df = partidos_politicos_extractor.parse_partidos(self.XML, servel_lookup)
        dc = df.filter(pl.col("id_partido") == "DC").row(0, named=True)
        self.assertEqual(dc["estado_legal"], "constituido")
        self.assertEqual(dc["fecha_constitucion"], "1988-05-02")
        self.assertIsNone(dc["ambito"])
        fa = df.filter(pl.col("id_partido") == "FA").row(0, named=True)
        self.assertIsNone(fa["estado_legal"])  # sin match en SERVEL, no se inventa

    def test_parse_fecha_espanol(self):
        f = partidos_politicos_extractor._parse_fecha_espanol
        self.assertEqual(f("2 de mayo de 1988"), "1988-05-02")
        self.assertEqual(f("13 de abril de 2017"), "2017-04-13")
        self.assertIsNone(f("fecha desconocida"))

    def test_norm_partido_quita_sufijo_de_chile(self):
        f = partidos_politicos_extractor._norm_partido
        self.assertEqual(f("Partido Comunista"), f("Partido Comunista de Chile"))

    def test_parse_servel_tabla_constituidos(self):
        html_fragment = (
            "<table>"
            "<tr><td>Partidos Políticos</td><td>Fecha de constitución</td><td>Ver</td></tr>"
            "<tr><td>Fecha Constitución Partidos Políticos</td><td>12 de junio de 2026</td>"
            "<td></td></tr>"
            "<tr><td>Partido Demócrata Cristiano</td><td>2 de mayo de 1988</td><td></td></tr>"
            "</table>"
        )
        filas = partidos_politicos_extractor._parse_servel_tabla(html_fragment, con_fecha=True)
        self.assertEqual(filas, [("Partido Demócrata Cristiano", "2 de mayo de 1988")])

    def test_fetch_servel_estado_legal_degrada_si_falla_la_red(self):
        with patch.object(
            partidos_politicos_extractor,
            "fetch_with_retry",
            side_effect=OSError("sin red"),
        ):
            lookup = partidos_politicos_extractor.fetch_servel_estado_legal()
        self.assertEqual(lookup, {})

    def test_fetch_servel_estado_legal_combina_ambas_paginas(self):
        constituidos_html = (
            "<table><tr><td>Partido Demócrata Cristiano</td>"
            "<td>2 de mayo de 1988</td><td></td></tr></table>"
        )
        en_formacion_html = (
            "<table><tr><td>Partido Cristiano de Chile</td><td>Ver</td></tr></table>"
        )

        def _fake_fetch(url, **kwargs):
            resp = MagicMock()
            resp.text = (
                constituidos_html
                if url == partidos_politicos_extractor.SERVEL_CONSTITUIDOS_URL
                else en_formacion_html
            )
            return resp

        with patch.object(
            partidos_politicos_extractor, "fetch_with_retry", side_effect=_fake_fetch
        ):
            lookup = partidos_politicos_extractor.fetch_servel_estado_legal()
        self.assertEqual(lookup["partido democrata cristiano"], ("constituido", "1988-05-02"))
        self.assertEqual(lookup["partido cristiano"], ("en_formacion", None))


class AutoridadesElectasExtractorTests(unittest.TestCase):
    """Tests del extractor de autoridades electas (Plan 023 · Ola A, cargo diputados)."""

    XML = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<DiputadosPeriodoColeccion xmlns="http://opendata.camara.cl/camaradiputados/v1">'
        "<DiputadoPeriodo><Diputado><Id>1009</Id><Nombre>Jorge</Nombre>"
        "<ApellidoPaterno>Alessandri</ApellidoPaterno><ApellidoMaterno>Vergara</ApellidoMaterno>"
        "<RUT>1</RUT><Sexo>Masculino</Sexo>"
        "<Militancias><Militancia><FechaInicio>2018-03-11T00:00:00</FechaInicio>"
        "<FechaTermino>2022-03-10T23:59:59</FechaTermino>"
        "<Partido><Id>UDI</Id><Nombre>Unión Demócrata Independiente</Nombre></Partido></Militancia>"
        "<Militancia><FechaInicio>2022-03-11T00:00:00</FechaInicio><FechaTermino></FechaTermino>"
        "<Partido><Id>UDI</Id><Nombre>Unión Demócrata Independiente</Nombre></Partido></Militancia>"
        "</Militancias></Diputado></DiputadoPeriodo>"
        "<DiputadoPeriodo><Diputado><Id>1015</Id><Nombre>Jorge</Nombre>"
        "<ApellidoPaterno>Brito</ApellidoPaterno><ApellidoMaterno>Hasbún</ApellidoMaterno>"
        "<Militancias><Militancia><FechaInicio>2022-03-11T00:00:00</FechaInicio>"
        "<FechaTermino></FechaTermino>"
        "<Partido><Id>FA</Id><Nombre>Frente Amplio</Nombre></Partido></Militancia></Militancias>"
        "</Diputado></DiputadoPeriodo>"
        "</DiputadosPeriodoColeccion>"
    ).encode("utf-8")

    DISTRITOS = {"1009": "10", "1015": "7"}

    def test_build_maps_schema_partido_y_distrito(self):
        df = autoridades_electas_extractor.build_autoridades_df(self.XML, self.DISTRITOS)
        self.assertEqual(df.height, 2)
        self.assertEqual(df["cargo"].unique().to_list(), ["diputado"])
        row = df.filter(pl.col("id_autoridad") == "diputado_1009").row(0, named=True)
        self.assertEqual(row["nombre"], "Jorge Alessandri Vergara")
        self.assertEqual(row["partido"], "Unión Demócrata Independiente")  # militancia vigente
        self.assertEqual(row["distrito_electoral"], "10")  # unido por id
        self.assertEqual(row["estado_mandato"], "vigente")

    def test_distrito_nulo_si_no_hay_mapa(self):
        df = autoridades_electas_extractor.build_autoridades_df(self.XML, {})
        self.assertTrue(df["distrito_electoral"].is_null().all())

    def test_sin_columnas_personales(self):
        df = autoridades_electas_extractor.build_autoridades_df(self.XML, self.DISTRITOS)
        personales = {"rut", "rutdv", "run", "fecha_nacimiento", "domicilio", "sexo"}
        self.assertFalse(personales & {c.lower() for c in df.columns})

    def test_incluye_senadores(self):
        senadores = [
            {
                "ID_PARLAMENTARIO": 1110,
                "NOMBRE_COMPLETO": "Pedro Araya Guerrero",
                "PARTIDO": "P.P.D.",
                "CIRCUNSCRIPCION_ID": 3,
                "REGION": "Región de Antofagasta",
                "EMAIL": "x@senado.cl",
                "PERIODOS": [
                    {"DESDE": "2022", "HASTA": "2026", "VIGENTE": 0},
                    {"DESDE": "2026", "HASTA": "2030", "VIGENTE": 1},
                ],
            }
        ]
        df = autoridades_electas_extractor.build_autoridades_df(self.XML, self.DISTRITOS, senadores)
        self.assertEqual(df.height, 3)  # 2 diputados + 1 senador
        sen = df.filter(pl.col("cargo") == "senador").row(0, named=True)
        self.assertEqual(sen["id_autoridad"], "senador_1110")
        self.assertEqual(sen["nombre"], "Pedro Araya Guerrero")
        self.assertEqual(sen["circunscripcion_senatorial"], "3")
        self.assertEqual(sen["institucion"], "Senado")
        self.assertEqual(sen["codigo_region"], "02")
        self.assertEqual(sen["periodo_inicio"], "2026-03-11")
        self.assertEqual(sen["periodo_fin"], "2030-03-10")
        # el email no debe filtrarse como columna
        self.assertNotIn("email", {c.lower() for c in df.columns})

    def test_senador_sin_periodo_vigente_queda_nulo(self):
        senadores = [
            {
                "ID_PARLAMENTARIO": 1111,
                "NOMBRE_COMPLETO": "Alguien Sin Vigente",
                "REGION": "Región de Magallanes y la Antártica Chilena",
                "PERIODOS": [{"DESDE": "2018", "HASTA": "2022", "VIGENTE": 0}],
            }
        ]
        df = autoridades_electas_extractor.build_autoridades_df(self.XML, self.DISTRITOS, senadores)
        sen = df.filter(pl.col("id_autoridad") == "senador_1111").row(0, named=True)
        self.assertIsNone(sen["periodo_inicio"])
        self.assertIsNone(sen["periodo_fin"])
        self.assertEqual(sen["codigo_region"], "12")  # variante real de senado.cl


class AutoridadesLocalesExtractorTests(unittest.TestCase):
    """Tests del extractor de autoridades locales (Plan 023 · Ola A, gobernadores)."""

    def test_region_from_title_variantes(self):
        f = autoridades_locales_extractor._region_from_title
        self.assertEqual(f("Gobernador regional de Arica y Parinacota"), "Arica y Parinacota")
        self.assertEqual(f("Gobernadora regional de Atacama"), "Atacama")  # femenino
        self.assertEqual(f("Gobernador regional del Maule"), "Maule")  # "del"
        self.assertEqual(f("Gobernador regional Metropolitano de Santiago"), "Santiago")
        self.assertIsNone(f("Renovación Nacional"))  # no es región

    def test_normalize_mapea_codigo_region(self):
        gobs = [
            {"region": "Maule", "nombre": "Pedro X", "partido": "UDI", "pacto": "Chile Vamos"},
            {"region": "Santiago", "nombre": "Claudio Y", "partido": "Ind.", "pacto": ""},
        ]
        df = autoridades_locales_extractor.build_autoridades_locales_df(gobs)
        self.assertEqual(df.height, 2)
        maule = df.filter(pl.col("nombre") == "Pedro X").row(0, named=True)
        self.assertEqual(maule["codigo_region"], "07")
        self.assertEqual(maule["cargo"], "gobernador_regional")
        self.assertEqual(maule["id_autoridad"], "gobernador_07")
        stgo = df.filter(pl.col("nombre") == "Claudio Y").row(0, named=True)
        self.assertEqual(stgo["codigo_region"], "13")

    def test_licencia_cc_by_sa(self):
        self.assertEqual(autoridades_locales_extractor.REUSE_POLICY["license"], "CC-BY-SA")

    def test_comuna_name_from_title_quita_prefijo_y_desambiguador(self):
        f = autoridades_locales_extractor._comuna_name_from_title
        self.assertEqual(f("Anexo:Alcaldes de Concepción (Chile)"), "Concepción")
        self.assertEqual(f("Anexo:Alcaldes de Alhué"), "Alhué")

    def test_extract_alcalde_actual_via_titular(self):
        wikitext = (
            "{{Ficha de cargo\n"
            "| titular = [[Orlando Vargas Pizarro]]\n"
            "| inicio = {{fecha|6|12|2024}}\n"
            "| imagen =\n"
            "}}"
        )
        nombre, inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertEqual(nombre, "Orlando Vargas Pizarro")
        self.assertEqual(inicio, "2024-12-06")

    def test_extract_alcalde_actual_titular_una_linea_no_se_come_otros_campos(self):
        # Bug real: infobox con varios campos en una sola línea separados por "|".
        wikitext = "|titular=Patricio Ferreira Rivera|inicio=6 de diciembre de 2016|dirige=[[Alto Hospicio]]|sede=[[Municipalidad]]"
        nombre, inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertEqual(nombre, "Patricio Ferreira Rivera")

    def test_extract_alcalde_actual_titular_vacio_cae_a_fallback(self):
        # Bug real: "titular = " vacío no debe tragarse el campo siguiente.
        wikitext = (
            "|titular         = \n"
            "|inicio          = {{Fecha|6|12|2012}}\n"
            "|en_ejercicio    = Centauri Dios\n"
        )
        nombre, _inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertNotIn("inicio", (nombre or ""))
        self.assertNotIn("{{Fecha", (nombre or ""))

    def test_extract_alcalde_actual_fallback_tabla_respeta_wikilink_con_pipe(self):
        # Bug real: dividir la fila por "|" a secas rompe [[X|Y]] a la mitad.
        wikitext = (
            "{|\n"
            "|-\n"
            "|[[Sacha Razmilic|Sacha Razmilic Burgos]]\n"
            "|6 de diciembre de 2024\n"
            "|''En el cargo''\n"
            "|[[Evolución Política|Evópoli]]\n"
            "|}"
        )
        nombre, _inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertEqual(nombre, "Sacha Razmilic Burgos")

    def test_extract_alcalde_actual_fallback_descarta_ordinal_e_imagen(self):
        # Bug real: la celda ordinal ("31°") y la de imagen no deben tomarse como nombre.
        wikitext = (
            "{|\n"
            "|-\n"
            "|'''31°'''\n"
            "|[[Archivo:Falta_imagen_hombre.svg|100x100px]]\n"
            "|Rodrigo Cornejo Inostroza\n"
            "|2024-en el cargo\n"
            "|}"
        )
        nombre, _inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertEqual(nombre, "Rodrigo Cornejo Inostroza")

    def test_extract_alcalde_actual_sin_evidencia_retorna_none(self):
        # Sin titular ni marca de vigencia explícita: no se inventa un nombre.
        wikitext = "{|\n|-\n|Juan Pérez\n|1990\n|1994\n|}"
        nombre, inicio = autoridades_locales_extractor._extract_alcalde_actual(wikitext)
        self.assertIsNone(nombre)
        self.assertIsNone(inicio)

    def test_normalize_alcaldes_matchea_codigo_comuna(self):
        alcaldes = [{"comuna": "Arica", "nombre": "X Y", "periodo_inicio": None}]
        lookup = {"arica": ("01101", "15")}
        df = autoridades_locales_extractor.build_autoridades_locales_df([], alcaldes, lookup)
        row = df.filter(pl.col("cargo") == "alcalde").row(0, named=True)
        self.assertEqual(row["codigo_comuna"], "01101")
        self.assertEqual(row["codigo_region"], "15")
        self.assertEqual(row["estado_mandato"], "vigente")

    def test_normalize_alcaldes_sin_nombre_marca_sin_identificar(self):
        alcaldes = [{"comuna": "Antofagasta", "nombre": None, "periodo_inicio": None}]
        df = autoridades_locales_extractor.build_autoridades_locales_df([], alcaldes, {})
        row = df.filter(pl.col("cargo") == "alcalde").row(0, named=True)
        self.assertIsNone(row["nombre"])
        self.assertEqual(row["estado_mandato"], "sin_identificar")

    def test_sin_columnas_personales_alcaldes(self):
        alcaldes = [{"comuna": "Arica", "nombre": "X Y", "periodo_inicio": None}]
        df = autoridades_locales_extractor.build_autoridades_locales_df([], alcaldes, {})
        personales = {"rut", "run", "domicilio", "fecha_nacimiento"}
        self.assertFalse(personales & {c.lower() for c in df.columns})


if __name__ == "__main__":
    import sys

    import pytest

    sys.exit(pytest.main(sys.argv))
