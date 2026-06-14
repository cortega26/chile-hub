"""Extrae un perfil demografico comunal desde resultados oficiales Censo 2024."""

import datetime
import os
import sys
from pathlib import Path

import openpyxl
import polars as pl
import requests

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

try:
    from src.extractors.base import (
        BaseExtractor,
        ensure_staging_directories,
        write_staging_metadata,
    )
except ModuleNotFoundError:
    from base import BaseExtractor, ensure_staging_directories, write_staging_metadata

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data"))
RAW_DIR = os.path.join(DATA_DIR, "raw")
STAGING_DIR = os.path.join(DATA_DIR, "staging")
STAGING_CSV_PATH = os.path.join(STAGING_DIR, "censo_comunal.csv")
METADATA_PATH = os.path.join(STAGING_DIR, "censo_comunal.metadata.json")
CENSO_URL = (
    "https://censo2024.ine.gob.cl/wp-content/uploads/2025/03/"
    "D1_Poblacion-censada-por-sexo-y-edad-en-grupos-quinquenales.xlsx"
)
REUSE_POLICY = {
    "status": "open-attribution",
    "license": "CC BY 4.0",
    "license_url": "https://www.ine.gob.cl/terminos-de-uso",
    "attribution_required": True,
    "redistribution_ok": True,
    "summary": "Resultados oficiales del Censo 2024 publicados por el INE; atribucion requerida.",
}

AGE_BANDS = {
    "poblacion_0_14": {"0 a 4", "5 a 9", "10 a 14"},
    "poblacion_15_29": {"15 a 19", "20 a 24", "25 a 29"},
    "poblacion_30_44": {"30 a 34", "35 a 39", "40 a 44"},
    "poblacion_45_64": {"45 a 49", "50 a 54", "55 a 59", "60 a 64"},
    "poblacion_65_mas": {
        "65 a 69",
        "70 a 74",
        "75 a 79",
        "80 a 84",
        "85 o más",
    },
}


def _snapshot_path() -> Path:
    stamp = datetime.datetime.now(datetime.UTC).strftime("%Y%m%dT%H%M%SZ")
    return Path(RAW_DIR) / f"ine_censo2024_comunal_{stamp}.xlsx"


def fetch_workbook() -> tuple[Path, str]:
    ensure_staging_directories()
    target = _snapshot_path()
    try:
        response = requests.get(CENSO_URL, timeout=60)
        response.raise_for_status()
        target.write_bytes(response.content)
        return target, "live"
    except Exception:
        snapshots = sorted(Path(RAW_DIR).glob("ine_censo2024_comunal_*.xlsx"))
        if not snapshots:
            raise
        return snapshots[-1], "fallback"


def parse_workbook(path: Path) -> pl.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    totals = {}
    for row in workbook["2"].iter_rows(min_row=6, values_only=True):
        if not row[4]:
            continue
        code = str(int(row[4])).zfill(5)
        totals[code] = {
            "codigo_region": str(int(row[0])).zfill(2),
            "nombre_region": row[1],
            "codigo_provincia": str(int(row[2])).zfill(3),
            "nombre_provincia": row[3],
            "codigo_comuna": code,
            "nombre_comuna": row[5],
            "poblacion_censada": int(row[6]),
            "hombres": int(row[7]),
            "mujeres": int(row[8]),
            "razon_hombre_mujer": float(row[9]),
            **{band: 0 for band in AGE_BANDS},
        }

    for row in workbook["4"].iter_rows(min_row=6, values_only=True):
        if not row[4]:
            continue
        code = str(int(row[4])).zfill(5)
        if code not in totals:
            continue
        for band, labels in AGE_BANDS.items():
            if row[6] in labels:
                totals[code][band] += int(row[7])

    return pl.DataFrame(list(totals.values())).sort("codigo_comuna")


def process_censo() -> str:
    path, source_mode = fetch_workbook()
    df = parse_workbook(path)
    validation = CensoExtractor().validate(df, {"source_mode": source_mode})
    if validation["status"] == "error":
        raise SystemExit(f"Validacion fallida: {validation['errors']}")
    metadata = {
        "dataset": "censo_comunal",
        "source_name": "Instituto Nacional de Estadisticas - Censo 2024",
        "source_url": CENSO_URL,
        "source_mode": source_mode,
        "source_detail": "official_xlsx" if source_mode == "live" else "raw_snapshot_recovery",
        "refreshed_at_utc": datetime.datetime.now(datetime.UTC).isoformat(),
        "record_count": df.height,
        "fields": df.columns,
        "notes": ["age_bands_derived_from_quinquennial_groups"],
        "reuse_policy": REUSE_POLICY,
    }
    CensoExtractor().write_staging(df, metadata)
    print(f"Censo comunal guardado en: {STAGING_CSV_PATH} ({df.height} registros, {source_mode})")
    return STAGING_CSV_PATH


class CensoExtractor(BaseExtractor):
    @property
    def dataset_name(self) -> str:
        return "censo_comunal"

    def fetch(self, **kwargs):
        return fetch_workbook()

    def normalize(self, raw_data):
        return parse_workbook(raw_data[0])

    def validate(self, df, metadata: dict) -> dict:
        from src.validation import validate_censo_comunal

        return validate_censo_comunal(df, metadata)

    def write_staging(self, df, metadata: dict) -> Path:
        ensure_staging_directories()
        output = Path(STAGING_CSV_PATH)
        df.write_csv(output)
        write_staging_metadata(METADATA_PATH, metadata)
        return output


if __name__ == "__main__":
    process_censo()
