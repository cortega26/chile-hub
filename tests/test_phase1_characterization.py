"""Phase 1 characterization harness tests.

Assertions are classified in the accompanying matrix, not inferred from this
module's implementation.  The helpers are intentionally strict so later
shadow migrations cannot normalize away unexplained artifact changes.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import zipfile
from contextlib import ExitStack
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

import duckdb
import pytest
from openpyxl import load_workbook

import chile_hub.pipeline_status_utils as runtime_pipeline_status_utils
from chile_hub import ChileHub
from src import build_dev_db
from src.builders import artifacts, catalog, data_package, dcat_catalog, formats, metadata, reports
from src.chile_hub import pipeline_status_utils
from tests.pipeline_characterization import (
    artifact_tree_fingerprints,
    assert_artifact_trees_equivalent,
    assert_semantic_zip_equivalent,
    write_legacy_build_staging_fixture,
)

ROOT_DIR = Path(__file__).resolve().parents[1]
REQUIRED_METADATA_PATHS = (
    "COMUNAS_METADATA_PATH",
    "INDICADORES_METADATA_PATH",
    "CENSO_METADATA_PATH",
    "SALUD_METADATA_PATH",
    "CENSO_HOGARES_METADATA_PATH",
    "ELECTORAL_METADATA_PATH",
    "FINANZAS_METADATA_PATH",
    "RESULTADOS_EDUCACIONALES_METADATA_PATH",
    "SIEDU_METADATA_PATH",
    "EMPRESAS_METADATA_PATH",
    "POBREZA_COMUNAL_METADATA_PATH",
    "CONSUMO_ELECTRICO_COMUNAL_METADATA_PATH",
    "PARTIDOS_POLITICOS_METADATA_PATH",
    "AUTORIDADES_ELECTAS_METADATA_PATH",
)
FROZEN_BUILD_TIME = datetime(2026, 8, 23, 12, 0, tzinfo=timezone.utc)
_SERIALIZATION_VARIANT_PATHS = {
    # Frozen input data still reaches these paths through implementation-defined
    # JSON dict iteration and ZIP member timestamps respectively.
    "data/normalized/pipeline_metadata.json",
    "data/normalized/chile-hub-publishable-bundle.zip",
}


class _FrozenDateTime(datetime):
    """Narrow test seam for legacy builders that call ``datetime.now``."""

    @classmethod
    def now(cls, tz=None):
        if tz is None:
            return FROZEN_BUILD_TIME.replace(tzinfo=None)
        return FROZEN_BUILD_TIME.astimezone(tz)


def _write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _assert_manifest_checksums_match_files(normalized: Path, manifest: dict) -> None:
    for artifact in manifest["artifacts"]:
        _assert_declared_reference_matches_file(normalized, artifact)
    for package in manifest["packages"]:
        _assert_declared_reference_matches_file(normalized, package)


def _assert_declared_reference_matches_file(normalized: Path, entry: dict) -> None:
    """Verify every declared hash/size before semantic comparison projects it."""
    path = normalized.parent.parent / entry["path"]
    payload = path.read_bytes()
    if "size_bytes" in entry:
        assert len(payload) == entry["size_bytes"]
    if "sha256" in entry:
        assert hashlib.sha256(payload).hexdigest() == entry["sha256"]


def _semantic_json_sha256(path: Path) -> str:
    """Hash JSON semantics when source byte key order is non-contractual."""
    canonical = json.dumps(
        json.loads(path.read_text(encoding="utf-8")),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(canonical).hexdigest()


def _semantic_zip_sha256(path: Path) -> str:
    """Hash uncompressed member semantics, excluding ZIP header timestamps."""
    digest = hashlib.sha256()
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            payload = archive.read(name)
            if Path(name).suffix == ".json":
                payload = json.dumps(
                    json.loads(payload), ensure_ascii=False, sort_keys=True, separators=(",", ":")
                ).encode("utf-8")
            digest.update(name.encode("utf-8"))
            digest.update(b"\0")
            digest.update(payload)
            digest.update(b"\0")
    return digest.hexdigest()


def _project_serialization_variant(normalized: Path, entry: dict) -> dict:
    """Project only two verified, known non-semantic byte representations."""
    _assert_declared_reference_matches_file(normalized, entry)
    projected = json.loads(json.dumps(entry))
    path = projected.get("path")
    if path not in _SERIALIZATION_VARIANT_PATHS:
        return projected
    artifact_path = normalized.parent.parent / path
    if artifact_path.suffix == ".json":
        projected["sha256"] = _semantic_json_sha256(artifact_path)
    elif artifact_path.suffix == ".zip":
        projected["sha256"] = _semantic_zip_sha256(artifact_path)
        with zipfile.ZipFile(artifact_path) as archive:
            projected["size_bytes"] = sum(len(archive.read(name)) for name in archive.namelist())
    return projected


def _semantic_manifest_projection(normalized: Path, manifest: dict) -> dict:
    """Project only unavoidable serialization variance to semantic equivalents.

    Every declared byte hash/size is first checked against its emitted file.
    This projection then changes only the two known byte-variant paths, never
    a general class of JSON/ZIP references.
    """
    projected = json.loads(json.dumps(manifest))
    projected["artifacts"] = [
        _project_serialization_variant(normalized, artifact) for artifact in manifest["artifacts"]
    ]
    projected["packages"] = [
        _project_serialization_variant(normalized, package) for package in manifest["packages"]
    ]
    return projected


def _semantic_inventory_projection(normalized: Path, payload: object) -> object:
    """Project only verified references with known serialization variance."""
    if isinstance(payload, list):
        return [_semantic_inventory_projection(normalized, value) for value in payload]
    if not isinstance(payload, dict):
        return payload
    projected = {
        key: _semantic_inventory_projection(normalized, value) for key, value in payload.items()
    }
    if isinstance(projected.get("path"), str) and (
        "sha256" in projected or "size_bytes" in projected
    ):
        return _project_serialization_variant(normalized, projected)
    return projected


def _workbook_values(path: Path) -> dict[str, list[tuple[object, ...]]]:
    """Compare the public spreadsheet's cells, not ZIP serialization metadata."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return {
            sheet.title: list(sheet.iter_rows(values_only=True)) for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _duckdb_semantics(path: Path) -> dict[str, object]:
    """Compare DuckDB's public schemas/rows, not nondeterministic file pages."""
    connection = duckdb.connect(str(path), read_only=True)
    try:
        tables = [
            row[0]
            for row in connection.execute(
                "SELECT table_name FROM information_schema.tables "
                "WHERE table_schema = 'main' ORDER BY table_name"
            ).fetchall()
        ]
        return {
            table: {
                "schema": connection.execute(f'DESCRIBE "{table}"').fetchall(),
                "rows": sorted(connection.execute(f'SELECT * FROM "{table}"').fetchall()),
            }
            for table in tables
        }
    finally:
        connection.close()


def _assert_zip_sidecar_matches(normalized: Path) -> None:
    bundle = normalized / "chile-hub-publishable-bundle.zip"
    sidecar = normalized / "chile-hub-publishable-bundle.zip.sha256"
    declared = sidecar.read_text(encoding="utf-8").split()[0]
    assert declared == hashlib.sha256(bundle.read_bytes()).hexdigest()


def _offline_synthetic_staging(
    destination: Path, *, include_candidate: bool = True
) -> dict[str, str]:
    """Create a test-owned staging tree from static literals only."""
    write_legacy_build_staging_fixture(destination, include_candidate=include_candidate)
    overrides = {"STAGING_DIR": str(destination)}
    for attribute in REQUIRED_METADATA_PATHS:
        overrides[attribute] = str(destination / Path(getattr(build_dev_db, attribute)).name)
    return overrides


def _write_pipeline_markdown_to(normalized: Path):
    """Return test-only wrappers for functions with import-time path defaults."""
    return {
        "write_status_markdown_file": lambda payload, health=None: (
            pipeline_status_utils.write_status_markdown_file(
                payload, normalized / "pipeline_status.md", health=health
            )
        ),
        "write_hub_health_markdown_file": lambda payload: (
            pipeline_status_utils.write_hub_health_markdown_file(
                payload, normalized / "hub_health.md"
            )
        ),
        "write_dataset_catalog_markdown_file": lambda payload: (
            pipeline_status_utils.write_dataset_catalog_markdown_file(
                payload, normalized / "dataset_catalog.md"
            )
        ),
        "write_redistribution_report_markdown_file": lambda payload: (
            pipeline_status_utils.write_redistribution_report_markdown_file(
                payload, normalized / "redistribution_report.md"
            )
        ),
        "write_provenance_report_markdown_file": lambda payload: (
            pipeline_status_utils.write_provenance_report_markdown_file(
                payload, normalized / "provenance_report.md"
            )
        ),
        "write_drift_report_markdown_file": lambda payload: (
            pipeline_status_utils.write_drift_report_markdown_file(
                payload, normalized / "drift_report.md"
            )
        ),
        "write_source_readiness_markdown_file": lambda payload: (
            pipeline_status_utils.write_source_readiness_markdown_file(
                payload, normalized / "source_readiness.md"
            )
        ),
        "write_dataset_quality_markdown_file": lambda payload: (
            pipeline_status_utils.write_dataset_quality_markdown_file(
                payload, normalized / "dataset_quality.md"
            )
        ),
        "write_overview_markdown_file": lambda payload: (
            pipeline_status_utils.write_overview_markdown_file(payload, normalized / "overview.md")
        ),
    }


def _isolated_main_paths(tmp_path: Path, *, include_candidate: bool = True) -> ExitStack:
    """Patch only file locations so legacy ``main`` can run in a temp tree."""
    sandbox = tmp_path / "sandbox"
    data = sandbox / "data"
    staging = data / "staging"
    normalized = data / "normalized"
    staging.mkdir(parents=True)
    normalized.mkdir()
    overrides = _offline_synthetic_staging(staging, include_candidate=include_candidate)
    shutil.copy2(ROOT_DIR / "data" / "dataset_catalog_config.json", data)
    shutil.copy2(ROOT_DIR / "data" / "source_registry.json", data)
    shutil.copy2(ROOT_DIR / "pyproject.toml", sandbox)
    shutil.copytree(ROOT_DIR / "contracts", sandbox / "contracts")

    stack = ExitStack()
    stack.enter_context(patch.multiple(build_dev_db, NORMALIZED_DIR=str(normalized), **overrides))
    for module in (artifacts, reports):
        stack.enter_context(
            patch.multiple(module, DATA_DIR=str(data), NORMALIZED_DIR=str(normalized))
        )
    for module in (catalog, formats, data_package, dcat_catalog):
        stack.enter_context(patch.object(module, "NORMALIZED_DIR", str(normalized)))
    for module in (catalog, metadata, reports):
        stack.enter_context(patch.object(module, "ROOT_DIR", str(sandbox)))
    stack.enter_context(
        patch.object(pipeline_status_utils, "SOURCE_REGISTRY_PATH", data / "source_registry.json")
    )
    stack.enter_context(patch.object(catalog, "sync_landing_metadata"))
    for module in (
        artifacts,
        catalog,
        data_package,
        metadata,
        reports,
        pipeline_status_utils,
        runtime_pipeline_status_utils,
    ):
        stack.enter_context(patch.object(module, "datetime", _FrozenDateTime))
    stack.enter_context(patch.multiple(build_dev_db, **_write_pipeline_markdown_to(normalized)))
    stack.enter_context(patch.object(build_dev_db, "sync_readme_layers_table"))
    stack.enter_context(patch.object(build_dev_db, "sync_all_docs"))
    stack.sandbox = sandbox  # type: ignore[attr-defined]
    return stack


def test_offline_staging_loads_direct_and_derived_datasets_without_extractors(
    tmp_path: Path,
) -> None:
    """Compatibility guarantee: the current build boundary accepts offline staging.

    No extractor/network entry point is invoked. This deliberately covers the
    normal direct path, its derived geography/profile outputs, and a present
    optional candidate input in one vertical slice.
    """
    staging = tmp_path / "staging"
    staging.mkdir()
    overrides = _offline_synthetic_staging(staging)

    with patch.multiple(build_dev_db, **overrides):
        datasets, metadata, _ = build_dev_db._load_inputs()
        validations = build_dev_db._compute_validations(datasets, metadata)

    assert datasets["comunas"].height > 0
    assert datasets["regiones"].height > 0
    assert datasets["provincias"].height > 0
    assert datasets["perfil_territorial"].height > 0
    assert datasets["empresas"] is None
    assert datasets["consumo_electrico"] is not None
    assert validations["comunas"]["status"] == "ok"
    assert validations["perfil_territorial_comunal"]["status"] == "ok"


def test_offline_staging_invalid_data_fails_closed_before_artifact_writing(tmp_path: Path) -> None:
    """Compatibility guarantee: invalid direct data terminates before publication."""
    staging = tmp_path / "staging"
    staging.mkdir()
    overrides = _offline_synthetic_staging(staging)

    with patch.multiple(build_dev_db, **overrides):
        datasets, metadata, _ = build_dev_db._load_inputs()
        # One comuna violates established coverage and gives dependent
        # validators an invalid CUT universe.
        datasets["comunas"] = datasets["comunas"].head(1)
        with pytest.raises(SystemExit, match="Validaciones fallidas"):
            build_dev_db._compute_validations(datasets, metadata)


def test_missing_required_csv_and_metadata_fail_closed(tmp_path: Path) -> None:
    """Compatibility guarantee: both required staging boundary failures abort."""
    for omitted_name in ("comunas.csv", "comunas.metadata.json"):
        staging = tmp_path / omitted_name
        staging.mkdir()
        overrides = _offline_synthetic_staging(staging)
        (staging / omitted_name).unlink()

        with patch.multiple(build_dev_db, **overrides), pytest.raises(SystemExit):
            build_dev_db._load_inputs()


def test_optional_candidate_omission_keeps_required_build_path_available(tmp_path: Path) -> None:
    """Compatibility guarantee: omitted optional input does not block a core build."""
    with _isolated_main_paths(tmp_path, include_candidate=False) as stack:
        build_dev_db.main()
        normalized = stack.sandbox / "data" / "normalized"  # type: ignore[attr-defined]

    bundle = json.loads((normalized / "hub_bundle.json").read_text(encoding="utf-8"))
    assert (normalized / "comunas.parquet").is_file()
    assert not (normalized / "consumo_electrico_comunal.parquet").exists()
    assert "consumo_electrico_comunal" not in {
        entry["dataset"] for entry in bundle["candidate_datasets"]
    }


def test_declared_artifact_checksum_rejects_tampered_output(tmp_path: Path) -> None:
    """Compatibility guarantee: injected emitted-artifact corruption is visible."""
    with _isolated_main_paths(tmp_path) as stack:
        build_dev_db.main()
        normalized = stack.sandbox / "data" / "normalized"  # type: ignore[attr-defined]

    manifest = json.loads((normalized / "artifact_manifest.json").read_text(encoding="utf-8"))
    artifact = next(
        entry for entry in manifest["artifacts"] if entry["path"].endswith("comunas.parquet")
    )
    path = normalized.parent.parent / artifact["path"]
    path.write_bytes(path.read_bytes() + b"tampered")

    with pytest.raises(AssertionError):
        _assert_manifest_checksums_match_files(normalized, manifest)


def test_current_main_builds_isolated_offline_artifacts(tmp_path: Path) -> None:
    """Compatibility guarantee: legacy ``main`` remains executable offline.

    This executes load, validation, artifact writing, metadata, report, manifest,
    ZIP, and checksum stages on the bounded fixture. Only output locations and
    documentation writers are isolated; no production builder is substituted.
    """
    with _isolated_main_paths(tmp_path) as stack:
        build_dev_db.main()
        normalized = stack.sandbox / "data" / "normalized"  # type: ignore[attr-defined]

    required = {
        "comunas.parquet",
        "perfil_territorial_comunal.parquet",
        "dataset_catalog.json",
        "pipeline_metadata.json",
        "artifact_manifest.json",
        "hub_bundle.json",
        "chile-hub-publishable-bundle.zip",
        "chile-hub-publishable-bundle.zip.sha256",
    }
    assert required <= {path.name for path in normalized.iterdir()}
    manifest = json.loads((normalized / "artifact_manifest.json").read_text(encoding="utf-8"))
    with zipfile.ZipFile(normalized / "chile-hub-publishable-bundle.zip") as archive:
        members = set(archive.namelist())
        bundle_catalog = json.loads(
            archive.read("data/normalized/dataset_catalog.json").decode("utf-8")
        )

    manifest_paths = {entry["path"] for entry in manifest["artifacts"]}
    assert manifest_paths <= members
    assert "data/normalized/comunas.parquet" in members
    assert "comunas" in {entry["dataset"] for entry in bundle_catalog["datasets"]}


def test_fixed_input_rebuild_is_semantically_equivalent_and_filters_candidates(
    tmp_path: Path,
) -> None:
    """Compatibility guarantee: legacy public artifacts are reproducible.

    The full emitted artifact trees are compared. ZIP containers are compared
    by member payload, not timestamp-bearing ZIP headers; their sidecars and
    manifest declarations are independently verified against emitted bytes.
    """
    outputs = []
    for name in ("first", "second"):
        with _isolated_main_paths(tmp_path / name) as stack:
            build_dev_db.main()
            outputs.append(stack.sandbox / "data" / "normalized")  # type: ignore[attr-defined]

    first, second = outputs
    first_manifest = json.loads((first / "artifact_manifest.json").read_text(encoding="utf-8"))
    second_manifest = json.loads((second / "artifact_manifest.json").read_text(encoding="utf-8"))
    _assert_manifest_checksums_match_files(first, first_manifest)
    _assert_manifest_checksums_match_files(second, second_manifest)
    _assert_zip_sidecar_matches(first)
    _assert_zip_sidecar_matches(second)
    assert _semantic_manifest_projection(first, first_manifest) == _semantic_manifest_projection(
        second, second_manifest
    )
    first_bundle = json.loads((first / "hub_bundle.json").read_text(encoding="utf-8"))
    second_bundle = json.loads((second / "hub_bundle.json").read_text(encoding="utf-8"))
    assert _semantic_inventory_projection(first, first_bundle) == _semantic_inventory_projection(
        second, second_bundle
    )
    assert _workbook_values(first / "chile_data_latest.xlsx") == _workbook_values(
        second / "chile_data_latest.xlsx"
    )
    assert _duckdb_semantics(first / "chile_data.duckdb") == _duckdb_semantics(
        second / "chile_data.duckdb"
    )
    first_overview = _semantic_inventory_projection(
        first, json.loads((first / "overview.json").read_text(encoding="utf-8"))
    )
    second_overview = _semantic_inventory_projection(
        second, json.loads((second / "overview.json").read_text(encoding="utf-8"))
    )
    assert first_overview == second_overview
    for overview in (first / "overview.md", second / "overview.md"):
        markdown = overview.read_text(encoding="utf-8")
        assert "`path`: `data/normalized/chile-hub-publishable-bundle.zip`" in markdown
        assert "`package_type`: `zip`" in markdown
    # Every regular build output must be equal. ZIP bytes and their sidecar
    # legitimately vary with archive member timestamps, so their semantics are
    # verified below rather than hiding those files from the comparison.
    first_regular = tmp_path / "first-regular"
    second_regular = tmp_path / "second-regular"
    for source, destination in ((first, first_regular), (second, second_regular)):
        destination.mkdir()
        for path in source.iterdir():
            if path.name not in {
                "chile-hub-publishable-bundle.zip",
                "chile-hub-publishable-bundle.zip.sha256",
                "artifact_manifest.json",
                "hub_bundle.json",
                "chile_data_latest.xlsx",
                "chile_data.duckdb",
                "overview.json",
                "overview.md",
            }:
                shutil.copy2(path, destination / path.name)
    assert_artifact_trees_equivalent(first_regular, second_regular)
    assert_semantic_zip_equivalent(
        first / "chile-hub-publishable-bundle.zip",
        second / "chile-hub-publishable-bundle.zip",
        volatile_json_paths={},
    )

    for normalized in outputs:
        bundle = json.loads((normalized / "hub_bundle.json").read_text(encoding="utf-8"))
        manifest = json.loads((normalized / "artifact_manifest.json").read_text(encoding="utf-8"))
        artifact_datasets = {
            entry["dataset"] for entry in manifest["artifacts"] if entry["dataset"]
        }
        # The optional candidate is present and valid. It remains visible in
        # public inventory metadata but is not permitted to own a publishable
        # artifact or enter the ZIP.
        assert "consumo_electrico_comunal" not in artifact_datasets
        candidates = {entry["dataset"] for entry in bundle["candidate_datasets"]}
        assert "consumo_electrico_comunal" in candidates
        with zipfile.ZipFile(normalized / "chile-hub-publishable-bundle.zip") as archive:
            assert "data/normalized/consumo_electrico_comunal.parquet" not in archive.namelist()


def test_alias_resolves_to_canonical_physical_artifact(tmp_path: Path) -> None:
    """Compatibility guarantee: alias ownership and resolution remain explicit."""
    with _isolated_main_paths(tmp_path) as stack:
        build_dev_db.main()
        normalized = stack.sandbox / "data" / "normalized"  # type: ignore[attr-defined]

    catalog_payload = json.loads((normalized / "dataset_catalog.json").read_text(encoding="utf-8"))
    catalog_by_dataset = {entry["dataset"]: entry for entry in catalog_payload["datasets"]}
    canonical = catalog_by_dataset["comunas"]
    alias = catalog_by_dataset["comunas_enriquecidas"]
    manifest = json.loads((normalized / "artifact_manifest.json").read_text(encoding="utf-8"))
    owners = {entry["dataset"] for entry in manifest["artifacts"] if entry["dataset"]}

    assert alias["alias_for"] == "comunas"
    assert alias["outputs"]["parquet"] == canonical["outputs"]["parquet"]
    assert "comunas" in owners
    assert "comunas_enriquecidas" not in owners
    hub = ChileHub(data_dir=normalized)
    assert hub.get_output_path("comunas_enriquecidas", "parquet") == hub.get_output_path(
        "comunas", "parquet"
    )
    assert hub.load_polars("comunas_enriquecidas").equals(hub.load_polars("comunas"))


def test_only_explicit_volatile_json_paths_are_ignored(tmp_path: Path) -> None:
    """Implementation detail: generated timestamps do not define equivalence."""
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()
    _write_json(first / "report.json", {"generated_at_utc": "2026-01-01T00:00:00Z", "status": "ok"})
    _write_json(
        second / "report.json", {"generated_at_utc": "2026-01-02T00:00:00Z", "status": "ok"}
    )

    with pytest.raises(AssertionError, match="report.json"):
        assert_artifact_trees_equivalent(first, second)
    assert_artifact_trees_equivalent(first, second, volatile_json_paths={("generated_at_utc",)})


def test_allowlist_does_not_hide_other_public_report_changes(tmp_path: Path) -> None:
    """Compatibility guarantee: volatile metadata cannot mask policy changes."""
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()
    _write_json(first / "report.json", {"generated_at_utc": "a", "status": "ok"})
    _write_json(second / "report.json", {"generated_at_utc": "b", "status": "error"})

    with pytest.raises(AssertionError, match="report.json"):
        assert_artifact_trees_equivalent(first, second, volatile_json_paths={("generated_at_utc",)})


def test_fingerprints_preserve_binary_artifacts_byte_for_byte(tmp_path: Path) -> None:
    """Compatibility guarantee: binary package/checksum differences are visible."""
    root = tmp_path / "artifacts"
    root.mkdir()
    (root / "bundle.zip").write_bytes(b"first")
    first = artifact_tree_fingerprints(root)
    (root / "bundle.zip").write_bytes(b"second")
    second = artifact_tree_fingerprints(root)

    assert first["bundle.zip"] != second["bundle.zip"]
